#!/usr/bin/env python3
"""Generate Surrounding Counties of Smith County TX Churches Contacts spreadsheet (email required)."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
county_font = Font(name="Calibri", bold=True, size=13, color="1B4332")
county_fill = PatternFill(start_color="B7E4C7", end_color="B7E4C7", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=11, color="2D6A4F")
section_fill = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Organization", "County", "Denomination", "Name", "Title", "Phone", "Email", "Address", "Website"]

# ── Data — ONLY entries with a confirmed email ──────────────────────────
# (Organization, County, Denomination, Name, Title, Phone, Email, Address, Website)
# None in Denomination = section header row; "COUNTY:" prefix = county header row

churches = [
    # ═══════════════════════════════════════════════════════════
    # GREGG COUNTY
    # ═══════════════════════════════════════════════════════════
    ("GREGG COUNTY (Longview, Kilgore, Gladewater)", None, None, None, None, None, None, None, None),

    ("First Christian Church - Kilgore", "Gregg", "Christian (Disciples of Christ)",
     "Pastor Jeneille LaGrone", "Senior Minister",
     "(903) 984-3963", "info@fcckilgore.org",
     "609 E Main St, Kilgore, TX 75662",
     "https://fcckilgore.org"),

    ("First Christian Church - Kilgore", "Gregg", "Christian (Disciples of Christ)",
     "Rev. Krystal Morrell", "Minister of Faith Development",
     "(903) 984-3963", "kmorrell@fcckilgore.org",
     "609 E Main St, Kilgore, TX 75662",
     "https://fcckilgore.org"),

    ("First Presbyterian Church - Kilgore", "Gregg", "Presbyterian (PCUSA)",
     "Camelia Griffin", "Administrative Assistant",
     "(903) 984-1502", "firstpreskilgore@gmail.com",
     "815 East Main, Kilgore, TX 75662",
     "https://firstpreskilgore.org"),

    ("First Baptist Church - Kilgore", "Gregg", "Baptist (SBC)",
     "B. Turner", "Staff",
     "(903) 984-3531", "bturner@fbckilgore.org",
     "501 E North St., Kilgore, TX 75662",
     "https://fbckilgore.org"),

    ("Christ the King Catholic Church - Kilgore", "Gregg", "Catholic",
     "", "Church Office",
     "(903) 483-2500", "ctkoffice@ctkkilgore.org",
     "1407 Broadway Blvd, Kilgore, TX",
     "https://ctkkilgore.org"),

    ("New Beginnings Baptist Church - Longview", "Gregg", "Baptist",
     "Pastor Todd Kaunitz", "Senior Pastor",
     "(903) 759-5552", "jarnold@nbbctx.org",
     "2137 E George Richey Rd, Longview, TX 75604",
     "https://www.nbbctx.org"),

    # ═══════════════════════════════════════════════════════════
    # RUSK COUNTY
    # ═══════════════════════════════════════════════════════════
    ("RUSK COUNTY (Henderson)", None, None, None, None, None, None, None, None),

    ("First Baptist Church - Henderson", "Rusk", "Baptist (SBC)",
     "Dr. David Higgs", "Pastor",
     "(903) 657-1646", "website@thefbc.org",
     "207 West Main St., Henderson, TX 75652",
     "https://thefbc.org"),

    ("South Main Church of Christ - Henderson", "Rusk", "Church of Christ",
     "Bob Payne", "Contact",
     "(903) 657-1408", "office@southmainchurch.com",
     "402 S Main Street, Henderson, TX 75652",
     "https://www.southmainchurch.com"),

    # ═══════════════════════════════════════════════════════════
    # UPSHUR COUNTY
    # ═══════════════════════════════════════════════════════════
    ("UPSHUR COUNTY (Gilmer, Big Sandy)", None, None, None, None, None, None, None, None),

    ("First Methodist Church - Gilmer", "Upshur", "Global Methodist",
     "Matthew Smith", "Pastor",
     "(903) 843-2610", "matt@firstgilmer.org",
     "105 North Montgomery Street, Gilmer, TX 75644",
     "https://www.firstgilmer.org"),

    ("First Methodist Church - Gilmer", "Upshur", "Global Methodist",
     "Michael Lucas", "Youth Minister / Worship Leader",
     "(903) 843-2610", "michael@firstgilmer.org",
     "105 North Montgomery Street, Gilmer, TX 75644",
     "https://www.firstgilmer.org"),

    ("First Methodist Church - Gilmer", "Upshur", "Global Methodist",
     "Chandra Sorrells", "Children's Minister",
     "(903) 843-2610", "chandra@firstgilmer.org",
     "105 North Montgomery Street, Gilmer, TX 75644",
     "https://www.firstgilmer.org"),

    # ═══════════════════════════════════════════════════════════
    # WOOD COUNTY
    # ═══════════════════════════════════════════════════════════
    ("WOOD COUNTY (Mineola, Quitman, Winnsboro)", None, None, None, None, None, None, None, None),

    ("Walnut Street Baptist Church - Winnsboro", "Wood", "Baptist",
     "", "Church Office",
     "(903) 342-6193", "wsbcoffice@yahoo.com",
     "919 N Walnut Street, Winnsboro, TX 75494",
     "http://www.walnutstreetbaptistchurch.org"),

    ("First Baptist Church - Mineola", "Wood", "Baptist (SBC)",
     "Mark Neeley", "Contact",
     "(903) 569-3873", "info@fbcmineola.org",
     "204 N Johnson St, Mineola, TX 75773",
     "https://fbcmineola.org"),

    # ═══════════════════════════════════════════════════════════
    # VAN ZANDT COUNTY
    # ═══════════════════════════════════════════════════════════
    ("VAN ZANDT COUNTY (Canton, Grand Saline, Wills Point)", None, None, None, None, None, None, None, None),

    ("Radiant Church VZ - Canton", "Van Zandt", "Foursquare",
     "Adam Henderson", "Senior Pastor",
     "(430) 307-5980", "office@radiantchurchvz.com",
     "19830 Interstate 20, Canton, TX 75103",
     "https://radiantchurchvz.com"),

    # ═══════════════════════════════════════════════════════════
    # HENDERSON COUNTY
    # ═══════════════════════════════════════════════════════════
    ("HENDERSON COUNTY (Athens, Gun Barrel City, Malakoff)", None, None, None, None, None, None, None, None),

    ("Allen Chapel AME Church - Athens", "Henderson", "African Methodist Episcopal",
     "P. Myria Bailey Whitcomb", "Pastor",
     "(903) 675-2425", "Allenchapelpastors@outlook.com",
     "512 N. Underwood Street, Athens, TX 75751",
     ""),

    ("Athens United Pentecostal Church", "Henderson", "United Pentecostal (UPCI)",
     "Rev. Johnny Grissom", "Pastor",
     "(903) 675-1244", "info@athensupc.com",
     "101 McArthur, Athens, TX 75751",
     "https://www.athensupc.com"),

    ("Lake Athens Baptist Church", "Henderson", "Baptist",
     "", "Church Office",
     "(903) 675-4008", "office@lakeathensbaptist.com",
     "5151 FM 2495, Athens, TX 75752",
     "https://lakeathensbaptist.com"),

    # ═══════════════════════════════════════════════════════════
    # CHEROKEE COUNTY
    # ═══════════════════════════════════════════════════════════
    ("CHEROKEE COUNTY (Jacksonville, Rusk)", None, None, None, None, None, None, None, None),

    ("Central Baptist Church - Jacksonville", "Cherokee", "Baptist (SBC)",
     "Dr. Mike Miller", "Senior Pastor",
     "(903) 586-2215", "Mike@cbcjax.net",
     "1909 East Rusk Street, Jacksonville, TX 75766",
     "https://cbcjax.net"),

    ("Central Baptist Church - Jacksonville", "Cherokee", "Baptist (SBC)",
     "Tom Myers", "Education and Media",
     "(903) 586-2215", "Tom@cbcjax.net",
     "1909 East Rusk Street, Jacksonville, TX 75766",
     "https://cbcjax.net"),

    ("Central Baptist Church - Jacksonville", "Cherokee", "Baptist (SBC)",
     "John Armitage", "Music and Worship",
     "(903) 586-2215", "John@cbcjax.net",
     "1909 East Rusk Street, Jacksonville, TX 75766",
     "https://cbcjax.net"),

    ("Central Baptist Church - Jacksonville", "Cherokee", "Baptist (SBC)",
     "Steve Edwards", "Sr. Adults, Missions & Admin",
     "(903) 586-2215", "Steve@cbcjax.net",
     "1909 East Rusk Street, Jacksonville, TX 75766",
     "https://cbcjax.net"),

    ("Central Baptist Church - Jacksonville", "Cherokee", "Baptist (SBC)",
     "Coby Duren", "Students",
     "(903) 586-2215", "Coby@cbcjax.net",
     "1909 East Rusk Street, Jacksonville, TX 75766",
     "https://cbcjax.net"),

    ("First Baptist Church - Jacksonville", "Cherokee", "Baptist (BMA)",
     "Bro. Eric Johnson", "Senior Pastor",
     "(903) 586-4844", "erictjnsn@gmail.com",
     "210 Philip Street, Jacksonville, TX 75766",
     "https://www.firstjax.org"),

    ("First Baptist Church - Jacksonville", "Cherokee", "Baptist (BMA)",
     "Jeanna", "Office Manager",
     "(903) 586-4844", "reception@firstjax.org",
     "210 Philip Street, Jacksonville, TX 75766",
     "https://www.firstjax.org"),
]


# ── Build worksheet ─────────────────────────────────────────────────────
ws = wb.active
ws.title = "Surrounding Counties Churches"

# Column widths
col_widths = [42, 14, 30, 26, 28, 18, 32, 48, 38]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Write headers
for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

# Write data
row_num = 2
for record in churches:
    org = record[0]
    # County header rows (Denomination is None)
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=org)
        cell.font = county_font
        cell.fill = county_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 28
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

# ── Save ────────────────────────────────────────────────────────────────
output_path = "/home/user/Claude/Surrounding_Counties_Churches.xlsx"
wb.save(output_path)
print(f"Saved {row_num - 2} rows to {output_path}")
