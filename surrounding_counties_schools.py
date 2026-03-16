#!/usr/bin/env python3
"""Generate Surrounding Counties School Districts spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=13, color="1B4332")
section_fill = PatternFill(start_color="B7E4C7", end_color="B7E4C7", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["District", "County", "Superintendent", "Phone", "Email", "Address", "Website", "Notes"]

districts = [
    # ── GREGG COUNTY ─────────────────────────────────────────────
    ("GREGG COUNTY", None),

    ("Longview ISD", "Gregg",
     "Dr. Marla Sheppard",
     "(903) 381-2200", "ask@lisd.org",
     "410 N. Center St., Longview, TX 75601 (PO Box 3268, 75606)",
     "https://www.lisd.org",
     "Supt office: tbeard@lisd.org (assistant)"),

    ("Pine Tree ISD", "Gregg",
     "Steve Clugston",
     "(903) 295-5000", "sclugston@ptisd.org",
     "1701 Pine Tree Rd., Longview, TX 75604 (PO Box 5878, 75608)",
     "https://www.ptisd.org",
     ""),

    ("Spring Hill ISD", "Gregg",
     "Penny Fleet",
     "(903) 759-4404", "pfleet@shisd.net",
     "3101 Spring Hill Rd., Longview, TX 75605",
     "https://www.shisd.net",
     "Supt direct: (903) 323-7730"),

    # ── GREGG / RUSK COUNTY ──────────────────────────────────────
    ("GREGG / RUSK COUNTY", None),

    ("Kilgore ISD", "Gregg/Rusk",
     "Dr. Andy Baker",
     "(903) 988-3900", "info@kisd.org",
     "301 N. Kilgore St., Kilgore, TX 75662",
     "https://www.kisd.org",
     "Supt assistant: Shelley Turner (sturner@kisd.org)"),

    # ── GREGG / UPSHUR COUNTY ────────────────────────────────────
    ("GREGG / UPSHUR COUNTY", None),

    ("Gladewater ISD", "Gregg/Upshur",
     "Rae Ann Patty",
     "(903) 845-6991", "pattyr@gladewaterisd.com",
     "200 E. Broadway, Gladewater, TX 75647",
     "https://www.gladewaterisd.com",
     ""),

    # ── RUSK COUNTY ──────────────────────────────────────────────
    ("RUSK COUNTY", None),

    ("Henderson ISD", "Rusk",
     "Brian Bowman",
     "(903) 655-5000", "",
     "300 Crosby Drive, Henderson, TX 75652",
     "https://www.hendersonisd.org",
     "Hired Feb 2025. Likely bbowman@hendersonisd.net"),

    # ── CHEROKEE COUNTY ──────────────────────────────────────────
    ("CHEROKEE COUNTY", None),

    ("Jacksonville ISD", "Cherokee",
     "Brad Stewart",
     "(903) 586-6511", "bradley.stewart@jisd.org",
     "800 College Ave., Jacksonville, TX 75766 (PO Box 631)",
     "https://www.jisd.org",
     ""),

    ("Rusk ISD", "Cherokee",
     "Dr. Mid Johnson",
     "(903) 683-5592", "",
     "203 East 7th St., Rusk, TX 75785",
     "https://www.ruskisd.net",
     "Leadership in transition. Confirm by calling"),

    ("Alto ISD", "Cherokee",
     "Kelly West",
     "(936) 858-7101", "",
     "244 County Rd. 2429, Alto, TX 75925",
     "https://www.alto.esc7.net",
     "Likely kwest@alto.esc7.net"),

    ("Bullard ISD", "Smith/Cherokee",
     "Dr. Micah Dyer (incoming)",
     "(903) 894-6639", "",
     "Bullard, TX 75757",
     "https://www.bullardisd.net",
     "Dr. Jack Lee outgoing. Dr. Dyer named lone finalist Feb 2026"),

    # ── HENDERSON COUNTY ─────────────────────────────────────────
    ("HENDERSON COUNTY", None),

    ("Athens ISD", "Henderson",
     "Dr. Janie Sims",
     "(903) 677-6900", "jsims@athensisd.net",
     "104 Hawn St., Athens, TX 75751",
     "https://www.athensisd.net",
     ""),

    ("Mabank ISD", "Henderson/Kaufman/Van Zandt",
     "Dr. Russell Marshall",
     "(903) 880-1300", "rmarshall@mabankisd.net",
     "310 E. Market St., Mabank, TX 75147",
     "https://www.mabankisd.net",
     "Covers Gun Barrel City area"),

    ("Brownsboro ISD", "Henderson",
     "Dr. Keri Hampton",
     "(903) 852-2321", "",
     "13942 State Hwy 31 E, Brownsboro, TX 75756 (PO Box 465)",
     "https://www.gobearsgo.net",
     "Admin asst: egoode@gobearsgo.net"),

    ("Malakoff ISD", "Henderson",
     "Dr. PJ Winters",
     "(903) 489-1152", "",
     "310 N. Terry, Malakoff, TX 75148",
     "https://www.malakoffisd.org",
     ""),

    ("Trinidad ISD", "Henderson",
     "Matt Mizell",
     "(903) 778-2673", "",
     "105 W. Eaton St., Trinidad, TX 75163",
     "https://www.trinidadisd.com",
     ""),

    ("Cross Roads ISD", "Henderson",
     "Richard Tedder",
     "(903) 489-2001", "",
     "14434 FM 59, Malakoff, TX 75148",
     "https://www.crossroadsisd.org",
     ""),

    ("Eustace ISD", "Henderson",
     "Dr. Coy Holcombe",
     "(903) 425-5151", "cholcombe@eustaceisd.net",
     "320 FM 316 S, Eustace, TX 75124",
     "https://www.eustaceisd.net",
     ""),

    # ── WOOD COUNTY ──────────────────────────────────────────────
    ("WOOD COUNTY", None),

    ("Mineola ISD", "Wood",
     "Cody Mize",
     "(903) 569-2448", "mizec@mineolaisd.net",
     "1695 West Loop 564, Mineola, TX 75773",
     "https://www.mineolaisd.net",
     ""),

    ("Quitman ISD", "Wood",
     "Christopher Mason",
     "(903) 763-5000", "masonc@quitmanisd.net",
     "600 N. Winnsboro St., Quitman, TX 75783",
     "https://www.quitmanisd.net",
     "Supt direct: (903) 760-5022"),

    ("Winnsboro ISD", "Wood/Franklin",
     "Acting: Dr. Cody Holloway",
     "(903) 342-3737", "",
     "505 S. Chestnut St., Winnsboro, TX 75494",
     "https://www.winnsboroisd.org",
     "Leadership in flux — prior supt placed on leave Jan 2026"),

    # ── VAN ZANDT COUNTY ─────────────────────────────────────────
    ("VAN ZANDT COUNTY", None),

    ("Canton ISD", "Van Zandt",
     "Dr. Brian Nichols",
     "(903) 567-4179", "",
     "225 W. Elm, Canton, TX 75103",
     "https://www.cantonisd.net",
     "Secretary: Stacie Wilkerson (stwilk@cantonisd.com). Likely bnichols@cantonisd.com"),

    ("Grand Saline ISD", "Van Zandt",
     "Micah Lewis",
     "(903) 962-7546", "",
     "400 Stadium Dr., Grand Saline, TX 75140",
     "https://www.grandsalineisd.net",
     "Ext 101. Likely mlewis@grandsalineisd.net"),

    ("Wills Point ISD", "Van Zandt",
     "Richard Cooper",
     "(903) 873-5100", "richard.cooper@wpisd.com",
     "338 W. North Commerce St., Wills Point, TX 75169",
     "https://www.wpisd.com",
     ""),

    # ── UPSHUR COUNTY ────────────────────────────────────────────
    ("UPSHUR COUNTY", None),

    ("Gilmer ISD", "Upshur",
     "Rick Albritton",
     "(903) 841-7400", "albrittonr@gilmerisd.org",
     "245 N. Bradford St., Gilmer, TX 75644",
     "https://www.gilmerisd.org",
     ""),

    ("Big Sandy ISD", "Upshur",
     "Mike Burns",
     "(903) 636-5318", "mburns@bigsandyisd.org",
     "401 N. Wildcat Dr., Big Sandy, TX 75755",
     "https://www.bigsandyisd.org",
     "Verify current — Donna Varnado may have served as interim"),

    ("Harmony ISD", "Upshur",
     "David Cochran",
     "(903) 725-5492", "cochrand@harmonyisd.net",
     "9788 SH 154 W, Big Sandy, TX 75755",
     "https://www.harmonyisd.net",
     ""),

    ("Union Grove ISD", "Upshur",
     "Dr. Greg Bower (incoming Apr 2026)",
     "", "",
     "11220 Union Grove Rd., Gladewater, TX 75647",
     "https://www.ugisd.org",
     "Leadership in transition"),

    ("New Diana ISD", "Upshur",
     "Marshall Moore",
     "(903) 663-8000", "mmoore@ndisd.org",
     "1373 US Hwy 259 South, Diana, TX 75640",
     "https://www.ndisd.org",
     ""),

    ("Ore City ISD", "Upshur",
     "Lynn Heflin",
     "(903) 968-3300", "",
     "100 Rebel Rd., Ore City, TX 75683",
     "https://www.ocisd.net",
     "Likely lheflin@ocisd.net"),
]

ws = wb.active
ws.title = "School Districts"

col_widths = [30, 26, 28, 18, 32, 56, 36, 56]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

row_num = 2
for record in districts:
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=record[0])
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 28
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

output = "/home/user/Claude/Surrounding_Counties_School_Districts.xlsx"
wb.save(output)
data_rows = sum(1 for r in districts if r[1] is not None)
print(f"Saved {data_rows} districts to {output}")
