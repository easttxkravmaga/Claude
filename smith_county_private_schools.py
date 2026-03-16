#!/usr/bin/env python3
"""Generate Smith County TX Private Schools – All Employees spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4A235A", end_color="4A235A", fill_type="solid")
school_font = Font(name="Calibri", bold=True, size=13, color="4A235A")
school_fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
dept_font = Font(name="Calibri", bold=True, size=11, color="6C3483")
dept_fill = PatternFill(start_color="F4ECF7", end_color="F4ECF7", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["School", "Department", "Name", "Title", "Phone", "Email", "Address"]

# ── Markers ──
# SCHOOL HEADER: (text, None, None, None, None, None, None) → school_font/fill
# DEPT HEADER:   (text, "DEPT", None, None, None, None, None) → dept_font/fill
# DATA ROW:      (school, dept, name, title, phone, email, address)

data = [
    # ═══════════════════════════════════════════════════════════════════
    ("GRACE COMMUNITY SCHOOL", None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════════════

    # ── Administration ──
    ("Grace Community School — Administration", "DEPT", None, None, None, None, None),

    ("Grace Community School", "Administration",
     "Jay Ferguson, PhD", "Head of School",
     "(903) 566-5678", "jay.ferguson@gracetyler.org",
     "3025 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Administration",
     "Kayla Franks", "Admissions Coordinator",
     "(903) 566-5678 x2114", "kayla.franks@gracetyler.org",
     "3025 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Administration",
     "Amber Bowler", "Director of Admissions – Upper Campus",
     "(903) 566-5661", "amber.bowler@gracetyler.org",
     "3001 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Administration",
     "Carl Hamm", "Director of Advancement",
     "(903) 566-5678", "carl.hamm@gracetyler.org",
     "3025 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Administration",
     "Jessica Huddleston", "Director of Curriculum & Instruction",
     "(903) 566-5678", "jessica.huddleston@gracetyler.org",
     "3025 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Administration",
     "Ijeoma 'EJ' Unegbu, EdD", "Director of Community & Culture",
     "(903) 566-5678", "ijeoma.unegbu@gracetyler.org",
     "3025 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Administration",
     "Tonya Boyd", "Variable Tuition / Finance",
     "(903) 566-5678", "tonya.boyd@gracetyler.org",
     "3025 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Administration",
     "Shari Wilson", "Upper Campus Contact",
     "(903) 566-5661", "shari.wilson@gracetyler.org",
     "3001 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Administration",
     "Wendee Parker", "After School Program / Substitutes",
     "(903) 566-5678", "wendee.parker@gracetyler.org",
     "3025 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Administration",
     "Heather Garrison", "Development Operations & Stewardship",
     "(903) 566-5678", "heather.garrison@gracetyler.org",
     "3025 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Administration",
     "", "General Admissions",
     "(903) 566-5678 x2114", "admissions@gracetyler.org",
     "3025 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Administration",
     "", "Early Education Center",
     "(903) 566-5678", "EE@gracetyler.org",
     "3025 University Blvd., Tyler, TX 75701"),

    # ── Faculty & Coaches ──
    ("Grace Community School — Faculty & Coaches", "DEPT", None, None, None, None, None),

    ("Grace Community School", "Athletics",
     "Eddie Francis", "Athletic Director",
     "(903) 566-5661", "eddie.francis@gracetyler.org",
     "3001 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Athletics",
     "Phil Castles", "Football Defensive Coordinator",
     "(903) 566-5661", "phil.castles@gracetyler.org",
     "3001 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Athletics",
     "Josue Sabillon", "Men's Soccer Head Coach / Spanish Teacher",
     "(903) 566-5661", "josue.sabillon@gracetyler.org",
     "3001 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Faculty",
     "Julie Aldredge", "Teacher",
     "(903) 566-5678", "julie.aldredge@gracetyler.org",
     "Tyler, TX 75701"),

    ("Grace Community School", "Faculty",
     "Michael Anderson", "High School ASC Teacher",
     "(903) 566-5661", "michael.anderson@gracetyler.org",
     "3001 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Faculty",
     "Shelly Chaudoir", "7th Grade TX History / 8th Grade English Teacher",
     "(903) 566-5661", "shelly.chaudoir@gracetyler.org",
     "3001 University Blvd., Tyler, TX 75701"),

    ("Grace Community School", "Faculty",
     "Amy Fleet", "Office Manager",
     "(903) 566-5678", "amy.fleet@gracetyler.org",
     "3025 University Blvd., Tyler, TX 75701"),

    # ═══════════════════════════════════════════════════════════════════
    ("THE BROOK HILL SCHOOL", None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════════════

    # ── Administration ──
    ("The Brook Hill School — Administration", "DEPT", None, None, None, None, None),

    ("The Brook Hill School", "Administration",
     "Braxton Brady", "Head of School",
     "(903) 894-5000", "bbrady@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "Steve Dement", "Founder / Chairman of the Board / US History Teacher",
     "(903) 894-5000", "sdement@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "Matthew Beasley", "Upper School Principal",
     "(903) 894-5000", "mbeasley@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "Michelle Rozell", "Middle School Principal",
     "(903) 894-5000", "mrozell@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "Travis Albea", "Director of Admissions",
     "(903) 894-5000", "talbea@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "Kim Webber", "Director of Human Resources",
     "(903) 894-5000", "kwebber@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "Heidii Godbold", "Dean of Academics",
     "(903) 894-5000", "hgodbold@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "Laurie Humphries", "Director of Advancement",
     "(903) 894-5000", "lhumphries@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "Madison Hanks", "Associate Director of Enrollment & Financial Aid",
     "(903) 894-5000", "mhanks@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "Aimee Lowthorp", "Admissions Associate",
     "(903) 894-5000", "alowthorp@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "Meri Mullicane", "Learning Specialist",
     "(903) 894-5000", "mmullicane@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "Melissa Adkins", "Administrative Assistant – Upper School",
     "(903) 894-5000", "madkins@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "", "General Admissions",
     "(903) 894-5000", "admissions@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Administration",
     "", "General Info",
     "(903) 894-5000", "info@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    # ── Counseling ──
    ("The Brook Hill School — Counseling", "DEPT", None, None, None, None, None),

    ("The Brook Hill School", "Counseling",
     "Ashley Bouwer", "Director of Academic & College Counseling",
     "(903) 894-5000", "abouwer@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Counseling",
     "Jaime Sturdivant", "Academic College Counselor",
     "(903) 894-5000", "jsturdivant@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Counseling",
     "Celia Tucker", "Academic / College Counselor",
     "(903) 894-5000", "ctucker@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    # ── Residence Life ──
    ("The Brook Hill School — Residence Life", "DEPT", None, None, None, None, None),

    ("The Brook Hill School", "Residence Life",
     "Shawn Rhoads", "Senior Director of Residence Life & Ministry",
     "(903) 894-5000", "srhoads@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Residence Life",
     "Hope Cooper", "Assoc. Dir. of Boarding Admissions / Marketing",
     "(903) 894-5000", "hcooper@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Residence Life",
     "Lisa Callens", "Administrative Assistant – Boarding Admissions",
     "(903) 894-5000", "lcallens@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Residence Life",
     "Carrie Brady", "Residence Life Team",
     "(903) 894-5000", "cbrady@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Residence Life",
     "Jessica Longenecker", "Boarding Parent",
     "(903) 894-5000", "jlongenecker@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Residence Life",
     "Nicole Shigley", "Girls Boarding Parent",
     "(903) 894-5000", "nshigley@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Residence Life",
     "Kim Kelley", "Boys Boarding Parent",
     "(903) 894-5000", "kkelley@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    # ── Security ──
    ("The Brook Hill School — Security", "DEPT", None, None, None, None, None),

    ("The Brook Hill School", "Security",
     "Bobby Brasher", "Director of School Security",
     "(903) 894-5000", "bbrasher@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    # ── Athletics & Coaching ──
    ("The Brook Hill School — Athletics & Coaching", "DEPT", None, None, None, None, None),

    ("The Brook Hill School", "Athletics",
     "Wally Dawkins", "Athletic Director",
     "(903) 894-5000", "wdawkins@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Athletics",
     "Don English", "Dir. of Faculty & Staff Culture / Head Baseball Coach",
     "(903) 894-5000", "denglish@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Athletics",
     "Scott Ryle", "Head Football Coach / Academic Advocate",
     "(903) 894-5000", "sryle@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Athletics",
     "David Collins", "Head Varsity Soccer / Asst. Football / Science Teacher (MS)",
     "(903) 894-5000", "dcollins@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Athletics",
     "Jacob Agnew", "Head Boys Basketball Coach",
     "(903) 894-5000", "jagnew@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Athletics",
     "Jereme Hubbard", "Head Girls Basketball / Sports Info Dir. / Science Teacher",
     "(903) 894-5000", "jhubbard@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Athletics",
     "Savanna Wilson", "Head Volleyball / MS Basketball / Athletic Marketing",
     "(903) 894-5000", "swilson@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Athletics",
     "Anthony Springer", "Head Softball / Cross Country / PreK-5 PE",
     "(903) 894-5000", "aspringer@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Athletics",
     "Nick Harrison", "Dean of Students / Head Golf Coach",
     "(903) 894-5000", "nharrison@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Athletics",
     "Tiffany Hubbard", "Director of Guard Spirit Teams",
     "(903) 894-5000", "thubbard@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    # ── Faculty ──
    ("The Brook Hill School — Faculty", "DEPT", None, None, None, None, None),

    ("The Brook Hill School", "Faculty",
     "Jason Hooker", "Dean of Students (MS) / Director of Outreach",
     "(903) 894-5000", "jhooker@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Austin Reed", "Bible Department Chair / Teacher",
     "(903) 894-5000", "areed@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Steven Soldi", "Chair of Business Dept. / Teacher of the Year 2023",
     "(903) 894-5000", "ssoldi@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Kent Travis", "Humanities Teacher (US)",
     "(903) 894-5000", "ktravis@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Michael Bontrager", "Humanities Teacher (US)",
     "(903) 894-5000", "mbontrager@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Ashleigh Emery", "Bible & Humanities Teacher (MS)",
     "(903) 894-5000", "aemery@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Emma French", "Math Teacher (MS) / Drill Team Director / Girls Golf Coach",
     "(903) 894-5000", "efrench@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Sarah Ethridge", "English Teacher (US)",
     "(903) 894-5000", "sethridge@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Keri McDonald", "Spanish Teacher",
     "(903) 894-5000", "kmcdonald@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Lisa Larios", "Art Teacher (LS)",
     "(903) 894-5000", "llarios@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Itzel Davis", "Spanish Teacher (LS)",
     "(903) 894-5000", "idavis@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Krystal Swink", "Innovation / Technology Instructional Support",
     "(903) 894-5000", "kswink@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Jonna Parker", "Pre-K Teacher",
     "(903) 894-5000", "jparker@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Rachel Carlton", "Strings Teacher",
     "(903) 894-5000", "rcarlton@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "Stephanie Attebery", "Lower School Teacher",
     "(903) 894-5000", "sattebery@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    ("The Brook Hill School", "Faculty",
     "", "SPARC Program",
     "(903) 894-5000 x1101", "sparc@brookhill.org",
     "1051 N Houston St, Bullard, TX 75757"),

    # ═══════════════════════════════════════════════════════════════════
    ("ALL SAINTS EPISCOPAL SCHOOL", None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════════════

    # ── Administration ──
    ("All Saints Episcopal School — Administration", "DEPT", None, None, None, None, None),

    ("All Saints Episcopal School", "Administration",
     "Dr. Mark Desjardins", "Head of School",
     "(903) 579-6000", "mdesjardins@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Administration",
     "Amy Alsip, Ed.D.", "Head of Lower School",
     "(903) 579-6000", "aalsip@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Administration",
     "Billy Beasley", "Director of A3 Learning / Dean of Grades 7-12",
     "(903) 579-6000", "bbeasley@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Administration",
     "Gretchen Mercer", "Director of Athletics / Admissions Advisor",
     "(903) 579-6000", "gmercer@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Administration",
     "J. Juarez", "Director of Admissions and Enrollment",
     "(903) 579-6000", "jjuarez@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Administration",
     "J. Walker", "Admissions Contact",
     "(903) 579-6001", "jwalker@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Administration",
     "Kristen Courtney", "Asst. to Head of Early Learning & Lower School / Dir. Extended Care",
     "(903) 579-6000", "kcourtney@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Administration",
     "Emily Allen", "Assistant Director of the LEC",
     "(903) 579-6000", "eallen@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Administration",
     "Drew Starnes", "Associate Director of Athletics / Accounting Asst.",
     "(903) 579-6000", "dstarnes@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Administration",
     "", "Admissions",
     "(903) 579-6009", "admissions@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    # ── Faith Formation ──
    ("All Saints Episcopal School — Faith & Chaplains", "DEPT", None, None, None, None, None),

    ("All Saints Episcopal School", "Faith Formation",
     "Tim Kennedy", "Senior Chaplain / Faith Formation Dept. Chair",
     "(903) 579-6000", "tkennedy@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faith Formation",
     "Leesa Lewis", "Chaplain",
     "(903) 579-6000", "llewis@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faith Formation",
     "Mika Campbell", "Faith Formation Teacher",
     "(903) 579-6000", "mcampbell@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    # ── Faculty ──
    ("All Saints Episcopal School — Faculty", "DEPT", None, None, None, None, None),

    ("All Saints Episcopal School", "Faculty",
     "Jean Arnold", "Collaboratory Teacher",
     "(903) 579-6000 x3212", "jarnold@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "Sarah White", "US Math Teacher / Math Dept. Chair",
     "(903) 579-6000", "swhite@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "Tom Marsh", "Upper School History Teacher",
     "(903) 579-6000", "tmarsh@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "Josh Jones", "Upper School Math Teacher",
     "(903) 579-6000", "jjones@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "David Lambert", "Upper School Science Teacher",
     "(903) 579-6000", "dlambert@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "DebraAnn Parham", "MS & US Science / Yearbook / MS Contest Coordinator",
     "(903) 579-6000", "dparham@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "Jay Reitman", "Chemistry Teacher",
     "(903) 579-6000", "jreitman@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "T. Collings", "US Science Teacher / Varsity Baseball Coach",
     "(903) 579-6000", "tcollings@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "E. Dunaway", "Intermediate/MS History Teacher",
     "(903) 579-6000", "edunaway@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "Silvia Capistran", "Upper School Spanish Teacher",
     "(903) 579-6000", "scapistran@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "Moni Cobb", "MS & US Spanish Teacher",
     "(903) 579-6000", "mcobb@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "J. Courtney", "MS/US Art Teacher",
     "(903) 579-6000", "jcourtney@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "Jay Jordan", "MS Elective Teacher / US Entrepreneur Class",
     "(903) 579-6000", "jjordan@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "J. Way", "US Math Teacher / Math Dept. Chair",
     "(903) 579-6000", "jway@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "Desmond Barron", "4th Grade Teacher",
     "(903) 579-6000", "dbarron@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "M. Meredith", "4th Grade Teacher",
     "(903) 579-6000", "mmeredith@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "Katie Cargill", "Lower School Teacher",
     "(903) 579-6000", "kcargill@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Faculty",
     "Casey Copfer", "Lower School Teacher",
     "(903) 579-6000", "ccopfer@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    # ── Athletics / Coaching ──
    ("All Saints Episcopal School — Athletics & Coaching", "DEPT", None, None, None, None, None),

    ("All Saints Episcopal School", "Athletics",
     "B. Winegeart", "Asst. Director of Athletics / Head Volleyball",
     "(903) 579-6000", "bwinegeart@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Kyle Freeman", "Head Football / Powerlifting",
     "(903) 579-6000", "kfreeman@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "David Benbow", "MS/US Football / MS/US Track",
     "(903) 579-6000", "dbenbow@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Ricky Collins", "MS/US Football / Track / Strength & Speed Coach",
     "(903) 579-6000", "rcollins@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Aston Francis", "Head Boys Basketball / Head Golf",
     "(903) 579-6000", "afrancis@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Jeremy Tatum", "Head Girls Basketball / Football / Tennis",
     "(903) 579-6000", "jtatum@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "D. Havard", "PE Teacher / Varsity Boys Soccer Coach",
     "(903) 579-6000", "dhavard@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Stacey Bramlet", "Head Cross Country / Head Track",
     "(903) 579-6000", "sbramlet@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Kyle Edgemon", "Head Tennis Coach",
     "(903) 579-6000", "kedgemon@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "C. Singer", "Varsity Tennis / Asst. Track Coach",
     "(903) 579-6000", "csinger@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "M. Rishel", "Varsity Swimming Coach",
     "(903) 579-6000", "mrishel@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Geovana Almeida", "Asst. Volleyball / Asst. Girls Soccer",
     "(903) 579-6000", "galmeida@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Emily Adams", "Head Cheer",
     "(903) 579-6000", "eadams@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Betsy Stith", "MS Cheer Head Coach",
     "(903) 579-6000", "bstith@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Marsha Phillips", "MS Volleyball / MS Girls Basketball / MS Track",
     "(903) 579-6000", "mphillips@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Allison Motto", "MS Volleyball",
     "(903) 579-6000", "amotto@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Brian Motto", "Head Coach",
     "(903) 579-6000", "bmotto@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Emily Brady", "Athletic Trainer / Varsity Softball Coach",
     "(903) 579-6000", "ebrady@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Christina Huffman", "Sponsor",
     "(903) 579-6000", "chuffman@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    ("All Saints Episcopal School", "Athletics",
     "Madison Granberry", "Sponsor",
     "(903) 579-6000", "mgranberry@all-saints.org",
     "2695 S SW Loop 323, Tyler, TX 75701"),

    # ═══════════════════════════════════════════════════════════════════
    ("BISHOP THOMAS K. GORMAN CATHOLIC SCHOOL", None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════════════

    # ── Administration ──
    ("Bishop Gorman — Administration", "DEPT", None, None, None, None, None),

    ("Bishop T.K. Gorman Catholic School", "Administration",
     "Very Rev. Hank Lanik", "President, Tyler Catholic Schools",
     "(903) 561-2424", "hlanik@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Administration",
     "John Kimec", "Principal",
     "(903) 579-9400", "jkimec@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Administration",
     "Mary Schick", "Assistant Principal / Science Dept.",
     "(903) 579-9404", "mschick@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Administration",
     "Ed Burns", "Dean of Discipline / Head Football Coach",
     "(903) 561-2424", "eburns@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Administration",
     "Deacon Keith Fournier", "Dean of Catholic Identity",
     "(903) 579-9413", "kfournier@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Administration",
     "Jan Pilgrim", "Director of Business Operations",
     "(903) 561-2424", "jpilgrim@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Administration",
     "Molly Saenz", "Director of Admissions & Advancement",
     "(903) 579-9408", "admissions@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Administration",
     "Michael Broaddus", "Director of Technology",
     "(903) 579-9430", "mbroaddus@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Administration",
     "Jill Coan", "Director of Academic & College Advising",
     "(903) 579-9409", "jcoan@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Administration",
     "", "General Info",
     "(903) 561-2424", "info@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    # ── Faculty ──
    ("Bishop Gorman — Faculty", "DEPT", None, None, None, None, None),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Christine Martin", "Choir Director",
     "(903) 561-2424", "cmartin@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Howard Galletly", "Band Director",
     "(903) 561-2424", "hgalletly@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Dwight Anderson", "Instructor of Orchestra",
     "(903) 561-2424", "danderson@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Angel Arambula", "Instructor of Spanish",
     "(903) 561-2424", "aarambula@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "John Ballew", "Instructor of Industrial Arts",
     "(903) 561-2424", "jballew@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Michael Calhoun", "Instructor of Math",
     "(903) 561-2424", "mcalhoun@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Shai Calhoun", "Instructor of Math",
     "(903) 561-2424", "scalhoun@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Angela Carney", "Instructor of Science",
     "(903) 561-2424", "acarney@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Cathy Carney", "Instructor of Math",
     "(903) 561-2424", "ccarney@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Erin Carney", "Instructor, Gorman Learning Center",
     "(903) 561-2424", "ecarney@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Judy Carney", "Instructor of Theology",
     "(903) 561-2424", "jcarney@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Brian Crawford", "Instructor, Gorman Learning Center",
     "(903) 561-2424", "bcrawford@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Lisa Carlisle", "Librarian / News & Broadcasting",
     "(903) 579-9426", "lcarlisle@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Jeanie Oxler", "Instructor of Choir",
     "(903) 561-2424", "joxler@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Amanda Wade", "Instructor of Drama",
     "(903) 561-2424", "awade@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Faculty",
     "Brady McCoy", "Government / Economics / World History",
     "(903) 561-2424", "bmccoy@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    # ── Athletics ──
    ("Bishop Gorman — Athletics", "DEPT", None, None, None, None, None),

    ("Bishop T.K. Gorman Catholic School", "Athletics",
     "Joseph Murray", "Athletic Director",
     "(903) 561-2424", "jmurray@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Athletics",
     "Monica Davis", "Athletic Director & Coach (Cross Country, Girls Track)",
     "(903) 579-9440", "mdavis@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Athletics",
     "Kevin Murray", "Asst. Athletic Director / Head Boys Basketball",
     "(903) 561-2424", "kmurray@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Athletics",
     "Bernard Anderson", "Head Varsity Women's Basketball Coach",
     "(903) 561-2424", "banderson@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Athletics",
     "Jose Carmona", "Head Varsity Soccer Coach",
     "(903) 561-2424", "jcarmona@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Athletics",
     "Michael Holochuck", "Head Baseball Coach",
     "(903) 561-2424", "mholochuck@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Athletics",
     "Tynnecia Malbroue", "Head Varsity Volleyball Coach",
     "(903) 561-2424", "tmalbroue@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Athletics",
     "Julia Eddings", "Head Varsity & MS Cheer Coach",
     "(903) 561-2424", "jeddings@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Athletics",
     "Marcus Townsend", "Head MS Football / Head MS Baseball",
     "(903) 561-2424", "mtownsend@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Athletics",
     "Felipe Natera", "IT Assistant / Visual & Media Director / Coach",
     "(903) 579-9449", "fnatera@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),

    ("Bishop T.K. Gorman Catholic School", "Athletics",
     "Dale Carney", "Director of Transportation",
     "(903) 561-2424", "dcarney@bishopgorman.net",
     "1405 E SE Loop 323, Tyler, TX 75701"),
]


# ── Build worksheet ─────────────────────────────────────────────────────
ws = wb.active
ws.title = "Private Schools – All Employees"

# Column widths
col_widths = [38, 22, 28, 44, 22, 36, 44]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Write headers
for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

# Write data
row_num = 2
for record in data:
    label = record[0]

    # SCHOOL HEADER
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=label)
        cell.font = school_font
        cell.fill = school_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 30
        row_num += 1
        continue

    # DEPT HEADER
    if record[1] == "DEPT":
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=label)
        cell.font = dept_font
        cell.fill = dept_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 22
        row_num += 1
        continue

    # DATA ROW
    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

# ── Save ────────────────────────────────────────────────────────────────
output_path = "/home/user/Claude/Smith_County_TX_Private_Schools_Employees.xlsx"
wb.save(output_path)

# Count data rows (not headers)
data_rows = sum(1 for r in data if r[1] is not None and r[1] != "DEPT")
print(f"Saved {data_rows} employee contacts to {output_path}")
