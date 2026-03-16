#!/usr/bin/env python3
"""Generate Surrounding Counties of Smith County TX Homeschool Groups spreadsheet (email required)."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
county_font = Font(name="Calibri", bold=True, size=13, color="1B4332")
county_fill = PatternFill(start_color="B7E4C7", end_color="B7E4C7", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Organization", "County", "Type", "Name", "Title", "Phone", "Email", "Address", "Website"]

# ── Data — ONLY entries with a confirmed email ──────────────────────────
# None in County = county header row

groups = [
    # ═══════════════════════════════════════════════════════════
    ("GREGG COUNTY (Longview, Kilgore, Gladewater, Hallsville)", None, None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("CHEC - Christian Home Educators Community of Longview", "Gregg", "Co-op / Support Group",
     "", "General Contact",
     "", "checlongview@gmail.com",
     "P.O. Box 224, Longview, TX 75606",
     "https://www.checlongview.com"),

    ("CHEC - East Texas Homeschool Sports", "Gregg", "Athletics",
     "Shane Wilson", "Athletic Director",
     "", "etxhssports@gmail.com",
     "Longview, TX",
     "https://www.checlongview.com"),

    ("LARHE - Longview Area Relaxed Home Educators", "Gregg", "Support Group",
     "", "Group Contact",
     "", "larhehomeschool@gmail.com",
     "Longview, TX",
     "https://www.larhe.com"),

    ("East Texas Homeschool Sports (Chargers)", "Gregg", "Athletics",
     "Brandon Allee", "Athletic Director",
     "", "etxchargers@gmail.com",
     "104 Persimmon Hl, Hallsville, TX 75650",
     "https://etxchargers.org"),

    # ═══════════════════════════════════════════════════════════
    ("RUSK COUNTY (Henderson)", None, None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("Henderson / Rusk County Homeschool Group", "Rusk", "Support Group",
     "Amy Dumas", "Contact",
     "(903) 847-6303", "aldumas@gmail.com",
     "9570 CR 4108 W, Henderson, TX 75654",
     ""),

    # ═══════════════════════════════════════════════════════════
    ("HENDERSON COUNTY (Athens, Gun Barrel City)", None, None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("Athens Homeschool Group", "Henderson", "Support Group",
     "Gayle Furlong", "Contact",
     "(903) 469-4031", "furlongs9@earthlink.net",
     "P.O. Box 312, Murchison, TX 75778",
     "http://www.athenshomeschool.com"),

    ("East Texas Arboretum - Homeschool Classes", "Henderson", "Enrichment Program",
     "", "Program Contact",
     "(903) 675-5630", "eta@easttexasarboretum.org",
     "1601 Patterson Rd., Athens, TX 75751",
     "https://www.easttexasarboretum.org/homeschool.html"),

    # ═══════════════════════════════════════════════════════════
    ("WOOD COUNTY (Mineola, Quitman, Winnsboro)", None, None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("CHAT of Mineola (Christian Homeschoolers Assembling Together)", "Wood", "Co-op",
     "Jennifer Little", "Director",
     "", "chatofmineola@gmail.com",
     "Harvest Acres Baptist Church, 460 NW Loop 564, Mineola, TX 75773",
     "https://chatofmineola.weebly.com"),

    # ═══════════════════════════════════════════════════════════
    ("VAN ZANDT COUNTY (Canton, Grand Saline, Wills Point)", None, None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("EXEL Homeschool Co-op", "Van Zandt", "Co-op",
     "", "Co-op Contact",
     "", "Exelcoop@gmail.com",
     "Canton, TX 75103",
     "https://www.homeschool-life.com/1874/"),

    ("EXEL Athletic Association", "Van Zandt", "Athletics",
     "", "Athletic Association",
     "", "ExelAthletics@gmail.com",
     "Canton, TX (serves Van Zandt, Wood, Smith, Gregg, Cherokee counties)",
     "https://exelathletics.weebly.com"),

    # ═══════════════════════════════════════════════════════════
    ("UPSHUR COUNTY (Gilmer, Big Sandy)", None, None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════
    # Note: East Texas Homeschool Convention (Gilmer) — no confirmed email found.
    # No groups with confirmed emails found in Upshur County.

    # ═══════════════════════════════════════════════════════════
    ("CHEROKEE COUNTY (Jacksonville, Rusk)", None, None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════
    # Note: No homeschool groups with confirmed emails found in Cherokee County.
    # Families in the area tend to use Tyler-based groups like TACHE.
]


# ── Build worksheet ─────────────────────────────────────────────────────
ws = wb.active
ws.title = "Surrounding Counties Homeschool"

# Column widths
col_widths = [50, 14, 22, 22, 20, 16, 32, 52, 42]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Write headers
for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

# Write data
row_num = 2
for record in groups:
    org = record[0]
    # County header rows (County is None)
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=org)
        cell.font = county_font
        cell.fill = county_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 28
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

# ── Save ────────────────────────────────────────────────────────────────
output_path = "/home/user/Claude/Surrounding_Counties_Homeschool_Groups.xlsx"
wb.save(output_path)

data_rows = sum(1 for r in groups if r[1] is not None)
print(f"Saved {data_rows} contacts to {output_path}")
