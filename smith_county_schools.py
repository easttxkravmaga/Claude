#!/usr/bin/env python3
"""Generate Smith County TX Schools Contacts spreadsheet (email required)."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="7B241C", end_color="7B241C", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=13, color="7B241C")
section_fill = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
subsection_font = Font(name="Calibri", bold=True, size=11, color="922B21")
subsection_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Organization", "Name", "Title", "Phone", "Email", "Address"]

# SECTION = (text, None, None, None, None, None) with style marker
# SUBSECTION = (text, "SUB", None, None, None, None)
# DATA = (org, name, title, phone, email, address)

data = [
    # ═══════════════════════════════════════════════════════════
    ("UNIVERSITIES & COLLEGES", None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("University of Texas at Tyler", "Katie Ramirez", "AVP of Recruitment & Admissions",
     "(903) 566-7204", "kramirez@uttyler.edu",
     "3900 University Blvd., Tyler, TX 75799"),

    ("University of Texas at Tyler", "Rudy Alarcon", "Admissions Officer - Greater Tyler",
     "(903) 566-5863", "ralarcon@uttyler.edu",
     "3900 University Blvd., Tyler, TX 75799"),

    ("University of Texas at Tyler", "", "Undergraduate Admissions",
     "(903) 566-7203", "admissions@uttyler.edu",
     "3900 University Blvd., Tyler, TX 75799"),

    ("University of Texas at Tyler", "", "Graduate Admissions",
     "(903) 566-7457", "gradadmissions@uttyler.edu",
     "3900 University Blvd., Tyler, TX 75799"),

    ("University of Texas at Tyler", "", "Office of the Registrar",
     "(903) 566-7180", "enroll@uttyler.edu",
     "3900 University Blvd., Tyler, TX 75799"),

    ("Tyler Junior College", "", "Admissions Office",
     "(903) 510-2523", "Admissions@TJC.edu",
     "1327 S Baxter Ave, Tyler, TX 75701"),

    ("Tyler Junior College", "Britt Sabota", "Registrar",
     "(903) 510-2200", "bsab@tjc.edu",
     "1327 S Baxter Ave, Tyler, TX 75701"),

    ("Texas College (HBCU)", "John Roberts", "Dean of Enrollment Mgmt & Registrar",
     "(903) 593-8311 x2297", "admissions1@texascollege.edu",
     "2404 N. Grand Avenue, Tyler, TX 75702"),

    ("Texas College (HBCU)", "Mrs. A. Copeland", "Asst. Director of Financial Aid",
     "(903) 593-8311 x2299", "financialaid@texascollege.edu",
     "2404 N. Grand Avenue, Tyler, TX 75702"),

    ("Texas College (HBCU)", "Dr. C. Marshall-Biggins", "VP for Student Affairs",
     "(903) 593-8311 x2710", "studentaffairs1@texascollege.edu",
     "2404 N. Grand Avenue, Tyler, TX 75702"),

    ("Texas College (HBCU)", "", "Office of the Registrar",
     "(903) 593-8311 x2233", "registrar2@texascollege.edu",
     "2404 N. Grand Avenue, Tyler, TX 75702"),

    # ═══════════════════════════════════════════════════════════
    ("PRIVATE SCHOOLS", None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("All Saints Episcopal School", "", "Admissions",
     "(903) 579-6009", "admissions@all-saints.org",
     "2695 S Southwest Loop 323, Tyler, TX"),

    ("All Saints Episcopal School", "J. Walker", "Admissions Contact",
     "(903) 579-6001", "jwalker@all-saints.org",
     "2695 S Southwest Loop 323, Tyler, TX"),

    ("Bishop Gorman Catholic School", "", "Admissions",
     "", "admissions@bishopgorman.net",
     "Tyler, TX (Diocese of Tyler)"),

    ("Grace Community School", "", "Admissions",
     "(903) 566-5678 x2114", "admissions@gracetyler.org",
     "3001 University Blvd., Tyler, TX 75701"),

    # ═══════════════════════════════════════════════════════════
    ("TYLER ISD", None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("Tyler ISD - District Office", "Dr. Marty Crawford", "Superintendent",
     "(903) 262-1000", "marty.crawford@tylerisd.org",
     "1319 Earl Campbell Pkwy, Tyler, TX 75701"),

    ("Tyler ISD - Communications", "Jennifer Hines", "Chief Communications Officer",
     "(903) 262-1064", "jennifer.hines@tylerisd.org",
     "1319 Earl Campbell Pkwy, Tyler, TX 75701"),

    ("Tyler ISD - Tyler High School", "", "Campus Administration",
     "(903) 262-2850", "craig.weaver@tylerisd.org",
     "Tyler, TX"),

    # ═══════════════════════════════════════════════════════════
    ("LINDALE ISD", None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("Lindale ISD - District Office", "Stan Surratt", "Superintendent",
     "(903) 881-4001", "surrattsm@lisdeagles.net",
     "505 Pierce St, Lindale, TX 75771"),

    ("Lindale ISD - Communications", "Courtney Sanguinetti", "Director of Communications",
     "(903) 881-4008", "sanguinetticd@lisdeagles.net",
     "505 Pierce St, Lindale, TX 75771"),

    ("Lindale High School", "David French", "Assistant Principal",
     "(903) 881-4050", "frenchdl@lisdeagles.net",
     "920 East Hubbard, Lindale, TX 75771"),

    ("Lindale High School", "Ryan Tomlin", "Campus Behavior Coordinator",
     "(903) 881-4050", "tomlinmr@lisdeagles.net",
     "920 East Hubbard, Lindale, TX 75771"),

    ("Lindale Junior High", "Chad Hodges", "Campus Behavior Coordinator",
     "(903) 881-4150", "hodgesca@lisdeagles.net",
     "15000 CR 463, Lindale, TX 75771"),

    ("E.J. Moss Intermediate", "Steven Hitt", "Campus Behavior Coordinator",
     "(903) 881-4200", "hittsw@lisdeagles.net",
     "411 Eagle Spirit Drive, Lindale, TX 75771"),

    ("Velma Penny Elementary", "Katie Standerfer", "Campus Behavior Coordinator",
     "(903) 881-4250", "standerferka@lisdeagles.net",
     "1000 Mt. Sylvan Hwy, Lindale, TX 75771"),

    # ═══════════════════════════════════════════════════════════
    ("WHITEHOUSE ISD", None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("Whitehouse ISD - District Office", "Dr. Casey Whittle", "Superintendent",
     "(903) 839-5500", "whittlec@whitehouseisd.org",
     "Whitehouse, TX 75791"),

    ("Whitehouse ISD - Student Services", "Dr. Maricela Helm", "Exec. Dir. of Student Services",
     "(903) 839-5500", "helmm@whitehouseisd.org",
     "Whitehouse, TX 75791"),

    ("Whitehouse ISD - Communications", "Clint Williamson", "Dir. of Communications & PR",
     "(903) 839-5500 x6175", "williamsonc@whitehouseisd.org",
     "Whitehouse, TX 75791"),

    ("Whitehouse ISD - Communications", "Courtney Coleman", "Communications & PR Specialist",
     "(903) 839-5500", "colemanc@whitehouseisd.org",
     "Whitehouse, TX 75791"),

    ("Whitehouse ISD - Auxiliary Services", "Scott McFadden", "Dir. of Student Auxiliary Services",
     "(903) 839-5570", "mcfaddens@whitehouseisd.org",
     "Whitehouse, TX 75791"),

    # ═══════════════════════════════════════════════════════════
    ("BULLARD ISD", None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("Bullard ISD - District Office", "Jan Hill", "Superintendent",
     "(903) 894-6639", "superintendent@bullardisd.net",
     "1426-B S Houston, Bullard, TX 75757"),

    # ═══════════════════════════════════════════════════════════
    ("ARP ISD", None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("Arp ISD - District Office", "", "District Office",
     "(903) 859-8482", "web@arpisd.org",
     "101 Toney Drive, Arp, TX 75750"),

    ("Arp High School", "Mike Miller", "Principal",
     "(903) 859-4917", "mikem@arpisd.org",
     "101 Toney Drive, Arp, TX 75750"),

    ("Arp Jr. High School", "Jason Shieldes", "Principal",
     "(903) 859-4936", "jasons@arpisd.org",
     "Arp, TX 75750"),

    # ═══════════════════════════════════════════════════════════
    ("TROUP ISD", None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("Troup ISD - District Office", "Tammy Jones", "Superintendent",
     "(903) 842-3067", "tjones@troupisd.org",
     "201 North Carolina, PO Box 578, Troup, TX 75789"),

    ("Troup ISD - District Office", "Cindy Wilson", "Secretary to Superintendent",
     "(903) 842-3067", "cwilson@troupisd.org",
     "201 North Carolina, Troup, TX 75789"),

    ("Troup ISD - Curriculum", "David Smith", "Exec. Dir. of Curriculum & Instruction",
     "(903) 842-3067", "dsmith@troupisd.org",
     "201 North Carolina, Troup, TX 75789"),

    # ═══════════════════════════════════════════════════════════
    ("WINONA ISD", None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("Winona ISD - District Office", "Justin Cowart", "Superintendent",
     "(903) 939-4010", "jcowart@winonaisd.org",
     "611 Wildcat Dr, Winona, TX 75792"),

    ("Winona ISD - Business Office", "Sheila Bowie", "Business Manager",
     "(903) 939-4021", "sbowie@winonaisd.org",
     "611 Wildcat Dr, Winona, TX 75792"),

    ("Winona ISD - Business Office", "Kayla Lee", "Accounts Payable",
     "(903) 939-4001", "klee@winonaisd.org",
     "611 Wildcat Dr, Winona, TX 75792"),

    ("Winona ISD - Payroll", "Michele Philpot", "Payroll",
     "(903) 939-4022", "mphilpot@winonaisd.org",
     "611 Wildcat Dr, Winona, TX 75792"),
]


# ── Build worksheet ─────────────────────────────────────────────────────
ws = wb.active
ws.title = "Smith County TX Schools"

col_widths = [40, 28, 36, 24, 34, 48]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

row_num = 2
for record in data:
    org = record[0]
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=org)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 28
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

output_path = "/home/user/Claude/Smith_County_TX_Schools.xlsx"
wb.save(output_path)
print(f"Saved {row_num - 2} rows to {output_path}")
