#!/usr/bin/env python3
"""Generate Smith County TX Churches Contacts spreadsheet (email required)."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="5B2C6F", end_color="5B2C6F", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=12, color="4A235A")
section_fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Organization", "Denomination", "Name", "Title", "Phone", "Email", "Address", "Website"]

# ── Data — ONLY entries with a confirmed email ──────────────────────────
# Format: (Organization, Denomination, Name, Title, Phone, Email, Address, Website)
churches = [
    # ── BAPTIST ──
    ("BAPTIST CHURCHES", None, None, None, None, None, None, None),

    ("Tyler Primitive Baptist Church", "Primitive Baptist",
     "Elder Robert Adam", "Pastor",
     "(903) 534-1998", "tylerpbc@gmail.com",
     "6502 Old Jacksonville Hwy, Tyler, TX 75703",
     "https://www.tylerpb.org"),

    ("Tyland Baptist Church", "Baptist",
     "Bro. Scott Moore", "Pastor",
     "", "broscott@tylandbaptist.com",
     "2818 Silver Creek Dr., Tyler, TX 75702",
     "https://tylandbaptist.com"),

    ("Grace Baptist Church", "Baptist",
     "M. Ray Bratcher", "Pastor",
     "(903) 561-7664", "Contactus@gracebaptisttyler.com",
     "5512 Old Jacksonville Hwy, Tyler, TX 75703",
     "https://www.gracebaptisttyler.com"),

    ("The Woods Baptist Church", "Baptist",
     "Warren Ethridge", "Pastor",
     "", "warren@thewoodsbaptist.com",
     "Tyler, TX (East side near UT Tyler)",
     "https://thewoodsbaptist.com"),

    ("Central Baptist Church", "Baptist",
     "", "Church Office",
     "(903) 561-6361", "info@centraltyler.org",
     "1343 E Grande Blvd, Tyler, TX",
     "https://centraltyler.org"),

    ("Family Way Baptist Cathedral of Praise", "Baptist",
     "Dr. Stanley S. Wooden", "Pastor",
     "(903) 245-1022", "stanley.wooden@yahoo.com",
     "1102 Augusta Street, Tyler, TX 75701",
     ""),

    ("Noonday Baptist Church", "Baptist (SBC)",
     "Scott Gorbett", "Pastor",
     "(903) 561-0457", "noondaybaptistchurch1876@gmail.com",
     "16701 County Road 196, Tyler, TX 75703",
     "http://www.noondaybaptistchurch.org"),

    ("Mt. Carmel Baptist Church", "Baptist (SBC)",
     "", "Church Office",
     "(903) 839-2606", "mcbcwhitehouse@gmail.com",
     "10519 FM 344 E, Whitehouse, TX 75791",
     "https://mcbcwhitehouse.com"),

    ("Green Acres Baptist Church", "Baptist (SBC)",
     "Dr. Michael Gossett", "Senior Pastor",
     "(903) 525-1100", "women@mail.gabc.org",
     "1607 Troup Hwy, Tyler, TX 75701",
     "https://www.gabc.org"),

    # ── METHODIST ──
    ("METHODIST CHURCHES", None, None, None, None, None, None, None),

    ("Marvin Methodist Church", "Global Methodist",
     "", "Church Office",
     "(903) 592-7396", "info@marvin.church",
     "300 W. Erwin Street, Tyler, TX 75702",
     "https://www.marvin.church"),

    ("Cedar Street UMC", "United Methodist",
     "Karen Jones", "Lead Pastor",
     "(903) 592-4704", "Revkaren45@aol.com",
     "1420 N Church Ave, Tyler, TX 75702",
     ""),

    # ── PRESBYTERIAN ──
    ("PRESBYTERIAN CHURCHES", None, None, None, None, None, None, None),

    ("First Presbyterian Church", "Presbyterian (PCUSA)",
     "Rev. Dr. Stuart G. Baskin", "Pastor",
     "(903) 597-6317", "church@fpctyler.com",
     "230 W Rusk St, Tyler, TX 75701",
     "http://www.fpctyler.com"),

    ("Tyler Orthodox Presbyterian Church", "Orthodox Presbyterian (OPC)",
     "Pastor Heaton", "Pastor",
     "(903) 839-3535", "heaton.1@opc.org",
     "4554 Farm to Market 2813, Tyler, TX 75762",
     "https://www.tyleropc.org"),

    ("Redeemer Presbyterian Church", "Presbyterian (PCA)",
     "Rev. Ben Wheeler", "Pastor",
     "(903) 805-1809", "ben@redeemertyler.com",
     "P.O. Box 132191, Tyler, TX 75713",
     "https://www.redeemertyler.com"),

    # ── EPISCOPAL ──
    ("EPISCOPAL CHURCHES", None, None, None, None, None, None, None),

    ("Christ Episcopal Church", "Episcopal",
     "", "Church Office",
     "(903) 597-9854", "office@christchurchtyler.org",
     "118 South Bois d'Arc, Tyler, TX 75702",
     "https://www.christchurchtyler.org"),

    ("St. Francis Episcopal Church", "Episcopal",
     "", "Church Office",
     "(903) 593-8459", "stfrancischurchtyler@gmail.com",
     "3232 Jan Avenue, Tyler, TX 75701",
     "https://www.stfrancistyler.org"),

    # ── CHURCH OF CHRIST ──
    ("CHURCHES OF CHRIST", None, None, None, None, None, None, None),

    ("Glenwood Church of Christ", "Church of Christ",
     "", "Church Office",
     "(903) 509-9494", "glenwoodcofc@tyler.net",
     "5210 Hollytree Drive, Tyler, TX 75703",
     ""),

    ("North Tenneha Church of Christ", "Church of Christ",
     "Edward Robinson", "Minister",
     "(903) 593-6868", "ntccminister@ntcconline.org",
     "1701 North Tenneha Avenue, Tyler, TX 75702",
     "https://www.ntcconline.org"),

    # ── CHRISTIAN (DISCIPLES) ──
    ("CHRISTIAN / DISCIPLES OF CHRIST", None, None, None, None, None, None, None),

    ("First Christian Church of Tyler", "Christian (Disciples of Christ)",
     "Dave Hill", "Missions Contact",
     "(903) 561-8138", "firstchristian@fcctyler.org",
     "4202 S Broadway Ave, Tyler, TX 75701",
     "https://fcctyler.org"),

    # ── ASSEMBLY OF GOD ──
    ("ASSEMBLY OF GOD CHURCHES", None, None, None, None, None, None, None),

    ("Calvary Assembly of God", "Assembly of God",
     "", "Church Office",
     "(903) 597-6120", "calvaryagchurch7@gmail.com",
     "1723 East Front, Tyler, TX 75702",
     "https://www.calvaryagtyler.com"),

    ("Tyler First Assembly of God", "Assembly of God",
     "", "Church Office",
     "(903) 597-9804", "info@firstagtyler.com",
     "5309 Rhones Quarter Road, Tyler, TX 75707",
     "https://www.firstagtyler.com"),

    # ── NON-DENOMINATIONAL ──
    ("NON-DENOMINATIONAL CHURCHES", None, None, None, None, None, None, None),

    ("New Life Worship Center", "Non-Denominational",
     "", "Church Office",
     "(903) 871-8700", "info@newlifetyler.com",
     "18535 Highway 69 South, Tyler, TX 75703",
     "https://newlifetyler.com"),

    ("Tyler Christian Fellowship", "Non-Denominational",
     "", "Church Office",
     "", "tcftyler@gmail.com",
     "Tyler, TX",
     "https://www.tcftyler.com"),

    ("Rose Heights Church", "Non-Denominational",
     "Doug Anderson", "Lead Pastor",
     "(903) 566-2080", "info@roseheights.org",
     "2120 Old Omen Rd, Tyler, TX 75701",
     "https://www.roseheights.org"),

    ("Rose Heights Church - Lindale Campus", "Non-Denominational",
     "Allen Townsend", "Campus Pastor",
     "(903) 881-5260", "info@roseheights.org",
     "12465 FM 16, Lindale, TX 75771",
     "https://www.roseheights.org/lindale/"),

    ("Redemption Church", "Non-Denominational",
     "", "Church Office",
     "(903) 561-0174", "info@redemptiontyler.com",
     "16844 CR 165, Gresham, TX 75703",
     "https://redemptiontyler.com"),

    ("Trinity Fellowship Church of Tyler", "Non-Denominational / Charismatic",
     "", "Church Office",
     "(903) 566-4226", "info@tfctyler.com",
     "10344 Hwy 31 E, Tyler, TX 75705",
     "https://www.tfctyler.com"),
]


# ── Build worksheet ─────────────────────────────────────────────────────
ws = wb.active
ws.title = "Smith County TX Churches"

# Column widths
col_widths = [40, 30, 28, 20, 18, 34, 48, 40]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Write headers
for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

# Write data
row_num = 2
for record in churches:
    org = record[0]
    # Section header rows
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=org)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 26
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:H{row_num - 1}"

# ── Save ────────────────────────────────────────────────────────────────
output_path = "/home/user/Claude/Smith_County_TX_Churches.xlsx"
wb.save(output_path)
print(f"Saved {row_num - 2} rows to {output_path}")
