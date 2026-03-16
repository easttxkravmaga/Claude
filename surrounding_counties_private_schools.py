#!/usr/bin/env python3
"""Generate Surrounding Counties Private Schools spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=13, color="1B4332")
section_fill = PatternFill(start_color="B7E4C7", end_color="B7E4C7", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["School", "County", "City", "Type", "Grades", "Head of School", "Title", "Phone", "Email", "Address", "Website", "Notes"]

schools = [
    # ── GREGG COUNTY ─────────────────────────────────────────────
    ("GREGG COUNTY (Longview, Kilgore)", None),

    ("Trinity School of Texas", "Gregg", "Longview", "Episcopal / College-Prep", "PK-12",
     "Cindy Verhalen", "Head of School",
     "(903) 753-0612", "office@trinityschooloftexas.com",
     "215 N Teague St, Longview, TX 75601",
     "https://trinityschooloftexas.com",
     "~255 students, 6:1 ratio, Niche A+. Also: info@trinityschooloftexas.com"),

    ("Longview Christian School (LCS)", "Gregg", "Longview", "Christian (Non-denom)", "PK-12",
     "Ben Cammack", "Administrator",
     "(903) 297-3501", "info@lcseagles.com",
     "1236 Pegues Place, Longview, TX 75601",
     "https://www.lcseagles.com",
     "~456-505 students. Founded 1982. Also: principal@lcseagles.com"),

    ("Christian Heritage Classical School", "Gregg", "Longview", "Christian Classical", "PK-12",
     "Stephanie Jones", "Head of School",
     "(903) 663-4151", "",
     "2715 FM 1844, Longview, TX 75605",
     "https://www.chcslongview.com",
     "~272 students, 6:1 ratio, Niche B+. Founded 1994"),

    ("St. Mary's Catholic School", "Gregg", "Longview", "Roman Catholic (Diocese of Tyler)", "PK-12",
     "Dr. Darbie Safford", "Principal",
     "(903) 753-1657", "dsafford@stmaryslgv.org",
     "405 Hollybrook Dr, Longview, TX 75605",
     "https://stmaryslgv.org",
     "~216 students, 8:1 ratio. Founded 1948. Classical education model 2025-26"),

    ("Longview Christian Academy (LCA)", "Gregg", "Longview", "Christian (Baptist)", "K-12",
     "", "Head of School",
     "(903) 759-0626", "lcaoffice1973@gmail.com",
     "2200 W Loop 281, Longview, TX 75604",
     "https://www.lcalongview.com",
     "~78 students, 6:1 ratio. Ministry of Emmanuel Baptist Church. Founded 1973"),

    ("Woodland Hills Christian Academy", "Gregg", "Longview", "Christian (Baptist/ACE)", "PK4-12",
     "Charles Hunt", "Head of School",
     "(430) 201-6537", "whca@whbchurch.com",
     "2105 E Loop 281, Longview, TX 75605",
     "https://whca.us",
     "Ministry of Woodland Hills Baptist Church. Founded 2021. Also: jennifer@whbchurch.com"),

    ("Thrasher School House", "Gregg", "Kilgore", "Private (Learning Differences)", "PK3-12",
     "", "",
     "", "",
     "2567 Danville Road, Kilgore, TX 75662",
     "https://thrasherschoolhouse.com",
     "One-room schoolhouse approach, focus on learning differences"),

    # ── RUSK COUNTY ──────────────────────────────────────────────
    ("RUSK COUNTY (Henderson)", None),

    ("Full Armor Christian Academy", "Rusk", "Henderson", "Christian (ACSI-accredited)", "K4-12",
     "Mika Jackson", "Head of School",
     "(903) 655-8489", "lisa.schweng@full-armor.org",
     "2324 FM 3135 E, Henderson, TX 75652",
     "https://full-armor.org",
     "~100 students, 13:1 ratio. Founded 1994. 4-day school week"),

    # ── CHEROKEE COUNTY ──────────────────────────────────────────
    ("CHEROKEE COUNTY (Jacksonville, Alto)", None),

    ("Westview Christian Academy", "Cherokee", "Jacksonville", "Christian (ACSI-affiliated)", "1-9",
     "", "",
     "(903) 363-6662", "",
     "20392 FM 747 N, Jacksonville, TX 75766",
     "",
     "Also supports homeschooling families. PO Box 2225"),

    ("Kids First Christian Academy", "Cherokee", "Jacksonville", "Christian (Baptist)", "PK-1",
     "", "",
     "", "",
     "210 Philip St, Jacksonville, TX 75766",
     "",
     "~99 students. Primarily early childhood/daycare. Facebook page only"),

    ("Alto Christian Academy", "Cherokee", "Alto", "Christian", "K-8",
     "", "",
     "(936) 858-2244", "",
     "385 Maggie Sessions St, Alto, TX 75925",
     "",
     "~20 students. Director: Jennifer Dearman (associated childcare)"),

    # ── HENDERSON COUNTY ─────────────────────────────────────────
    ("HENDERSON COUNTY (Athens)", None),

    ("Athens Christian Preparatory Academy", "Henderson", "Athens", "Christian (TEPSAC-accredited)", "K-12",
     "Teresa DeMay", "Principal",
     "(903) 386-0400", "jennifer@athensprep.org",
     "2621 US Highway 175 E, Athens, TX 75751",
     "https://www.athensprep.org",
     "~175 students, 9:1 ratio. Founded 2008. 49-acre campus"),

    ("Athens Christian Academy", "Henderson", "Athens", "Christian", "K-6",
     "Brent Williams", "Administrator",
     "(903) 675-5135", "",
     "105 S Carroll St, Athens, TX 75751",
     "",
     "~69-77 students"),

    # ── VAN ZANDT COUNTY ─────────────────────────────────────────
    ("VAN ZANDT COUNTY (Canton)", None),

    ("Canton Christian Academy", "Van Zandt", "Canton", "Christian (NAPS-accredited)", "K-12",
     "Julann Goldsmith", "Principal",
     "(903) 567-1133", "admin@cantonchristian.org",
     "1315 West Dallas Street, Canton, TX 75103",
     "https://www.cantonchristian.org",
     "Est. 2017. Mon-Thu 8am-3:30pm"),

    ("Holy Family Academy of Van Zandt County", "Van Zandt", "Canton", "Catholic (Diocese of Tyler)", "PK-4",
     "", "Principal",
     "(903) 567-4659", "holyfamilyvzc@gmail.com",
     "23951 Hwy 64, Canton, TX 75103",
     "https://www.holyfamilyvzc.org",
     "~15 students. Opened ~2023. TCCB ED accredited"),

    # ── WOOD & UPSHUR COUNTIES ───────────────────────────────────
    ("WOOD COUNTY & UPSHUR COUNTY — No active private schools found", None),
]

ws = wb.active
ws.title = "Private Schools"

col_widths = [40, 14, 16, 28, 10, 22, 18, 16, 32, 46, 38, 60]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

row_num = 2
for record in schools:
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=record[0])
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 28
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

output = "/home/user/Claude/Surrounding_Counties_Private_Schools.xlsx"
wb.save(output)
data_rows = sum(1 for r in schools if r[1] is not None)
print(f"Saved {data_rows} schools to {output}")
