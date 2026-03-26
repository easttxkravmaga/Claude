"""
Build Smith County TX — College Contacts XLSX
Tabs: UT Tyler | Tyler Junior College | Texas College | LeTourneau University
Output: /home/user/Claude/downloads/smith-county-college-contacts.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLACK = "FF000000"; WHITE = "FFFFFFFF"; RED = "FFCC0000"
GRAY  = "FF1A1A1A"; LGRAY = "FF2A2A2A"

def hdr_fill():   return PatternFill("solid", fgColor=RED)
def row_fill():   return PatternFill("solid", fgColor=BLACK)
def alt_fill():   return PatternFill("solid", fgColor=LGRAY)
def hdr_font():   return Font(name="Arial", bold=True, color=WHITE, size=10)
def body_font():  return Font(name="Arial", color=WHITE, size=9)
def title_font(): return Font(name="Arial", bold=True, color=RED, size=12)
def sub_font():   return Font(name="Arial", color="FF888888", size=8, italic=True)

def bdr():
    s = Side(border_style="thin", color="FF333333")
    return Border(left=s, right=s, top=s, bottom=s)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

HEADERS = ["#", "Name", "Title", "Department", "Phone", "Email", "Organization", "Notes"]

def build_sheet(wb, tab, title_text, subtitle, data):
    ws = wb.create_sheet(title=tab)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "CC0000"
    n = len(HEADERS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = title_font(); c.fill = PatternFill("solid", fgColor=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = sub_font(); c.fill = PatternFill("solid", fgColor=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font(); c.fill = hdr_fill()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr()
    ws.row_dimensions[3].height = 18

    for ri, row in enumerate(data, 4):
        fill = alt_fill() if ri % 2 == 0 else row_fill()
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = body_font(); c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=False)
            c.border = bdr()
        ws.row_dimensions[ri].height = 16
    return ws

# ═══════════════════════════════════════════════════
# UT TYLER
# ═══════════════════════════════════════════════════
UT = "University of Texas at Tyler"
ut_data = [
    (1,  "Dr. Julie V. Philley, MD", "President",                              "Office of the President",          "903-566-7325",  "mitziharris@uttyler.edu (admin)",    UT, "First woman president of UT Tyler"),
    (2,  "Mitzi Harris",             "President's Office Admin Contact",        "Office of the President",          "903-566-7325",  "mitziharris@uttyler.edu",            UT, ""),
    (3,  "",                         "Provost / Academic Affairs",              "Academic Affairs",                 "903-566-7103",  "academicaffairs@uttyler.edu",        UT, ""),
    (4,  "Dr. Troy G. White",        "University Registrar",                    "Office of the Registrar",          "903-566-7180",  "troywhite@uttyler.edu",              UT, ""),
    (5,  "Dennis Cliborn",           "Associate Registrar",                     "Office of the Registrar",          "903-566-7180",  "dcliborn@uttyler.edu",               UT, ""),
    (6,  "Andrew Barnson",           "Director, Enrollment Services",           "Office of the Registrar",          "903-566-7180",  "abarnson@uttyler.edu",               UT, ""),
    (7,  "Gena M. Smith",            "Graduation & Commencement Coordinator",   "Office of the Registrar",          "903-566-7180",  "genasmith@uttyler.edu",              UT, ""),
    (8,  "Katie Ramirez",            "Asst. VP, University Recruitment & Admissions","Undergraduate Admissions",   "903-566-7204",  "kramirez@uttyler.edu",               UT, ""),
    (9,  "Regan Ball",               "Exec. Director, Enrollment Operations",   "Undergraduate Admissions",         "903-566-6509",  "rball@uttyler.edu",                  UT, ""),
    (10, "Korbin Hamner",            "Asst. Director, Admissions-Welcome Ctr",  "Undergraduate Admissions",         "903-566-7080",  "Khamner@uttyler.edu",                UT, ""),
    (11, "Tanner Wold",              "Asst. Director, Undergraduate Recruitment","Undergraduate Admissions",        "903-566-7034",  "Twold@uttyler.edu",                  UT, ""),
    (12, "Rudy Alarcon",             "Admissions Officer, Greater Tyler",        "Undergraduate Admissions",         "903-566-5863",  "ralarcon@uttyler.edu",               UT, ""),
    (13, "Taylor Fray",              "Admissions Officer, Transfer",             "Undergraduate Admissions",         "903-566-6285",  "tfray@uttyler.edu",                  UT, ""),
    (14, "Jasmine Rodriguez",        "Houston Regional Manager",                 "Undergraduate Admissions",         "903-730-3839",  "jasminerodriguez@uttyler.edu",       UT, ""),
    (15, "Kendal Hill",              "Dallas Regional Manager",                  "Undergraduate Admissions",         "903-566-4451",  "khill@uttyler.edu",                  UT, ""),
    (16, "",                         "Graduate Admissions",                      "Graduate Admissions",              "903-566-7457",  "gradadmissions@uttyler.edu",         UT, ""),
    (17, "",                         "Human Resources",                          "Human Resources",                  "903-566-7234",  "humanresources@uttyler.edu",         UT, ""),
    (18, "",                         "Faculty Senate",                           "Faculty Senate",                   "903-566-7177",  "facultysenate@uttyler.edu",          UT, ""),
    (19, "",                         "Student Government Association",           "SGA",                              "903-565-5557",  "sga@uttyler.edu",                    UT, "UC 2160"),
    (20, "A. Dinkins",               "Scheduling & Conference Services",         "Scheduling & Conference Services", "903-565-5723",  "adinkins@uttyler.edu",               UT, ""),
    (21, "D. Bailey",                "Associate Prof / Dir, Center for Excellence","Social Sciences",               "903-566-7432",  "dbailey@uttyler.edu",                UT, "Criminal Justice"),
    (22, "",                         "Pharmacy / Pharmaceutical Sciences",       "Pharmacy",                         "903-566-6287",  "cabdullah@uttyler.edu",              UT, ""),
    (23, "H. Bloodsworth",           "Testing Center / Student Accessibility",   "Testing Center",                   "903-565-5693",  "hbloodsworth@uttyler.edu",           UT, ""),
    (24, "E. Battle",                "Strategic Communications & Media Relations","Marketing & Communications",      "903-565-5604",  "ebattle@uttyler.edu",                UT, ""),
]

# ═══════════════════════════════════════════════════
# TJC
# ═══════════════════════════════════════════════════
TJC = "Tyler Junior College"
tjc_data = [
    (1,  "Dr. Juan E. Mejia",    "President & CEO",                          "Office of the President",              "903-510-2200",  "",                        TJC, "12th year with TJC"),
    (2,  "Ellen Matthews",       "Secretary to the Board of Trustees",       "Board of Trustees",                    "903-510-2379",  "emat@tjc.edu",            TJC, ""),
    (3,  "Dr. Deana Sheppard",   "Provost & VP, Academic and Student Affairs","Provost's Office",                    "903-510-3017",  "deana.sheppard@tjc.edu",  TJC, ""),
    (4,  "Dr. Lee Grimes",       "Associate Vice Provost, Academic Affairs",  "Provost's Office",                    "903-510-2281",  "lee.grimes@tjc.edu",      TJC, ""),
    (5,  "Dr. Tim Drain",        "Associate Vice Provost, Student Affairs",   "Provost's Office",                    "903-510-2320",  "tdra@tjc.edu",            TJC, ""),
    (6,  "Kayla Cantey",         "Exec. Admin Asst., Provost's Office",       "Provost's Office",                    "903-510-2261",  "Kayla.Cantey@tjc.edu",    TJC, ""),
    (7,  "Dr. Jim Richey",       "Dean, Humanities, Communications & Fine Arts","School of Humanities",              "903-510-2468",  "jric@tjc.edu",            TJC, ""),
    (8,  "Dr. Elizabeth Olivier","Dean, Nursing and Health Sciences",          "School of Nursing & Health Sciences", "903-510-2362",  "eoli@tjc.edu",            TJC, ""),
    (9,  "Dr. Cliff Boucher",    "Dean, Engineering, Mathematics & Sciences",  "School of Engineering & Sciences",    "903-510-2546",  "cliff.boucher@tjc.edu",   TJC, ""),
    (10, "Dr. Richard Nichols",  "Dean, Professional and Technical Programs",  "School of Professional & Technical",  "903-533-5564",  "Richard.Nichols@tjc.edu", TJC, ""),
    (11, "Brent Wallace",        "Dean, TJC West & Skilled Trades",            "TJC West / Skilled Trades",           "903-510-2999",  "brent.wallace@tjc.edu",   TJC, ""),
    (12, "",                     "Admissions Office",                          "Admissions",                          "903-510-2523",  "Admissions@TJC.edu",      TJC, "Text: 903-371-0120"),
    (13, "",                     "Apache Enrollment Center (Fin. Aid)",        "Financial Aid",                       "903-510-2385",  "aec@tjc.edu",             TJC, "FAFSA School Code: 003648"),
    (14, "",                     "Campus Police (Non-Emergency)",              "Campus Safety",                       "903-510-2800",  "",                        TJC, "Emergency: 911"),
    (15, "",                     "Athletics",                                  "Athletics",                           "903-510-2458",  "",                        TJC, ""),
    (16, "Kelsi Weeks",          "Athletics / Transfer Coordinator",           "Athletics",                           "",              "kelsi.weeks@tjc.edu",     TJC, ""),
    (17, "Judie Bower",          "Director, Program Development",              "Program Development",                 "",              "",                        TJC, ""),
    (18, "",                     "TJC North Campus",                           "TJC North",                           "903-510-3100",  "",                        TJC, "75 Miranda Lambert Way, Suite 16, Lindale TX 75771"),
    (19, "",                     "TJC West Campus",                            "TJC West",                            "903-510-2992",  "",                        TJC, "1530 S SW Loop 323, Tyler TX 75701"),
]

# ═══════════════════════════════════════════════════
# TEXAS COLLEGE
# ═══════════════════════════════════════════════════
TC = "Texas College"
tc_data = [
    (1,  "Dr. Jan Duncan",        "VP for Academic Affairs",               "Academic Affairs",      "(903) 593-8311 x2704", "",                             TC, "Martin Hall, 1st Floor"),
    (2,  "Dr. C. Marshall-Biggins","VP for Student Affairs",               "Student Affairs",       "(903) 593-8311 x2710", "studentaffairs1@texascollege.edu", TC, ""),
    (3,  "John Roberts",          "Dean of Enrollment Management & Registrar","Admissions / Registrar","(903) 593-8311 x2297","admissions1@texascollege.edu", TC, "Also: registrar2@texascollege.edu ext 2233"),
    (4,  "A. Copeland",           "Asst. Director of Financial Aid",       "Financial Aid",         "(903) 593-8311 x2299", "financialaid@texascollege.edu", TC, ""),
    (5,  "",                      "Business and Finance / Student Accounts","Business & Finance",    "(903) 593-8311 x2200", "businessoffice@texascollege.edu",TC, ""),
    (6,  "Elissia Burwell",       "Athletic Director",                     "Athletics",             "(903) 593-8311 x4043", "eburwell@texascollege.edu",    TC, ""),
    (7,  "Yaw Labang",            "Director of Information Technology",    "Information Technology","(903) 593-8311 x2315", "",                             TC, ""),
    (8,  "",                      "Human Resources",                       "Human Resources",       "(903) 593-8311 x2201", "",                             TC, "Martin Hall, 2nd Floor; Fax: 903-363-1828"),
    (9,  "",                      "Main Switchboard",                      "General",               "(903) 593-8311",       "",                             TC, "2404 N. Grand Ave, Tyler TX 75702"),
]

# ═══════════════════════════════════════════════════
# LETOURNEAU UNIVERSITY
# ═══════════════════════════════════════════════════
LU = "LeTourneau University"
lu_data = [
    (1,  "",  "Admissions",                   "Admissions",               "800-759-8811",  "Admissions@LETU.edu",  LU, "Main campus: Longview TX; Tyler campus: 501 Shelley Dr, Tyler TX"),
    (2,  "",  "Main Switchboard",             "General",                  "903-233-3000",  "",                     LU, "2100 S. Mobberly Ave, Longview TX 75602"),
    (3,  "",  "Information Technology",       "IT Department",            "903-233-3000",  "",                     LU, "letu.edu/offices/administration-finance/it/contact-us.html"),
    (4,  "",  "Student Life",                 "Student Life",             "903-233-3000",  "",                     LU, "letu.edu/student-life/meet-the-staff.html"),
    (5,  "",  "Tyler Campus",                 "Tyler Location",           "",              "",                     LU, "501 Shelley Dr, Tyler TX — call main 903-233-3000 for hours"),
]

# ═══════════════════════════════════════════════════
# BUILD
# ═══════════════════════════════════════════════════
wb = openpyxl.Workbook()

ws1 = build_sheet(wb, "UT Tyler",
    "UNIVERSITY OF TEXAS AT TYLER — Contacts",
    "3900 University Blvd, Tyler TX 75799  |  800-UT-TYLER  |  Scraped 2026-03-26",
    ut_data)
set_widths(ws1, [4, 26, 34, 30, 16, 32, 28, 36])

ws2 = build_sheet(wb, "Tyler Junior College",
    "TYLER JUNIOR COLLEGE — Contacts",
    "1327 S. Baxter Ave, Tyler TX 75701  |  903-510-2200  |  Scraped 2026-03-26",
    tjc_data)
set_widths(ws2, [4, 22, 34, 30, 16, 26, 22, 36])

ws3 = build_sheet(wb, "Texas College",
    "TEXAS COLLEGE — Contacts",
    "2404 N. Grand Ave, Tyler TX 75702  |  (903) 593-8311  |  Scraped 2026-03-26",
    tc_data)
set_widths(ws3, [4, 22, 30, 26, 22, 32, 16, 36])

ws4 = build_sheet(wb, "LeTourneau University",
    "LETOURNEAU UNIVERSITY — Tyler Campus Contacts",
    "501 Shelley Dr, Tyler TX  |  Main: 903-233-3000  |  Scraped 2026-03-26",
    lu_data)
set_widths(ws4, [4, 22, 24, 22, 16, 24, 22, 42])

if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

out = "/home/user/Claude/downloads/smith-county-college-contacts.xlsx"
wb.save(out)
print(f"Saved: {out}")
