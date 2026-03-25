"""
Build East Texas HR Manager contact spreadsheet.
If output/smith_county_hr_managers.csv exists (from running the scraper),
converts that. Otherwise generates from the embedded seed company list.
"""

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)
CSV_INPUT  = OUTPUT_DIR / "smith_county_hr_managers.csv"
XLSX_OUT   = OUTPUT_DIR / "smith_county_hr_managers.xlsx"

# ── Colors ─────────────────────────────────────────────────────────────────────
BLACK  = "FF000000"
WHITE  = "FFFFFFFF"
RED    = "FFCC0000"
DARK   = "FF111111"
GRAY   = "FF1A1A1A"
LGRAY  = "FF2A2A2A"
ALTROW = "FF222222"

thin   = Side(style="thin", color="FF444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

SEED_COMPANIES = [
    # Smith County — Tyler
    ("UT Health Tyler",                     "uthct.edu",               "Tyler",        "Smith"),
    ("CHRISTUS Mother Frances Hospital",    "christushealth.org",      "Tyler",        "Smith"),
    ("Tyler Technologies",                  "tylertech.com",           "Tyler",        "Smith"),
    ("ALDI Distribution Center",            "aldi.us",                 "Tyler",        "Smith"),
    ("Brookshire Grocery Company",          "brookshires.com",         "Tyler",        "Smith"),
    ("Trane Technologies",                  "tranetechnologies.com",   "Tyler",        "Smith"),
    ("University of Texas at Tyler",        "uttyler.edu",             "Tyler",        "Smith"),
    ("Tyler ISD",                           "tylerisd.org",            "Tyler",        "Smith"),
    ("Tyler Junior College",                "tjc.edu",                 "Tyler",        "Smith"),
    ("East Texas Medical Center",           "etmc.org",                "Tyler",        "Smith"),
    ("Suddenlink / Altice",                 "alticeusa.com",           "Tyler",        "Smith"),
    ("Smith County",                        "smith-county.com",        "Tyler",        "Smith"),
    ("City of Tyler",                       "cityoftyler.org",         "Tyler",        "Smith"),
    ("Target Distribution Center",          "target.com",              "Tyler",        "Smith"),
    ("Walmart Distribution Center",         "walmart.com",             "Tyler",        "Smith"),
    ("Pilgrim's Pride",                     "pilgrims.com",            "Tyler",        "Smith"),
    ("HEB Tyler",                           "heb.com",                 "Tyler",        "Smith"),
    ("Home Depot Tyler",                    "homedepot.com",           "Tyler",        "Smith"),
    ("Regions Bank Tyler",                  "regions.com",             "Tyler",        "Smith"),
    ("JPMorgan Chase Tyler",                "jpmorganchase.com",       "Tyler",        "Smith"),
    ("Wells Fargo Tyler",                   "wellsfargo.com",          "Tyler",        "Smith"),
    ("Southside Bank",                      "southsidebank.com",       "Tyler",        "Smith"),
    ("Peoples Bank Tyler",                  "mypeoplesbank.com",       "Tyler",        "Smith"),
    ("First Bank and Trust East Texas",     "fbtet.com",               "Tyler",        "Smith"),
    ("BancorpSouth Tyler",                  "bancorpsouth.com",        "Tyler",        "Smith"),
    ("Security Bank Tyler",                 "securitybank.com",        "Tyler",        "Smith"),
    ("Broadway National Bank",              "broadwaynational.com",    "Tyler",        "Smith"),
    ("KLTV / East Texas Media Group",       "kltv.com",                "Tyler",        "Smith"),
    ("Texas Health & Human Services",       "hhs.texas.gov",           "Tyler",        "Smith"),
    ("Texas Workforce Commission Tyler",    "twc.texas.gov",           "Tyler",        "Smith"),
    ("East Texas Communities Foundation",   "etcf.org",                "Tyler",        "Smith"),
    ("Caliber Collision Tyler",             "calibercollision.com",    "Tyler",        "Smith"),
    ("Southern Tire Mart",                  "southerntiremart.com",    "Tyler",        "Smith"),
    ("Harvey Industries",                   "harveyindustries.com",    "Tyler",        "Smith"),
    ("HomeTeam Pest Defense",               "hometeam.com",            "Tyler",        "Smith"),
    ("Rose Capital Hotels",                 "rosecapitalhotels.com",   "Tyler",        "Smith"),
    ("Whitehouse ISD",                      "whitehouseisd.org",       "Whitehouse",   "Smith"),
    ("Lindale ISD",                         "lindaleeagles.org",       "Lindale",      "Smith"),
    # Gregg County — Longview
    ("Longview Regional Medical Center",    "longviewregional.com",    "Longview",     "Gregg"),
    ("Good Shepherd Medical Center",        "gsmc.org",                "Longview",     "Gregg"),
    ("Longview ISD",                        "lisd.org",                "Longview",     "Gregg"),
    ("Gregg County",                        "co.gregg.tx.us",          "Longview",     "Gregg"),
    ("City of Longview",                    "longviewtexas.gov",       "Longview",     "Gregg"),
    ("Eastman Chemical Longview",           "eastman.com",             "Longview",     "Gregg"),
    ("Halliburton Longview",                "halliburton.com",         "Longview",     "Gregg"),
    ("LeTourneau University",               "letu.edu",                "Longview",     "Gregg"),
    ("Kilgore College",                     "kilgore.edu",             "Kilgore",      "Gregg"),
    ("East Texas Council of Governments",   "etcog.org",               "Kilgore",      "Gregg"),
    ("Longview News-Journal",               "news-journal.com",        "Longview",     "Gregg"),
    # Henderson County — Athens
    ("Athens ISD",                          "athenstigers.net",        "Athens",       "Henderson"),
    ("Henderson County",                    "hendersoncounty.net",     "Athens",       "Henderson"),
    ("CHRISTUS Mother Frances Athens",      "christushealth.org",      "Athens",       "Henderson"),
    # Van Zandt County — Canton
    ("Canton ISD",                          "cantonisd.net",           "Canton",       "Van Zandt"),
    ("Van Zandt County",                    "vanzandtcounty.org",      "Canton",       "Van Zandt"),
    # Cherokee County — Jacksonville
    ("Jacksonville ISD",                    "jacksonvilleisd.net",     "Jacksonville", "Cherokee"),
    ("Jacksonville Medical Center",         "jacksonvillemc.com",      "Jacksonville", "Cherokee"),
    ("Rusk State Hospital",                 "hhs.texas.gov",           "Rusk",         "Cherokee"),
    # Upshur County — Gilmer
    ("Gilmer ISD",                          "gilmerisd.org",           "Gilmer",       "Upshur"),
    ("Upshur County",                       "upshurcounty.org",        "Gilmer",       "Upshur"),
    # Wood County — Quitman
    ("Quitman ISD",                         "quitmanisd.net",          "Quitman",      "Wood"),
    ("Wood County",                         "mywoodcounty.com",        "Quitman",      "Wood"),
]

COLUMNS = [
    ("Company",         22),
    ("City",            12),
    ("County",          12),
    ("Contact Name",    20),
    ("Title",           26),
    ("Email",           30),
    ("Confidence %",    14),
    ("Domain",          24),
    ("LinkedIn",        30),
]


def apply(cell, **kw):
    for attr, val in kw.items():
        setattr(cell, attr, val)


def build():
    # ── Load data ──────────────────────────────────────────────────────────────
    rows = []
    fieldnames = ["company", "city", "county", "name", "title",
                  "email", "confidence", "domain", "linkedin"]

    if CSV_INPUT.exists():
        print(f"[Data] Loading from {CSV_INPUT}")
        with open(CSV_INPUT, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for r in reader:
                rows.append([
                    r.get("company", ""),
                    r.get("city", ""),
                    r.get("county", ""),
                    r.get("name", ""),
                    r.get("title", ""),
                    r.get("email", ""),
                    r.get("confidence", ""),
                    r.get("domain", ""),
                    r.get("linkedin", ""),
                ])
        print(f"[Data] {len(rows)} contacts loaded")

    if not rows:
        print(f"[Data] CSV not found — building from seed company list ({len(SEED_COMPANIES)} companies)")
        seen_domains = set()
        for company, domain, city, county in SEED_COMPANIES:
            if domain in seen_domains:
                continue
            seen_domains.add(domain)
            rows.append([company, city, county, "", "", "", "", domain, ""])

    # ── Build workbook ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HR Contacts"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "East Texas — Corporate HR Contacts  |  Smith County & Surrounding Areas"
    title_cell.font      = Font(bold=True, color=WHITE, name="Arial", size=13)
    title_cell.fill      = PatternFill("solid", fgColor=RED)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-header row
    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = (
        f"Sources: Tyler Chamber · Longview Chamber · Indeed HR Postings · Hunter.io  "
        f"| {len(rows)} {'contacts' if CSV_INPUT.exists() else 'target companies'}"
    )
    sub.font      = Font(color="FFAAAAAA", name="Arial", size=9, italic=True)
    sub.fill      = PatternFill("solid", fgColor=DARK)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Column headers
    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font      = Font(bold=True, color=WHITE, name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor="FF880000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22

    # Data rows
    for row_idx, row_data in enumerate(rows, start=4):
        fill_color = ALTROW if row_idx % 2 == 0 else DARK
        row_fill   = PatternFill("solid", fgColor=fill_color)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Column-specific styling
            if col_idx == 1:   # Company — bold white
                cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
            elif col_idx == 6: # Email — red accent
                cell.font = Font(color="FFFF4444", name="Arial", size=10)
            elif col_idx == 7: # Confidence
                cell.font      = Font(color="FFAAAAAA", name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = Font(color="FFCCCCCC", name="Arial", size=10)

        ws.row_dimensions[row_idx].height = 16

    # Freeze panes below header rows
    ws.freeze_panes = "A4"

    # Auto-filter on header row
    ws.auto_filter.ref = f"A3:I{len(rows) + 3}"

    # Footer row
    footer_row = len(rows) + 4
    ws.merge_cells(f"A{footer_row}:I{footer_row}")
    footer = ws.cell(row=footer_row, column=1,
                     value="East Texas Krav Maga  |  Corporate Training Outreach  |  Smith County & Surrounding Areas")
    footer.font      = Font(color="FF666666", name="Arial", size=8, italic=True)
    footer.fill      = PatternFill("solid", fgColor=BLACK)
    footer.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[footer_row].height = 14

    wb.save(XLSX_OUT)
    print(f"[Done] Saved → {XLSX_OUT}  ({len(rows)} rows)")


if __name__ == "__main__":
    build()
