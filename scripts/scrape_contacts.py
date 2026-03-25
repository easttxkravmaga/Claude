#!/usr/bin/env python3
"""
ETKM Contact Scraper
Uses Playwright with a real Chromium browser to bypass 403 blocks.
Usage:
    python3 scripts/scrape_contacts.py <url> [url2] [url3] ...
    python3 scripts/scrape_contacts.py --file urls.txt
Output: output/scrape_contacts.xlsx (appends each run)
"""

import sys
import re
import argparse
from pathlib import Path
from datetime import datetime
from playwright.sync_api import sync_playwright

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

CHROMIUM_PATH = "/root/.cache/ms-playwright/chromium-1194/chrome-linux/chrome"
OUTPUT_DIR = Path(__file__).parent.parent / "output"
OUTPUT_XLSX = OUTPUT_DIR / "scrape_contacts.xlsx"

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
PHONE_RE = re.compile(r"(\(?\d{3}\)?[\s.\-]?\d{3}[\s.\-]?\d{4})")
FIELDS = ["Source URL", "Name", "Title", "Email", "Phone", "Address", "Department", "Scraped At"]


def fetch_page_text(url: str, timeout: int = 20000) -> tuple[str, str]:
    """Returns (page_text, page_title) using headless Chromium."""
    with sync_playwright() as p:
        browser = p.chromium.launch(
            executable_path=CHROMIUM_PATH,
            args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-blink-features=AutomationControlled"],
            headless=True,
        )
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800},
        )
        page = context.new_page()
        page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        try:
            page.goto(url, wait_until="networkidle", timeout=timeout)
        except Exception:
            page.goto(url, wait_until="domcontentloaded", timeout=timeout)
        title = page.title()
        text = page.inner_text("body")
        browser.close()
        return text, title


def extract_contacts(text: str, url: str) -> list[dict]:
    """Heuristic extraction of contact blocks from raw page text."""
    emails = EMAIL_RE.findall(text)
    phones = PHONE_RE.findall(text)
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    contacts = []
    used_emails = set()

    for email in emails:
        if email in used_emails:
            continue
        used_emails.add(email)

        name, title, address, department = "", "", "", ""
        for i, line in enumerate(lines):
            if email in line:
                window = lines[max(0, i - 5):i + 5]
                for wline in window:
                    if email in wline:
                        continue
                    if PHONE_RE.fullmatch(wline.strip()):
                        continue
                    if wline.startswith("http") or wline.startswith("www"):
                        continue
                    if not name and len(wline) < 60 and not any(
                        w in wline.lower() for w in ["©", "privacy", "cookie", "menu", "navigation"]
                    ):
                        name = wline
                    elif not title and len(wline) < 80:
                        title = wline
                break

        phone = phones[0] if len(phones) == 1 else ""
        contacts.append({
            "Source URL": url,
            "Name": name,
            "Title": title,
            "Email": email,
            "Phone": phone,
            "Address": address,
            "Department": department,
            "Scraped At": datetime.now().isoformat(timespec="seconds"),
        })

    if not contacts and phones:
        contacts.append({
            "Source URL": url,
            "Name": "",
            "Title": "",
            "Email": "",
            "Phone": phones[0],
            "Address": "",
            "Department": "",
            "Scraped At": datetime.now().isoformat(timespec="seconds"),
        })

    return contacts


def save_contacts(contacts: list[dict]):
    OUTPUT_DIR.mkdir(exist_ok=True)

    # Load existing workbook or create new one
    if OUTPUT_XLSX.exists():
        wb = openpyxl.load_workbook(OUTPUT_XLSX)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"

        # Header row styling
        header_fill = PatternFill("solid", fgColor="CC0000")
        header_font = Font(bold=True, color="FFFFFF")
        for col, field in enumerate(FIELDS, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        col_widths = [40, 25, 30, 35, 18, 35, 25, 22]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Append rows
    for contact in contacts:
        ws.append([contact[f] for f in FIELDS])

    wb.save(OUTPUT_XLSX)


def scrape(url: str):
    print(f"\n→ Scraping: {url}")
    try:
        text, title = fetch_page_text(url)
        print(f"  Page: {title}")
        contacts = extract_contacts(text, url)
        if contacts:
            save_contacts(contacts)
            print(f"  Found {len(contacts)} contact(s):")
            for c in contacts:
                print(f"    {c['Name'] or '(unnamed)':30s}  {c['Title'] or '':35s}  {c['Email']}  {c['Phone']}")
        else:
            print("  No structured contacts found. Raw text sample:")
            print("  " + text[:500].replace("\n", " "))
    except Exception as e:
        print(f"  ERROR: {e}")


def main():
    parser = argparse.ArgumentParser(description="ETKM Contact Scraper")
    parser.add_argument("urls", nargs="*", help="One or more URLs to scrape")
    parser.add_argument("--file", help="Text file with one URL per line")
    args = parser.parse_args()

    urls = list(args.urls)
    if args.file:
        urls += [l.strip() for l in Path(args.file).read_text().splitlines() if l.strip()]

    if not urls:
        parser.print_help()
        sys.exit(1)

    for url in urls:
        scrape(url)

    print(f"\n✓ Contacts saved to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
