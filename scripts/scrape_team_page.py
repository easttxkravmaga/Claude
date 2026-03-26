"""
scrape_team_page.py
-------------------
Scrapes a staff/team page and pairs names, titles, emails, and phones
into a clean XLSX output.

Usage:
    python3 scrape_team_page.py <URL> [output_filename]

Examples:
    python3 scrape_team_page.py https://www.gabc.org/our-team/
    python3 scrape_team_page.py https://www.gabc.org/our-team/ gabc-team.xlsx

Requirements:
    pip install requests beautifulsoup4 openpyxl
"""

import sys
import re
import requests
from bs4 import BeautifulSoup, Tag
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse
from pathlib import Path

# ── helpers ────────────────────────────────────────────────────────────────

PHONE_RE = re.compile(
    r'(\(?\d{3}\)?[\s.\-]?\d{3}[\s.\-]?\d{4})'
)

def clean(text):
    return re.sub(r'\s+', ' ', text or '').strip()

def extract_phone_from_text(text):
    m = PHONE_RE.search(text or '')
    return m.group(1) if m else ''

def walk_up_for_container(tag, max_levels=6):
    """Walk up the DOM to find the nearest meaningful block container."""
    BLOCK_TAGS = {'div', 'article', 'section', 'li', 'tr', 'td', 'figure'}
    node = tag.parent
    for _ in range(max_levels):
        if node is None:
            break
        if isinstance(node, Tag) and node.name in BLOCK_TAGS:
            return node
        node = node.parent
    return tag.parent

def extract_name_title(container):
    """
    Try to extract a name and title from a container block.
    Strategy: look for heading tags (h1-h5), then <strong>, then first bold line.
    Title is the next non-empty text block after the name.
    """
    name = ''
    title = ''

    # Headings first
    for level in ['h2', 'h3', 'h4', 'h5', 'h1', 'h6']:
        h = container.find(level)
        if h:
            name = clean(h.get_text())
            break

    # Fall back to <strong> or <b>
    if not name:
        s = container.find(['strong', 'b'])
        if s:
            name = clean(s.get_text())

    # Title: look for common class hints first
    for cls in ['title', 'position', 'role', 'job', 'subtitle', 'staff-title',
                'team-title', 'member-title', 'card-title', 'designation']:
        el = container.find(class_=re.compile(cls, re.I))
        if el:
            candidate = clean(el.get_text())
            if candidate and candidate != name:
                title = candidate
                break

    # Fall back: first <p> or <span> that isn't the name and isn't an email/phone
    if not title:
        for tag in container.find_all(['p', 'span', 'div'], recursive=True):
            t = clean(tag.get_text())
            if (t and t != name
                    and '@' not in t
                    and not PHONE_RE.search(t)
                    and len(t) < 120
                    and len(t) > 2):
                title = t
                break

    return name, title

def get_container_phone(container):
    """Extract a phone number from a container — tel: link first, then text."""
    tel = container.find('a', href=re.compile(r'^tel:', re.I))
    if tel:
        return clean(tel['href'].replace('tel:', '').replace('tel=', ''))
    return extract_phone_from_text(container.get_text())

# ── main scraper ───────────────────────────────────────────────────────────

def scrape(url):
    headers = {
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/120.0.0.0 Safari/537.36'
        )
    }
    print(f"Fetching: {url}")
    r = requests.get(url, headers=headers, timeout=20)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, 'html.parser')

    records = []
    seen_emails = set()

    # Strategy 1: anchor all records on mailto links
    mailto_links = [a for a in soup.find_all('a', href=True) if 'mailto:' in a['href']]

    for a in mailto_links:
        email = clean(a['href'].replace('mailto:', ''))
        if not email or email in seen_emails:
            continue
        seen_emails.add(email)

        container = walk_up_for_container(a, max_levels=8)
        name, title = extract_name_title(container)
        phone = get_container_phone(container)

        records.append({
            'name':  name,
            'title': title,
            'email': email,
            'phone': phone,
        })

    # Strategy 2: if no mailto links found, try tel: links and text extraction
    if not records:
        print("No mailto links found — falling back to tel: link scan...")
        tel_links = [a for a in soup.find_all('a', href=True) if 'tel:' in a['href']]
        for a in tel_links:
            phone = clean(a['href'].replace('tel:', ''))
            container = walk_up_for_container(a, max_levels=8)
            name, title = extract_name_title(container)
            email_tag = container.find('a', href=re.compile(r'mailto:', re.I))
            email = ''
            if email_tag:
                email = clean(email_tag['href'].replace('mailto:', ''))
            records.append({'name': name, 'title': title, 'email': email, 'phone': phone})

    print(f"Found {len(records)} contacts.")
    return records

# ── XLSX builder ───────────────────────────────────────────────────────────

BLACK = "FF000000"; WHITE = "FFFFFFFF"; RED = "FFCC0000"; LGRAY = "FF2A2A2A"

def build_xlsx(records, url, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "CC0000"

    def bdr():
        s = Side(border_style="thin", color="FF333333")
        return Border(left=s, right=s, top=s, bottom=s)

    headers = ["#", "Name", "Title", "Email", "Phone", "Source URL"]
    col_widths = [4, 28, 32, 34, 18, 48]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=f"STAFF CONTACTS — {urlparse(url).netloc}")
    c.font = Font(name="Arial", bold=True, color=RED, size=12)
    c.fill = PatternFill("solid", fgColor=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Sub-header
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    c = ws.cell(row=2, column=1, value=f"Source: {url}  |  Scraped 2026-03-26")
    c.font = Font(name="Arial", color="FF888888", size=8, italic=True)
    c.fill = PatternFill("solid", fgColor=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    # Header row
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr()
    ws.row_dimensions[3].height = 18

    # Data rows
    for ri, rec in enumerate(records, 4):
        fill = PatternFill("solid", fgColor=LGRAY) if ri % 2 == 0 else PatternFill("solid", fgColor=BLACK)
        row = [ri - 3, rec['name'], rec['title'], rec['email'], rec['phone'], url]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", color=WHITE, size=9)
            c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=False)
            c.border = bdr()
        ws.row_dimensions[ri].height = 16

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    print(f"Saved: {out_path}")

# ── entry point ────────────────────────────────────────────────────────────

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 scrape_team_page.py <URL> [output.xlsx]")
        sys.exit(1)

    url = sys.argv[1]

    # Default output filename derived from domain + path
    if len(sys.argv) >= 3:
        out_path = sys.argv[2]
    else:
        domain = urlparse(url).netloc.replace('www.', '').replace('.', '-')
        slug = urlparse(url).path.strip('/').replace('/', '-') or 'contacts'
        out_path = f"{domain}-{slug}.xlsx"

    records = scrape(url)

    if not records:
        print("No contacts found. The page may require JavaScript rendering.")
        print("Try opening the page in a browser, saving as HTML, then running:")
        print("  python3 scrape_team_page.py file:///path/to/saved.html")
        sys.exit(1)

    build_xlsx(records, url, out_path)

    # Print summary to console
    print(f"\n{'#':<4} {'Name':<28} {'Title':<30} {'Email':<34} {'Phone'}")
    print('-' * 110)
    for i, r in enumerate(records, 1):
        print(f"{i:<4} {r['name']:<28} {r['title']:<30} {r['email']:<34} {r['phone']}")
