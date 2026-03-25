"""
Convert smith_county_realtors.csv to a formatted Excel spreadsheet.
"""

import csv
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

INPUT  = "output/smith_county_realtors.csv"
OUTPUT = "output/smith_county_realtors.xlsx"

# ── Colors ────────────────────────────────────────────────────────────────────
BLACK  = "FF000000"
WHITE  = "FFFFFFFF"
RED    = "FFCC0000"
DARK   = "FF111111"
GRAY   = "FF222222"
LGRAY  = "FF333333"
ALTROW = "FF1A1A1A"

thin = Side(style="thin", color="FF444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def header_style():
    return {
        "font": Font(bold=True, color=WHITE, name="Arial", size=11),
        "fill": PatternFill("solid", fgColor=RED),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": border,
    }


def apply(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


def build():
    # Read CSV
    with open(INPUT, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fields = reader.fieldnames

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Smith County Realtors"

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(fields))}1")
    title_cell = ws["A1"]
    title_cell.value = "Smith County, TX — Realtor Contacts"
    title_cell.font = Font(bold=True, color=WHITE, name="Arial", size=14)
    title_cell.fill = PatternFill("solid", fgColor=BLACK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Subtitle row ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(len(fields))}2")
    sub_cell = ws["A2"]
    sub_cell.value = f"East Texas Krav Maga  |  {len(rows)} contacts  |  Sources: GTAR, HAR.com, cbapex.com, remax.com, ebby.com, eXp Realty, RealEdge, Gregory Realtors"
    sub_cell.font = Font(italic=True, color="FFAAAAAA", name="Arial", size=9)
    sub_cell.fill = PatternFill("solid", fgColor=DARK)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Header row ────────────────────────────────────────────────────────────
    COL_LABELS = {
        "name":        "Name",
        "phone":       "Phone",
        "email":       "Email",
        "office":      "Office",
        "brokerage":   "Brokerage",
        "address":     "Address",
        "city":        "City",
        "state":       "State",
        "zip":         "ZIP",
        "profile_url": "Profile URL",
        "source":      "Source",
        "notes":       "Notes",
    }
    hs = header_style()
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=3, column=col_idx, value=COL_LABELS.get(field, field.title()))
        apply(cell, **hs)
    ws.row_dimensions[3].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, 4):
        fill_color = ALTROW if row_idx % 2 == 0 else DARK
        for col_idx, field in enumerate(fields, 1):
            val = row.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(color=WHITE, name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border

            # Hyperlink for URL column
            if field == "profile_url" and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="FF4488FF", name="Arial", size=10, underline="single")

            # Red accent for name column (bold)
            if field == "name":
                cell.font = Font(color=WHITE, name="Arial", size=10, bold=True)

        ws.row_dimensions[row_idx].height = 18

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        "name":        28,
        "phone":       16,
        "email":       30,
        "office":      30,
        "brokerage":   22,
        "address":     28,
        "city":        12,
        "state":        7,
        "zip":          8,
        "profile_url": 40,
        "source":      22,
        "notes":       40,
    }
    for col_idx, field in enumerate(fields, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 18)

    # ── Freeze panes below header ─────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:{get_column_letter(len(fields))}3"

    # Set tab color
    ws.sheet_properties.tabColor = "CC0000"

    # Dark background for whole sheet
    ws.sheet_view.showGridLines = False

    wb.save(OUTPUT)
    print(f"Saved → {OUTPUT}  ({len(rows)} contacts)")


if __name__ == "__main__":
    build()
