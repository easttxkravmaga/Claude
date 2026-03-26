"""
Build Smith County TX Brokers XLSX
Generates output/smith-county-brokers.xlsx with styled tabs for
mortgage brokers and investment/wealth management firms.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Color palette (ETKM brand) ─────────────────────────────────────────────
BLACK  = "FF000000"
WHITE  = "FFFFFFFF"
RED    = "FFCC0000"
GRAY   = "FF1A1A1A"
LGRAY  = "FF2A2A2A"

def hdr_fill():  return PatternFill("solid", fgColor=RED)
def sub_fill():  return PatternFill("solid", fgColor=GRAY)
def row_fill():  return PatternFill("solid", fgColor=BLACK)
def alt_fill():  return PatternFill("solid", fgColor=LGRAY)

def hdr_font():  return Font(name="Arial", bold=True, color=WHITE, size=10)
def body_font(): return Font(name="Arial", color=WHITE, size=9)
def title_font():return Font(name="Arial", bold=True, color=RED, size=12)

def thin_border():
    s = Side(border_style="thin", color="FF333333")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Data ───────────────────────────────────────────────────────────────────

MORTGAGE_HEADERS = [
    "#", "Company", "Contact Name", "Address", "City", "State", "Zip",
    "Phone", "Email", "NMLS #", "Website", "Notes"
]

MORTGAGE_DATA = [
    (1,  "Legacy Mortgage Group",                    "",                        "Tyler",                      "Tyler",   "TX", "75703", "",                 "",                                    "355969",  "txhomeloans.com",                    "Highest-rated lender in Tyler on Google (350+ reviews)"),
    (2,  "Tyler Lending Group",                      "Kelvin Blake Woodfin",    "6649 Oak Hill Blvd",         "Tyler",   "TX", "75703", "",                 "",                                    "314553",  "tylerlending.com",                   "Independent brokerage; shops nationwide lender network"),
    (3,  "Mortgage Financial Services (Tyler Mortgage)","Jeanna Hill",           "455 Rice Rd., #111",         "Tyler",   "TX", "",      "(903) 360-1131",   "jhill@mfsus.com",                     "253767",  "jhill.mortgagefinancial.com",         ""),
    (4,  "CMG Home Loans – Tyler Branch",             "J. McCrary",              "1028 Asher Way, Suite 100",  "Tyler",   "TX", "75703", "(903) 316-2883",   "jmccrary@cmgfi.com",                  "",        "cmghomeloans.com",                   ""),
    (5,  "Southside Bank Mortgage",                   "Emmanuel Terrazas",       "1201 S. Beckham Ave.",       "Tyler",   "TX", "",      "(903) 531-7229",   "emmanuel.terrazas@southside.com",     "",        "southside.mymortgage-online.com",    ""),
    (6,  "Movement Mortgage, LLC",                    "Mallory Mazarakes",       "119 W Charnwood St",         "Tyler",   "TX", "75701", "",                 "",                                    "",        "movementmortgage.com",               ""),
    (7,  "Integrity Mortgage Services",               "Shonda Denise Lowe",      "305 S Broadway St.",         "Tyler",   "TX", "75702", "",                 "",                                    "268453",  "",                                   ""),
    (8,  "Lifeway Lending Group, Inc.",               "Waverly Ann Bratcher",    "305 E South St.",            "Lindale", "TX", "75771", "",                 "",                                    "283223",  "",                                   ""),
    (9,  "Finance of America Mortgage",               "Kristie Van Huis",        "212 Old Grande Blvd",        "Tyler",   "TX", "75703", "",                 "",                                    "",        "financeofamerica.com",               ""),
    (10, "Altra Federal Credit Union",                "Jesus Flores",            "5523 Troop Hwy",             "Tyler",   "TX", "75707", "",                 "",                                    "1389522", "altra.org",                          ""),
    (11, "Fairway Independent Mortgage Corp.",        "",                        "Tyler, TX 75702",            "Tyler",   "TX", "75702", "",                 "",                                    "",        "fairway.com",                        ""),
    (12, "Tyler Home Mortgage, LLC",                  "",                        "",                           "Tyler",   "TX", "",      "",                 "",                                    "",        "",                                   ""),
    (13, "Premier Nationwide Lending",                "",                        "",                           "Tyler",   "TX", "",      "",                 "",                                    "",        "",                                   ""),
    (14, "Highlands Residential Mortgage",            "Craig Nichols",           "",                           "Tyler",   "TX", "",      "",                 "",                                    "",        "",                                   ""),
    (15, "Cornerstone Brokerage, LLC",                "Lorri Loggins",           "",                           "Tyler",   "TX", "",      "",                 "",                                    "",        "",                                   ""),
    (16, "Coleman Mortgage Company",                  "",                        "",                           "Tyler",   "TX", "",      "",                 "",                                    "",        "",                                   ""),
    (17, "Supreme Lending",                           "",                        "",                           "Tyler",   "TX", "",      "",                 "",                                    "",        "",                                   ""),
    (18, "Maverick Residential Mortgage",             "",                        "",                           "Tyler",   "TX", "",      "",                 "",                                    "",        "",                                   ""),
    (19, "Prosperity Home Mortgage, LLC",             "Tyler Traeger",           "",                           "Tyler",   "TX", "",      "",                 "",                                    "",        "phmloans.com",                       ""),
]

INVEST_HEADERS = [
    "#", "Firm", "Key Contact", "Address", "City", "State", "Zip",
    "Phone", "Website", "Type", "Notes"
]

INVEST_DATA = [
    (1,  "Merrill Lynch Wealth Management",     "",                              "Tyler, TX 75703",                "Tyler", "TX", "75703", "",              "advisor.ml.com/sites/tx/tyler-tx",          "Wealth Management",         "Bank of America affiliate; retirement + estate planning"),
    (2,  "Morgan Stanley – Tyler Branch",       "",                              "1 America Bldg, Loop 323 & Copeland Rd", "Tyler", "TX", "", "",          "advisor.morganstanley.com/tyler-branch",    "Full-Service Broker-Dealer","NMLS# 1364474"),
    (3,  "RBC Wealth Management – Tyler",       "Thomas Smith, CAP/CFP/CLU/CPA", "1347 Dominion Plaza",            "Tyler", "TX", "75703", "",             "us.rbcwealthmanagement.com/tyler",          "Wealth Management",         "Managing Director & Branch Director"),
    (4,  "Stifel Financial – Tyler Branch",     "",                              "Tyler, TX",                      "Tyler", "TX", "",      "",              "stifel.com/branch/tx/tyler",               "Investment Banking + Wealth","Full-service firm"),
    (5,  "Trinity Capital Management",          "",                              "821 ESE Loop 323, Suite 110",    "Tyler", "TX", "75701", "(903) 747-3960","tcmtx.com",                                "Independent RIA",           "Independent wealth management"),
    (6,  "Financial Synergies Wealth Advisors", "Will Goodson, CFP",             "Tyler, TX",                      "Tyler", "TX", "",      "",              "financialadvisortylertx.com",              "Fee-Only RIA",              "Retirement + investment management"),
    (7,  "Texas Financial and Retirement",      "",                              "Tyler, TX 75703",                "Tyler", "TX", "75703", "",              "texasfinancialandretirement.com",           "IAR / Insurance",           "Series 65 IAR; insurance + investment advisory"),
    (8,  "Edward Jones",                        "Various advisors",              "Multiple locations",             "Tyler", "TX", "",      "",              "edwardjones.com",                          "Full-Service Broker-Dealer","Multiple Smith County offices"),
    (9,  "Raymond James Financial Services",    "",                              "Tyler, TX",                      "Tyler", "TX", "",      "",              "raymondjames.com",                         "Full-Service Broker-Dealer",""),
    (10, "Ameriprise Financial Services",       "Rod Boaz",                      "Tyler, TX",                      "Tyler", "TX", "",      "",              "ameriprise.com",                           "Wealth Management",         ""),
    (11, "Northwestern Mutual",                 "Travis Bearden",                "Tyler, TX",                      "Tyler", "TX", "",      "",              "northwesternmutual.com",                   "Wealth / Insurance",        ""),
    (12, "Northrock Investment Management",     "",                              "Tyler, TX",                      "Tyler", "TX", "",      "",              "",                                         "Investment Management",     ""),
    (13, "Income Portfolio Strategies",         "",                              "Tyler, TX",                      "Tyler", "TX", "",      "",              "",                                         "Investment Management",     ""),
    (14, "Drake Real Estate & Investments",     "",                              "11621 CR 166",                   "Tyler", "TX", "75703", "(903) 581-3737","draketexas.com",                           "Business Brokerage / M&A",  "Buying and selling small-to-midsize businesses"),
]

# ── Build workbook ─────────────────────────────────────────────────────────

wb = openpyxl.Workbook()

def build_sheet(wb, tab_name, title_text, headers, data):
    ws = wb.create_sheet(title=tab_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "CC0000"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = title_font()
    title_cell.fill = PatternFill("solid", fgColor=BLACK)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Sub-header row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sub_cell = ws.cell(row=2, column=1, value="Smith County, TX  |  Scraped 2026-03-26  |  Verify current info at nmlsconsumeraccess.org")
    sub_cell.font = Font(name="Arial", color="FF888888", size=8, italic=True)
    sub_cell.fill = PatternFill("solid", fgColor=BLACK)
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = hdr_font()
        cell.fill = hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border()
    ws.row_dimensions[3].height = 18

    # Data rows
    for row_idx, row_data in enumerate(data, 4):
        fill = alt_fill() if row_idx % 2 == 0 else row_fill()
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font()
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = thin_border()
        ws.row_dimensions[row_idx].height = 16

    return ws

# Sheet 1 – Mortgage Brokers
ws1 = build_sheet(wb, "Mortgage Brokers", "MORTGAGE BROKERS — Smith County TX", MORTGAGE_HEADERS, MORTGAGE_DATA)
set_col_widths(ws1, [4, 34, 24, 28, 10, 6, 7, 16, 30, 10, 28, 44])

# Sheet 2 – Investment / Wealth
ws2 = build_sheet(wb, "Investment & Wealth", "INVESTMENT BANKERS & WEALTH MANAGERS — Smith County TX", INVEST_HEADERS, INVEST_DATA)
set_col_widths(ws2, [4, 34, 26, 32, 10, 6, 7, 16, 36, 26, 44])

# Remove default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

out_path = "/home/user/Claude/output/smith-county-brokers.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
