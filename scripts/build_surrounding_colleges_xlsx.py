"""
Build Surrounding Counties College Contacts XLSX
Counties: Gregg, Cherokee, Henderson, Rusk, Upshur, Wood, Panola
Output: /home/user/Claude/downloads/surrounding-counties-college-contacts.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLACK = "FF000000"; WHITE = "FFFFFFFF"; RED = "FFCC0000"
GRAY  = "FF1A1A1A"; LGRAY = "FF2A2A2A"

def hdr_fill():   return PatternFill("solid", fgColor=RED)
def row_fill():   return PatternFill("solid", fgColor=BLACK)
def alt_fill():   return PatternFill("solid", fgColor=LGRAY)
def hdr_font():   return Font(name="Arial", bold=True, color=WHITE, size=10)
def body_font():  return Font(name="Arial", color=WHITE, size=9)
def title_font(): return Font(name="Arial", bold=True, color=RED, size=12)
def sub_font():   return Font(name="Arial", color="FF888888", size=8, italic=True)

def bdr():
    s = Side(border_style="thin", color="FF333333")
    return Border(left=s, right=s, top=s, bottom=s)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

HEADERS = ["#", "Name", "Title", "Department", "Phone", "Email", "Organization", "County", "Notes"]

def build_sheet(wb, tab, title_text, subtitle, data):
    ws = wb.create_sheet(title=tab)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "CC0000"
    n = len(HEADERS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = title_font(); c.fill = PatternFill("solid", fgColor=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = sub_font(); c.fill = PatternFill("solid", fgColor=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font(); c.fill = hdr_fill()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr()
    ws.row_dimensions[3].height = 18

    for ri, row in enumerate(data, 4):
        fill = alt_fill() if ri % 2 == 0 else row_fill()
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = body_font(); c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=False)
            c.border = bdr()
        ws.row_dimensions[ri].height = 16
    return ws

# ═══════════════════════════════════════════════════════════════════════
# ALL DATA — one master sheet + one per institution
# ═══════════════════════════════════════════════════════════════════════

# --- LeTourneau University (Gregg County / Longview) ---
LETU = "LeTourneau University"
letu = [
    (1,  "Tom Bevan",   "President / CEO",             "Office of the President",  "903-233-3000",  "tom.bevan@letu.edu",    LETU, "Gregg",   "Main campus: 2100 S. Mobberly Ave, Longview TX 75602"),
    (2,  "",            "Admissions",                  "Admissions",               "800-759-8811",  "Admissions@LETU.edu",   LETU, "Gregg",   ""),
    (3,  "",            "Athletics",                   "Athletics",                "903-233-3760",  "",                      LETU, "Gregg",   "Fax: 903-233-3751"),
    (4,  "",            "Main Switchboard",            "General",                  "903-233-3000",  "",                      LETU, "Gregg",   "Fax: 903-233-4301"),
    (5,  "",            "Tyler Campus",                "Tyler Location",           "",              "",                      LETU, "Smith",   "501 Shelley Dr, Tyler TX — call 903-233-3000"),
]

# --- UT Tyler Longview Campus (Gregg County) ---
UTL = "UT Tyler – Longview University Center"
utl = [
    (1,  "",  "Longview Campus – General",     "Longview University Center",  "903-566-7200",  "",                       UTL, "Gregg", "3900 University Blvd, Tyler TX 75799 (admin); campus in Longview"),
    (2,  "",  "Nursing Program",               "College of Nursing",          "903-566-7200",  "",                       UTL, "Gregg", "98% of BSN graduates enter workforce or grad school"),
    (3,  "",  "Admissions (main)",             "Undergraduate Admissions",    "903-566-7203",  "admissions@uttyler.edu", UTL, "Gregg", "Longview programs administered via UT Tyler main campus"),
]

# --- Kilgore College (Gregg / Rusk Counties) ---
KC = "Kilgore College"
kc = [
    (1,  "",  "President's Office",            "Office of the President",     "903-983-8209",  "",                       KC, "Gregg/Rusk", "1100 Broadway, Kilgore TX 75662"),
    (2,  "",  "Admissions",                    "Admissions",                  "903-983-8134",  "",                       KC, "Gregg/Rusk", ""),
    (3,  "",  "Kilgore Main Campus",           "General",                     "903-983-8209",  "",                       KC, "Gregg/Rusk", "1100 Broadway, Kilgore TX 75662"),
    (4,  "",  "KC-Longview Campus",            "General",                     "903-753-2642",  "",                       KC, "Gregg",      "300 S. High Street, Longview TX 75601"),
    (5,  "",  "TRIO Educational Opp. Center",  "TRIO / EOC",                  "903-983-8209",  "",                       KC, "Gregg/Harrison/Rusk/Upshur", "Serves 4-county area"),
]

# --- Jacksonville College (Cherokee County) ---
JC = "Jacksonville College"
jc = [
    (1,  "Danny Morris",  "Director of Admissions",    "Admissions",            "903-589-7110",  "admissions@jacksonville-college.org", JC, "Cherokee", "105 B.J. Albritton Dr, Jacksonville TX 75766"),
    (2,  "",              "Registrar",                 "Office of the Registrar","903-586-2518",  "registrar@jacksonville-college.edu",  JC, "Cherokee", "Norman Building front desk"),
    (3,  "",              "IT / Help Desk",            "Information Technology", "903-586-2518",  "help@jacksonville-college.edu",       JC, "Cherokee", ""),
    (4,  "",              "Main Switchboard",          "General",               "903-586-2518",  "",                                    JC, "Cherokee", "105 B.J. Albritton Dr, Jacksonville TX 75766"),
]

# --- Baptist Missionary Association Theological Seminary (Cherokee County) ---
BMA = "Baptist Missionary Association Theological Seminary"
bma = [
    (1,  "Dr. Philip Attebery", "President / CEO / CAO / Prof of Church Ministries", "Office of the President", "903-586-2501", "", BMA, "Cherokee", "1530 E Pine St, Jacksonville TX 75766; PO Box 670"),
    (2,  "",                    "Main Switchboard",                                   "General",                 "903-586-2501", "", BMA, "Cherokee", "www.bmats.edu"),
]

# --- Trinity Valley Community College (Henderson County) ---
TVCC = "Trinity Valley Community College"
tvcc = [
    (1,  "",  "Main Switchboard",           "General",           "903-675-6200",  "",               TVCC, "Henderson", "100 Cardinal Dr, Athens TX 75751"),
    (2,  "",  "Admissions / General Info",  "Admissions",        "903-677-8822",  "",               TVCC, "Henderson", "Toll-free: 1-866-882-2937"),
    (3,  "",  "Housing",                    "Housing",           "903-675-6256",  "housing@tvcc.edu", TVCC, "Henderson", ""),
    (4,  "",  "Kaufman Campus",             "Kaufman Campus",    "",              "",               TVCC, "Kaufman",   "Kaufman TX 75142"),
    (5,  "",  "Palestine Campus",           "Palestine Campus",  "",              "",               TVCC, "Anderson",  "Palestine TX 75802"),
]

# --- Panola College (Panola County) ---
PC = "Panola College"
pc = [
    (1,  "Jeremy Dorman", "Director of Human Resources",  "Human Resources",  "903-693-2000",  "jdorman@panola.edu",  PC, "Panola", "1109 W. Panola, Carthage TX 75633; Fax: 903-693-1122"),
    (2,  "",              "Admissions",                   "Admissions",       "903-693-2038",  "",                    PC, "Panola", "Text: 903-500-9110"),
    (3,  "",              "Main Switchboard",             "General",          "903-693-2000",  "",                    PC, "Panola", "1109 W. Panola, Carthage TX 75633"),
    (4,  "",              "Marshall Campus",              "Marshall Campus",  "903-935-5039",  "",                    PC, "Harrison","1300 E Pinecrest Dr #126, Marshall TX 75670"),
    (5,  "",              "Center Campus",                "Center Campus",    "",              "",                    PC, "Shelby",  "678 Rough Rider Dr, Center TX 75935"),
]

# --- Jarvis Christian University (Wood County) ---
JCU = "Jarvis Christian University"
jcu = [
    (1,  "Brandon Byrd",  "Director of Admissions & Enrollment", "Admissions",      "903-730-4890",         "recruitment@jarvis.edu",  JCU, "Wood", "PO Box 1470, Hawkins TX 75765; Fax: 903-769-4842"),
    (2,  "",              "Registrar",                           "Registrar",        "903-730-4890 x2453",   "registrar@jarvis.edu",    JCU, "Wood", ""),
    (3,  "",              "Financial Aid",                       "Financial Aid",    "903-730-4890 x4890",   "finaid@jarvis.edu",       JCU, "Wood", "Option 4"),
    (4,  "",              "Business Office",                     "Business Office",  "903-730-4890 x2709",   "svalentine@jarvis.edu",   JCU, "Wood", ""),
    (5,  "",              "Campus Tours / Admissions",           "Admissions",       "903-730-4890 x2202",   "recruitment@jarvis.edu",  JCU, "Wood", "FAFSA Code: 003637"),
]

# --- Northeast Texas Community College (Titus / Morris Counties) ---
NTCC = "Northeast Texas Community College"
ntcc = [
    (1,  "",  "Main Switchboard",                "General",              "903-434-8100",  "",  NTCC, "Titus/Morris", "2886 FM 1735 Chapel Hill Rd, Mt. Pleasant TX 75455"),
    (2,  "",  "Admissions",                      "Admissions",           "903-434-8100",  "",  NTCC, "Titus/Morris", "Toll-free: 800-870-0142"),
    (3,  "",  "Pittsburg / Culinary Arts",        "Culinary Arts",        "903-434-8333",  "",  NTCC, "Camp",         "114 Jefferson St, Pittsburg TX 75686"),
    (4,  "",  "Pittsburg / Hanson-Sewell",        "Hanson-Sewell Center", "903-434-8393",  "",  NTCC, "Camp",         "237 College St, Pittsburg TX 75686; Fax: 903-855-0442"),
    (5,  "",  "Industrial Technology Ctr",        "Industrial Tech",      "903-434-8390",  "",  NTCC, "Titus",        "1100 Lakewood Dr, Mt. Pleasant TX 75455"),
]

# ═══════════════════════════════════════════════════════════════════════
# MASTER / SUMMARY SHEET — all contacts combined
# ═══════════════════════════════════════════════════════════════════════
all_data = []
idx = 1
for block in [letu, utl, kc, jc, bma, tvcc, pc, jcu, ntcc]:
    for row in block:
        all_data.append((idx,) + row[1:])
        idx += 1

# ═══════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

WIDTHS = [4, 22, 30, 26, 18, 30, 30, 16, 38]

ws0 = build_sheet(wb, "ALL CONTACTS",
    "SURROUNDING COUNTIES — All College Contacts",
    "Smith County Surrounding Counties TX  |  Scraped 2026-03-26  |  Verify at trec.texas.gov & institution websites",
    all_data)
set_widths(ws0, WIDTHS)

ws1 = build_sheet(wb, "LeTourneau Univ",
    "LETOURNEAU UNIVERSITY — Gregg County (Longview TX)",
    "2100 S. Mobberly Ave, Longview TX 75602  |  903-233-3000  |  Scraped 2026-03-26",
    letu); set_widths(ws1, WIDTHS)

ws2 = build_sheet(wb, "UT Tyler Longview",
    "UT TYLER – LONGVIEW UNIVERSITY CENTER — Gregg County",
    "Longview, TX  |  903-566-7200  |  Scraped 2026-03-26",
    utl); set_widths(ws2, WIDTHS)

ws3 = build_sheet(wb, "Kilgore College",
    "KILGORE COLLEGE — Gregg/Rusk Counties",
    "1100 Broadway, Kilgore TX 75662  |  903-983-8209  |  Scraped 2026-03-26",
    kc); set_widths(ws3, WIDTHS)

ws4 = build_sheet(wb, "Jacksonville College",
    "JACKSONVILLE COLLEGE — Cherokee County",
    "105 B.J. Albritton Dr, Jacksonville TX 75766  |  903-586-2518  |  Scraped 2026-03-26",
    jc); set_widths(ws4, WIDTHS)

ws5 = build_sheet(wb, "BMA Seminary",
    "BAPTIST MISSIONARY ASSOC THEOLOGICAL SEMINARY — Cherokee County",
    "1530 E Pine St, Jacksonville TX 75766  |  903-586-2501  |  Scraped 2026-03-26",
    bma); set_widths(ws5, WIDTHS)

ws6 = build_sheet(wb, "Trinity Valley CC",
    "TRINITY VALLEY COMMUNITY COLLEGE — Henderson County (Athens TX)",
    "100 Cardinal Dr, Athens TX 75751  |  903-675-6200  |  Scraped 2026-03-26",
    tvcc); set_widths(ws6, WIDTHS)

ws7 = build_sheet(wb, "Panola College",
    "PANOLA COLLEGE — Panola County (Carthage TX)",
    "1109 W. Panola, Carthage TX 75633  |  903-693-2000  |  Scraped 2026-03-26",
    pc); set_widths(ws7, WIDTHS)

ws8 = build_sheet(wb, "Jarvis Christian Univ",
    "JARVIS CHRISTIAN UNIVERSITY — Wood County (Hawkins TX)",
    "PO Box 1470, Hawkins TX 75765  |  903-730-4890  |  Scraped 2026-03-26",
    jcu); set_widths(ws8, WIDTHS)

ws9 = build_sheet(wb, "NE Texas CC",
    "NORTHEAST TEXAS COMMUNITY COLLEGE — Titus County (Mt. Pleasant TX)",
    "2886 FM 1735 Chapel Hill Rd, Mt. Pleasant TX 75455  |  903-434-8100  |  Scraped 2026-03-26",
    ntcc); set_widths(ws9, WIDTHS)

if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

out = "/home/user/Claude/downloads/surrounding-counties-college-contacts.xlsx"
wb.save(out)
print(f"Saved: {out}")
