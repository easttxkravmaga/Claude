"""
Build Smith County TX — Real Estate Agents & Mortgage Brokers XLSX
Drops to /home/user/Downloads/smith-county-real-estate-agents.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Brand colors ──────────────────────────────────────────────────────────
BLACK  = "FF000000"
WHITE  = "FFFFFFFF"
RED    = "FFCC0000"
GRAY   = "FF1A1A1A"
LGRAY  = "FF2A2A2A"

def hdr_fill():   return PatternFill("solid", fgColor=RED)
def row_fill():   return PatternFill("solid", fgColor=BLACK)
def alt_fill():   return PatternFill("solid", fgColor=LGRAY)
def hdr_font():   return Font(name="Arial", bold=True, color=WHITE, size=10)
def body_font():  return Font(name="Arial", color=WHITE, size=9)
def title_font(): return Font(name="Arial", bold=True, color=RED, size=12)
def sub_font():   return Font(name="Arial", color="FF888888", size=8, italic=True)

def border():
    s = Side(border_style="thin", color="FF333333")
    return Border(left=s, right=s, top=s, bottom=s)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_sheet(wb, tab_name, title_text, headers, data):
    ws = wb.create_sheet(title=tab_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "CC0000"

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = title_font()
    c.fill = PatternFill("solid", fgColor=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Sub-header
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    c = ws.cell(row=2, column=1,
                value="Smith County, TX  |  Scraped 2026-03-26  |  Verify licenses at trec.texas.gov & nmlsconsumeraccess.org")
    c.font = sub_font()
    c.fill = PatternFill("solid", fgColor=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    # Headers
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font()
        c.fill = hdr_fill()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border()
    ws.row_dimensions[3].height = 18

    # Data
    for ri, row in enumerate(data, 4):
        fill = alt_fill() if ri % 2 == 0 else row_fill()
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = body_font()
            c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=False)
            c.border = border()
        ws.row_dimensions[ri].height = 16

    return ws

# ══════════════════════════════════════════════════════════════════════════
# SHEET 1 — REAL ESTATE AGENTS & BROKERAGES
# ══════════════════════════════════════════════════════════════════════════
RE_HEADERS = [
    "#", "Agency / Brokerage", "Agent / Contact", "Address",
    "City", "State", "Zip", "Phone", "Email", "Website", "Notes"
]

RE_DATA = [
    (1,  "Keller Williams Realty – Tyler",          "Shannon Robinson",         "6761 Old Jacksonville Hwy",  "Tyler",   "TX", "75703", "(903) 534-6600",  "shannon.robinson@kw.com",    "tylertx.kw.com",              "54 agents; buyers & sellers"),
    (2,  "Keller Williams Realty – Tyler",          "Karen Conn",               "6761 Old Jacksonville Hwy",  "Tyler",   "TX", "75703", "(430) 205-3996",  "",                            "tylertx.kw.com",              ""),
    (3,  "Keller Williams Realty – Tyler",          "Claudia Carroll",          "6761 Old Jacksonville Hwy",  "Tyler",   "TX", "75703", "",                "",                            "tylerhomes.com",              ""),
    (4,  "Keller Williams Realty – Tyler",          "Judy Kunzman",             "6761 Old Jacksonville Hwy",  "Tyler",   "TX", "75703", "",                "",                            "tylertx.kw.com",              "Top 5%; 41 closings in 2 yrs; luxury, HOA, first-time buyers"),
    (5,  "RE/MAX Tyler",                            "Mary Smith",               "4300 Kinsey Dr.",            "Tyler",   "TX", "75703", "(903) 530-9707",  "",                            "tylerpropertygroup.com",      "32 yrs exp; luxury, investment, farm/ranch, multi-family"),
    (6,  "RE/MAX Tyler",                            "The Burks Team",           "4300 Kinsey Dr.",            "Tyler",   "TX", "75703", "",                "",                            "remax.com/real-estate-agents/tyler-tx", "Top-rated team"),
    (7,  "RE/MAX Tyler",                            "Rene Griffin-Perry",       "4300 Kinsey Dr.",            "Tyler",   "TX", "75703", "(903) 474-2499",  "",                            "remax.com/real-estate-agents/tyler-tx", ""),
    (8,  "Coldwell Banker Apex, Realtors",          "Shannon Franzen",          "Tyler, TX",                  "Tyler",   "TX", "",      "(903) 500-9204",  "",                            "coldwellbanker.com",          "44 total sales; $80K–$725K price range"),
    (9,  "Coldwell Banker Apex, Realtors",          "Managing Broker",          "Tyler, TX",                  "Tyler",   "TX", "",      "",                "",                            "coldwellbanker.com",          "Family-owned; 20 licensed agents; Tyler/DFW/Henderson MLS"),
    (10, "Gregory Real Estate",                     "Sam Zigtema",              "113 Robert E. Lee Dr.",      "Tyler",   "TX", "75703", "(903) 530-5483",  "",                            "gregoryrealtors.com",         "In Tyler since 1979; donates 5% commission to client's charity"),
    (11, "Gregory Real Estate",                     "Emily (REALTOR)",          "113 Robert E. Lee Dr.",      "Tyler",   "TX", "75703", "(903) 581-0016",  "",                            "gregoryrealtors.com",         "REALTOR since 2004; residential buyers & sellers"),
    (12, "Gregory Real Estate / Broker-Owner",      "",                         "113 Robert E. Lee Dr.",      "Tyler",   "TX", "75703", "(903) 581-0016",  "",                            "gregoryrealtors.com",         "CCIM; broker license 2004; property mgmt, investment, commercial"),
    (13, "The Agency Tyler",                        "Vanessa Griffin",          "Tyler, TX",                  "Tyler",   "TX", "",      "",                "",                            "theagencytylertx.com",        "Managing Partner; 20+ yrs exp; local buyers & investors"),
    (14, "Leslie Cain Realty",                      "Leslie Cain",              "Lindale / Tyler, TX",        "Lindale", "TX", "",      "",                "",                            "",                            "Owner-broker; boutique; Rookie Realtor of Year 2012 (GTAR)"),
    (15, "Dee Martin Realty Group",                 "Dee Martin",               "Tyler, TX",                  "Tyler",   "TX", "",      "",                "",                            "",                            "Residential; single/multi-family, manufactured, condos, townhomes"),
    (16, "eXp Realty",                              "Shalon 'Shay' Garner",     "Tyler, TX",                  "Tyler",   "TX", "",      "",                "",                            "",                            "Highly rated; Tyler area"),
    (17, "RealEdge Real Estate",                    "Doc Deason",               "Tyler, TX",                  "Tyler",   "TX", "",      "",                "",                            "",                            "Well-reviewed; vast network of contacts"),
    (18, "RealEdge Real Estate",                    "Amanda Jackson (Broker)",  "Tyler, TX",                  "Tyler",   "TX", "",      "",                "",                            "",                            ""),
    (19, "Jones & Co. Realty Group",                "",                         "Tyler, TX",                  "Tyler",   "TX", "",      "",                "",                            "",                            "Top-listed company in Tyler"),
    (20, "Texas Real Estate Source",                "",                         "Tyler, TX",                  "Tyler",   "TX", "",      "(469) 885-7774",  "",                            "texasrealestatesource.com",   ""),
    (21, "Shilling Real Estate Company",            "",                         "Tyler, TX",                  "Tyler",   "TX", "",      "",                "",                            "",                            ""),
    (22, "SET Real Estate",                         "",                         "Tyler, TX",                  "Tyler",   "TX", "",      "",                "",                            "",                            ""),
    (23, "Realty One Group Rose",                   "",                         "Tyler, TX",                  "Tyler",   "TX", "",      "",                "",                            "",                            ""),
]

# ══════════════════════════════════════════════════════════════════════════
# SHEET 2 — MORTGAGE BROKERS & LENDERS
# ══════════════════════════════════════════════════════════════════════════
MTG_HEADERS = [
    "#", "Company", "Contact Name", "Address",
    "City", "State", "Zip", "Phone", "Email", "NMLS #", "Website", "Notes"
]

MTG_DATA = [
    (1,  "Legacy Mortgage Group",                       "",                     "Tyler, TX 75703",             "Tyler",   "TX", "75703", "",                "",                            "355969",  "txhomeloans.com",              "Highest-rated lender Tyler TX; 350+ Google reviews"),
    (2,  "Tyler Lending Group",                         "Kelvin Blake Woodfin", "6649 Oak Hill Blvd",          "Tyler",   "TX", "75703", "",                "",                            "314553",  "tylerlending.com",             "Independent broker; nationwide lender network"),
    (3,  "Mortgage Financial Services (Tyler Mortgage)","Jeanna Hill",          "455 Rice Rd., #111",          "Tyler",   "TX", "",      "(903) 360-1131",  "jhill@mfsus.com",             "253767",  "jhill.mortgagefinancial.com",  ""),
    (4,  "CMG Home Loans – Tyler Branch",               "J. McCrary",           "1028 Asher Way, Suite 100",   "Tyler",   "TX", "75703", "(903) 316-2883",  "jmccrary@cmgfi.com",          "",        "cmghomeloans.com",             ""),
    (5,  "Southside Bank Mortgage",                     "Johnna Hutchins",      "1201 S. Beckham Ave.",        "Tyler",   "TX", "",      "(903) 531-7229",  "",                            "580531",  "southside.mymortgage-online.com",""),
    (6,  "Southside Bank Mortgage",                     "Whitney Parks",        "16691 FM 2493",               "Tyler",   "TX", "75703", "",                "",                            "827730",  "southside.mymortgage-online.com",""),
    (7,  "Southside Bank Mortgage",                     "Emmanuel Terrazas",    "1201 S. Beckham Ave.",        "Tyler",   "TX", "",      "(903) 531-7229",  "emmanuel.terrazas@southside.com","",     "southside.mymortgage-online.com",""),
    (8,  "Movement Mortgage, LLC",                      "Mallory Mazarakes",    "119 W Charnwood St",          "Tyler",   "TX", "75701", "",                "",                            "",        "movementmortgage.com",         ""),
    (9,  "Integrity Mortgage Services",                 "Shonda Denise Lowe",   "305 S Broadway St.",          "Tyler",   "TX", "75702", "",                "",                            "268453",  "",                             ""),
    (10, "Lifeway Lending Group, Inc.",                 "Waverly Ann Bratcher", "305 E South St.",             "Lindale", "TX", "75771", "",                "",                            "283223",  "",                             ""),
    (11, "Finance of America Mortgage",                 "Kristie Van Huis",     "212 Old Grande Blvd",         "Tyler",   "TX", "75703", "",                "",                            "266847",  "financeofamerica.com",         ""),
    (12, "Altra Federal Credit Union",                  "Jesus Flores",         "5523 Troop Hwy",              "Tyler",   "TX", "75707", "",                "",                            "1389522", "altra.org",                    ""),
    (13, "Republic State Mortgage Co.",                 "Morris Norman",        "2341 Dueling Oaks",           "Tyler",   "TX", "75703", "",                "",                            "653983",  "",                             ""),
    (14, "Rate (formerly Guaranteed Rate)",             "Sherry Paull",         "Tyler, TX",                   "Tyler",   "TX", "",      "",                "",                            "270955",  "rate.com",                     "20+ yrs exp; branch manager/VP"),
    (15, "Fairway Independent Mortgage Corp.",          "",                     "Tyler, TX 75702",             "Tyler",   "TX", "75702", "",                "",                            "",        "fairway.com",                  ""),
    (16, "Prosperity Home Mortgage, LLC",               "Tyler Traeger",        "Tyler, TX",                   "Tyler",   "TX", "",      "",                "",                            "",        "phmloans.com",                 ""),
    (17, "Hallmark Home Mortgage – Tyler",              "",                     "Tyler, TX",                   "Tyler",   "TX", "",      "",                "",                            "",        "",                             ""),
    (18, "Premier Nationwide Lending",                  "",                     "Tyler, TX",                   "Tyler",   "TX", "",      "",                "",                            "",        "",                             ""),
    (19, "Highlands Residential Mortgage",              "Craig Nichols",        "Tyler, TX",                   "Tyler",   "TX", "",      "",                "",                            "",        "",                             ""),
    (20, "Cornerstone Brokerage, LLC",                  "Lorri Loggins",        "Tyler, TX",                   "Tyler",   "TX", "",      "",                "",                            "",        "",                             ""),
    (21, "Coleman Mortgage Company",                    "",                     "Tyler, TX",                   "Tyler",   "TX", "",      "",                "",                            "",        "",                             ""),
    (22, "Supreme Lending",                             "",                     "Tyler, TX",                   "Tyler",   "TX", "",      "",                "",                            "",        "",                             ""),
    (23, "Maverick Residential Mortgage",               "",                     "Tyler, TX",                   "Tyler",   "TX", "",      "",                "",                            "",        "",                             ""),
    (24, "Tyler Home Mortgage, LLC",                    "",                     "Tyler, TX",                   "Tyler",   "TX", "",      "",                "",                            "",        "",                             ""),
    (25, "Independent Mortgage (Christopher Fussell)",  "Christopher Fussell",  "6115 New Copeland Rd",        "Tyler",   "TX", "75703", "",                "",                            "518007",  "",                             ""),
]

# ══════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

ws1 = build_sheet(wb, "Real Estate Agents", "REAL ESTATE AGENTS & BROKERAGES — Smith County TX", RE_HEADERS, RE_DATA)
set_widths(ws1, [4, 30, 24, 26, 10, 6, 7, 16, 30, 26, 44])

ws2 = build_sheet(wb, "Mortgage Brokers", "MORTGAGE BROKERS & LENDERS — Smith County TX", MTG_HEADERS, MTG_DATA)
set_widths(ws2, [4, 32, 22, 26, 10, 6, 7, 16, 30, 10, 26, 44])

if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

out = "/home/user/Downloads/smith-county-real-estate-agents.xlsx"
wb.save(out)
print(f"Saved: {out}")
