#!/usr/bin/env python3
"""Generate Smith County Gyms & Fitness / Personal Trainers spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=13, color="1B4332")
section_fill = PatternFill(start_color="B7E4C7", end_color="B7E4C7", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Business", "Type", "Owner/Manager", "Phone", "Email", "Address", "Website", "Notes"]

gyms = [
    # ── CHAIN GYMS ───────────────────────────────────────────────
    ("CHAIN GYMS", None),

    ("Planet Fitness Tyler (East)", "Budget Gym",
     "", "(903) 771-0103", "",
     "2747 E 5th St, Tyler, TX 75701",
     "https://planetfitness.com/gyms/tyler-tx", "24 hours"),

    ("Planet Fitness Tyler (South)", "Budget Gym",
     "", "(903) 698-6345", "",
     "8950 S Broadway Ave, Tyler, TX 75703",
     "https://planetfitness.com/gyms/tyler-south-tx", "24 hours"),

    ("Anytime Fitness Tyler (Troup Hwy)", "24-Hour Gym",
     "", "(903) 509-1121", "",
     "1827 Troup Hwy, Tyler, TX 75701",
     "https://anytimefitness.com/locations/tyler-texas-1064", ""),

    ("Anytime Fitness Tyler (Old Jax Hwy)", "24-Hour Gym",
     "", "(903) 617-6590", "",
     "6435 Old Jacksonville Hwy, Tyler, TX 75703",
     "https://anytimefitness.com/locations/tyler-texas-1957", ""),

    ("Crunch Fitness Tyler", "Gym",
     "", "(903) 251-9194", "",
     "1909 E SE Loop 323, Tyler, TX 75701",
     "https://crunch.com/locations/tyler", "24 hrs. From $9.99/mo. Kids Crunch childcare"),

    ("Workout Anytime Tyler South", "24-Hour Gym",
     "", "(903) 571-1095", "tyler.south@workoutanytime.com",
     "6004 S Broadway Ste 208, Tyler, TX 75703",
     "https://workoutanytime.com/tyler-south", "No start-up fees, month-to-month"),

    ("F45 Training Legacy Trail", "Functional Training",
     "", "(817) 876-9745", "legacytrail@f45training.com",
     "6970 Arbor Ridge Drive, Tyler, TX 75703",
     "https://f45training.com/studio/legacytrail", "Team-based functional training"),

    # ── CROSSFIT ─────────────────────────────────────────────────
    ("CROSSFIT", None),

    ("CrossFit Tyler (CFT)", "CrossFit",
     "John Hersey II", "(361) 550-2736", "crossfittyler2011@gmail.com",
     "4906 Hightech Drive, Tyler, TX 75703",
     "https://crossfittyler.com", "Est. 2011. H4 Fitness LLC. BBB A+"),

    ("Premier CrossFit", "CrossFit",
     "", "(903) 534-7031", "info@premiercrossfit.com",
     "1906 Capital Drive, Tyler, TX 75701",
     "https://premiercrossfit.com", "Since 2009. Also houses Athleo Barbell Club (USAW)"),

    ("CrossFit Chief", "CrossFit",
     "", "(903) 780-6775", "rich_video@hotmail.com",
     "6391 Elkton Way, Tyler, TX 75703",
     "https://crossfitchief.com", "Founded 2015. Free trial"),

    # ── YOGA / PILATES / BARRE ───────────────────────────────────
    ("YOGA / PILATES / BARRE", None),

    ("Studio B Pilates + Barre", "Pilates / Barre / TRX",
     "", "(903) 245-1653", "",
     "2469 Earl Campbell Pkwy, Ste B, Tyler, TX 75701",
     "https://studiobtyler.com", "Also Pilates teacher training host"),

    ("Pure Barre Tyler", "Barre Studio",
     "", "(903) 630-6154", "",
     "5100 Old Bullard Rd, Ste A, Tyler, TX 75703",
     "https://purebarre.com/location/tyler-tx", ""),

    ("Club Pilates Tyler", "Pilates (Reformer)",
     "", "(903) 747-8190", "tyler@clubpilates.com",
     "8926 S Broadway Ave, Ste 140, Tyler, TX 75703",
     "https://clubpilates.com/location/tyler", "Village at Cumberland Park"),

    ("Pilates Collective Tyler", "Pilates",
     "Stacy H.", "(903) 207-8559", "studio@pilatescollectivetyler.com",
     "208 Shelley Dr, Tyler, TX 75701",
     "https://pilatescollectivetyler.com", "Women-owned"),

    ("HOTWORX Tyler", "Infrared Fitness (Hot Yoga/HIIT)",
     "Ashlee (Owner), Robbie Campbell (Co-Owner)", "(903) 630-5152", "",
     "8926 S Broadway Ave, Ste 128, Tyler, TX 75703",
     "https://hotworx.net/studio/tyler", "24-hour access"),

    # ── LOCAL / INDEPENDENT GYMS ─────────────────────────────────
    ("LOCAL & INDEPENDENT GYMS", None),

    ("Downtown Iron 24/7 Gym", "Gym (Independent)",
     "", "(903) 343-2878", "",
     "116 W Erwin St, Tyler, TX 75702",
     "https://downtowniron.com", ""),

    ("Elite Fitness Tyler", "Gym (Independent, 24/7)",
     "Angela & Ashten", "(903) 630-7829", "elitefitness@elitefitnesstyler.com",
     "4100 Troup Hwy, Suite 200, Tyler, TX 75703",
     "https://elitefitnesstyler.com", "16,000 sq ft. FREEMOTION & Rogue equipment. Family-run"),

    ("The Press Gym", "Gym (Independent, 24/7)",
     "", "(903) 462-0200", "",
     "410 W Erwin St, Tyler, TX 75702",
     "https://presstyler.com", "$40/mo, no contracts. In Tyler Morning Telegraph bldg"),

    ("Raw Iron Gym Tyler", "Gym (Powerlifting/Strongman, 24/7)",
     "Tara Rickert (Owner), Michael Mathis (Co-Owner)", "(903) 535-7867", "rawirongymtyler@gmail.com",
     "3320 Troup Hwy, Ste 160, Tyler, TX 75701",
     "https://rawirongymtyler.com", "Part of Raw Iron chain (11 locations)"),

    ("Rise RX Fitness", "Gym (Independent)",
     "Michelle & Cune Pena", "(903) 920-0009", "",
     "3500 S Broadway Ave, Suite 300, Tyler, TX 75701",
     "https://riserxfitness.com", "12,000 sq ft. Latino- & women-owned. On-site nutritionist"),

    ("Life Changing Fitness", "Gym / Personal Training (24/7)",
     "Kelly Hitchcock; Terry Bunker", "(903) 630-7497", "info@lcftyler.com",
     "5407 New Copeland Rd, Suite 300, Tyler, TX 75703",
     "https://lifechangingfitnesstyler.com", "Also: lcftyler@gmail.com. Free one-week trial"),

    ("CHRISTUS Health & Fitness Center", "Hospital-Affiliated Fitness",
     "", "(903) 939-4665", "",
     "3593 E Grande Blvd, Tyler, TX 75707",
     "https://christushealth.org/locations/health-fitness-center-herrington-ornelas",
     "14,000 sq ft. 25m salt water lap pool. 24 hours"),

    ("CHRISTUS Human Performance Center", "Performance Training",
     "", "(903) 606-8855", "",
     "8591 S Broadway Ave, Tyler, TX 75703",
     "https://christushealth.org/locations/human-performance-center",
     "Youth sports performance (ages 8-18) and adult training"),

    ("YMCA of Tyler", "Community Fitness / Recreation",
     "Jeremy M. Bumgardner (CEO)", "(903) 597-3573", "",
     "225 South Vine Ave, Tyler, TX 75702",
     "https://tylerymca.org", "Also: Prime Time YMCA, 2105 Garden Valley Rd"),

    # ── BOXING / KICKBOXING ──────────────────────────────────────
    ("BOXING & KICKBOXING", None),

    ("Kingdom Fitness and Boxing", "Boxing / Kickboxing / HIIT",
     "Akil Olatunji Augustus", "(903) 710-1980", "",
     "212 Old Grande Blvd, Suite A-100, Tyler, TX 75703",
     "https://kingdomfitbox.com", "15+ years. Youth boxing. BBB A+"),

    ("RockBox Fitness Tyler", "Kickboxing / Strength Training",
     "Greg & Renee (Owners)", "(903) 949-6933", "",
     "6555 Old Jacksonville Hwy, Tyler, TX 75703",
     "https://rockboxfitness.com/locations/tyler", "First class free"),

    # ── MMA / BJJ (Cross-Referral Partners) ──────────────────────
    ("MMA / BJJ (Cross-Referral Partners for Krav Maga)", None),

    ("Gracie Barra Tyler BJJ", "Brazilian Jiu-Jitsu",
     "Tim Thompson (Owner), Prof. Dionathan Santos (Head Coach)", "(903) 520-1112", "info@gbtyler.com",
     "6371 Elkton Trail, Tyler, TX 75703",
     "https://gbtyler.com", "10+ years. 160+ students. Free trial"),

    ("Ethos Jiu Jitsu & Martial Arts", "BJJ / MMA / Judo / Kickboxing",
     "Chad Mark Decker (Owner)", "(903) 740-5491", "",
     "4703 D.C. Dr, Ste 105, Tyler, TX 75701",
     "https://ethosjiujitsutyler.com", "Since 2015. BBB A+. Women's kickboxing classes"),

    ("Relson Gracie / Lone Star MMA Academy", "BJJ / MMA / JKD / Kali",
     "Drew & Hong Arthur", "(903) 509-1662", "zenmasta@worldblackbelt.com",
     "3508-B Westway St, Tyler, TX 75703",
     "https://gracietyler.com", "Since 2003. Owner is retired Deputy U.S. Marshal"),

    ("Tyler Kung Fu & Fitness", "Kung Fu / Tai Chi / Kickboxing",
     "Brandon W. Jones", "(903) 597-0275", "tylerkungfu@gmail.com",
     "2533 E 5th St, Tyler, TX 75701",
     "https://tkff.com", "Tyler's oldest martial arts school, since 1996. BBB-accredited"),

    ("Premier Martial Arts South Tyler", "Karate / Kickboxing / Krav Maga",
     "", "(903) 625-4118", "",
     "8938 S Broadway Ave, Tyler, TX 75703",
     "https://premiermartialarts.com/texas-tyler", "Franchise. Adult & children's programs"),

    # ── PERSONAL TRAINING STUDIOS ────────────────────────────────
    ("PERSONAL TRAINING STUDIOS", None),

    ("B-FIT Tyler", "Personal Training",
     "Johnny Ray Barrera", "(903) 617-6713", "",
     "6421 S Broadway Ave, Suite 600, Tyler, TX 75703",
     "https://bfittyler.com", "30-min sessions. Nutritional coaching. BBB A+. Since 2013"),

    ("Redefine Fitness & Lifestyle", "Small Group Personal Training",
     "David", "(903) 339-0622", "david@redefinefitnessetx.com",
     "2251 Three Lakes Pkwy, Suite 102, Tyler, TX 75703",
     "https://redefinefitnessetx.com", "Opened April 2025"),

    # ── BOOT CAMP / OUTDOOR ──────────────────────────────────────
    ("BOOT CAMP & OUTDOOR FITNESS", None),

    ("Camp Gladiator East Texas", "Outdoor Group Fitness",
     "", "(512) 494-6966", "",
     "GABC: 1607 Troup Hwy; The ROC: 18700 US-69, Tyler",
     "https://campgladiator.com", "60-min outdoor workouts"),

    # ── SURROUNDING AREAS ────────────────────────────────────────
    ("SURROUNDING AREAS (Lindale, Longview)", None),

    ("Anytime Fitness Lindale", "24-Hour Gym",
     "Brian Bradford", "(903) 882-0202", "",
     "618 N Main St, Lindale, TX 75771",
     "", "Since 2007"),

    ("Fusion Athletic Club (Lindale)", "Gym / Athletic Performance",
     "", "(903) 218-3122", "",
     "140 Eagle Spirit Dr, Lindale, TX 75771",
     "https://fusionathletic.club", "15,000 sq ft. 24/7. Youth ages 8+ (Sports Lab)"),

    ("TK Fitness Gym (Lindale)", "Women-Only Strength & Conditioning",
     "Trisha Kinney", "", "",
     "15710 FM 849, Lindale, TX",
     "https://tkfitnessgym.com", ""),

    ("CHRISTUS Health & Fitness Lindale", "Hospital-Affiliated Fitness",
     "", "", "",
     "Lindale, TX",
     "https://christushealth.org/locations/health-fitness-center-lindale", ""),
]

ws = wb.active
ws.title = "Gyms & Fitness"

col_widths = [42, 30, 36, 18, 34, 50, 52, 56]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

row_num = 2
for record in gyms:
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=record[0])
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 28
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

output = "/home/user/Claude/Smith_County_Gyms_Fitness_Trainers.xlsx"
wb.save(output)
data_rows = sum(1 for r in gyms if r[1] is not None)
print(f"Saved {data_rows} contacts to {output}")
