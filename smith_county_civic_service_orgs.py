#!/usr/bin/env python3
"""Generate Smith County Civic & Service Organizations spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=13, color="1B4332")
section_fill = PatternFill(start_color="B7E4C7", end_color="B7E4C7", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Organization", "Type", "Contact Name", "Title", "Phone", "Email", "Address", "Website", "Notes"]

groups = [
    # ── ROTARY CLUBS ─────────────────────────────────────────────
    ("ROTARY CLUBS", None),

    ("Rotary Club of Tyler", "Rotary",
     "Holly Lewis", "Administrator",
     "(903) 561-8056", "Tylerrotaryaa@outlook.com",
     "PO Box 131444, Tyler, TX 75713",
     "https://tylerrotary.org",
     "Meets Thursdays 11:45 AM at Hollytree Country Club. President: William Croom II"),

    ("Tyler Sunrise Rotary Club", "Rotary",
     "Sam Broyles", "President",
     "(903) 266-5606", "tylersunriserotary@gmail.com",
     "PO Box 8383, Tyler, TX 75711",
     "https://tylersunriserotary.org",
     "Meets Thursdays 6:30-7:30 AM at Pollard United Methodist Church"),

    ("South Tyler Rotary Club", "Rotary",
     "Alex Fruth", "President",
     "", "",
     "Tyler, TX",
     "https://southtylerrotary.org",
     "President-Elect: Tiffany Damskov. District 5830, Club #1883"),

    # ── LIONS / KIWANIS ──────────────────────────────────────────
    ("LIONS & KIWANIS CLUBS", None),

    ("Tyler Lions Club #1084", "Lions Club",
     "Tammie Barton", "President",
     "(903) 570-5264", "Quarteraggie@Yahoo.com",
     "Tyler, TX",
     "https://e-clubhouse.org/sites/tyler_tx/index.php",
     "Meets noon, 1st & 3rd Tuesdays at Traditions Restaurant, 6205 S. Broadway. Secretary: Edith Schneider (Ediepanda56@Gmail.com)"),

    ("Kiwanis Club of Tyler-Rose City", "Kiwanis",
     "", "General Contact",
     "(903) 561-6117", "",
     "Tyler, TX 75702",
     "https://www.tylerkiwanis.org",
     "Foundation PO Box 130033, Tyler, TX 75713. Chartered May 10, 1921"),

    # ── CHAMBERS OF COMMERCE ─────────────────────────────────────
    ("CHAMBERS OF COMMERCE", None),

    ("Tyler Area Chamber of Commerce", "Chamber of Commerce",
     "Scott Martinez", "CEO",
     "(903) 592-1661", "chamberinfo@tylertexas.com",
     "315 N Broadway Ave, Suite 100, Tyler, TX 75702",
     "https://tylertexas.com",
     "Toll-free: (800) 235-5712"),

    ("Lindale Area Chamber of Commerce", "Chamber of Commerce",
     "Shelbie Glover", "President & CEO",
     "(903) 882-7181", "shelbie.glover@lindalechamber.org",
     "205 S. Main St., Lindale, TX 75771",
     "https://lindalechamber.org",
     "Staff: Emily Sanchez (emily.sanchez@lindalechamber.org), Lisette Whatley (lisette.whatley@lindalechamber.org)"),

    ("Whitehouse Area Chamber of Commerce", "Chamber of Commerce",
     "", "General Contact",
     "(903) 941-5221", "info@whitehousetx.com",
     "PO Box 1041, Whitehouse, TX 75791",
     "https://whitehousetx.com",
     "Also: whitehouseareachamber@gmail.com"),

    ("Bullard Area Chamber of Commerce", "Chamber of Commerce",
     "", "General Contact",
     "(903) 894-4238", "",
     "114 S Phillips St, Bullard, TX 75757",
     "https://www.bullardchamber.com",
     ""),

    # ── VETERANS ORGANIZATIONS ───────────────────────────────────
    ("VETERANS ORGANIZATIONS", None),

    ("American Legion Post 12, Tyler", "American Legion",
     "", "Post Contact",
     "(903) 593-9751", "amlegion12@gmail.com",
     "5503 American Legion Rd, Tyler, TX 75708",
     "https://americanlegion12tyler.org",
     "Meets 2nd Tuesday at 6:00 PM. Dues: $45"),

    ("VFW Carl Webb Post 1799", "VFW",
     "", "Quartermaster",
     "(903) 561-3501", "Vfw1799qm@gmail.com",
     "14391 Rhones Quarter Road, Tyler, TX 75707",
     "",
     "Mon-Fri 9 AM-5 PM. Meets 2nd Monday"),

    # ── FRATERNAL ORGANIZATIONS ──────────────────────────────────
    ("FRATERNAL ORGANIZATIONS", None),

    ("Tyler Elks Lodge #215", "Elks (BPOE)",
     "", "Lodge Contact",
     "(903) 592-6508", "tylerelks215@yahoo.com",
     "13618 State Highway 31 W, Tyler, TX 75709",
     "https://www.elks.org/lodges/home.cfm?LodgeNumber=215",
     "Open Thu/Fri/Sat from 6 PM. Meets 2nd & 4th Thursdays"),

    ("St. John's Masonic Lodge #53", "Masonic Lodge",
     "", "Lodge Secretary",
     "(903) 597-6413", "",
     "323 W. Front St., Tyler, TX 75702",
     "",
     "Chartered 1849. Building on National Register of Historic Places. Also hosts York Rite, Scottish Rite, Eastern Star"),

    ("Knights of Columbus Council #1502", "Knights of Columbus",
     "Johnny", "Council Contact",
     "(903) 581-2323", "Johnny@kc1502.com",
     "3509 S Southwest Loop 323, Tyler, TX",
     "https://uknight.org/CouncilSite/?CNO=1502",
     "Bishop Charles E. Herzig Council. Est. 1910"),

    ("Rose City Civitan Club #2445", "Civitan",
     "", "",
     "", "",
     "Tyler, TX",
     "",
     "Contact via Civitan International: growth@civitan.org. First group in nation to adopt a highway (US 69, 1985)"),

    # ── WOMEN'S ORGANIZATIONS ────────────────────────────────────
    ("WOMEN'S & VOLUNTEER ORGANIZATIONS", None),

    ("Junior League of Tyler", "Junior League",
     "", "Nominating Chair",
     "(903) 595-5426", "nominating@juniorleagueoftyler.org",
     "1919 S. Donnybrook Ave., Tyler, TX 75701",
     "https://juniorleagueoftyler.org",
     "~215 active members, ~475 sustaining. $7.5M+ given back, 20K+ volunteer hours/yr"),

    ("Women in Tyler (WIT)", "Women's Organization",
     "", "",
     "", "womenintyler@gmail.com",
     "Tyler, TX",
     "https://womenintyler.org",
     ""),

    ("Tyler Executive Women's Network (TEWN)", "Women's Professional Network",
     "", "",
     "", "",
     "Tyler, TX",
     "https://www.tewntyler.com",
     "Meets at Hollytree Country Club"),
]

ws = wb.active
ws.title = "Civic & Service Organizations"

col_widths = [42, 26, 22, 22, 18, 34, 48, 42, 60]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

row_num = 2
for record in groups:
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=record[0])
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 28
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

output = "/home/user/Claude/Smith_County_Civic_Service_Organizations.xlsx"
wb.save(output)
data_rows = sum(1 for r in groups if r[1] is not None)
print(f"Saved {data_rows} contacts to {output}")
