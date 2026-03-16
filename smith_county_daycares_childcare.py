#!/usr/bin/env python3
"""Generate Smith County Daycares & Childcare Centers spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=13, color="1B4332")
section_fill = PatternFill(start_color="B7E4C7", end_color="B7E4C7", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Center", "Type", "Director/Owner", "Phone", "Email", "Address", "Website", "Ages/Capacity", "Notes"]

centers = [
    # ── CHAIN / FRANCHISE ────────────────────────────────────────
    ("CHAIN / FRANCHISE CENTERS", None),

    ("Tyler KinderCare", "Daycare / Preschool",
     "Joan Crone (Director)", "", "",
     "3130 E 5th St, Tyler, TX 75701",
     "https://www.kindercare.com/our-centers/tyler/tx/302125",
     "6 weeks - 12 yrs", "Mon-Fri 6am-6pm"),

    ("Kid City USA — Tyler", "Daycare / Preschool",
     "Tiffany Veasey", "(409) 757-8396", "tyler@kidcityusa.com",
     "2101 Shiloh Road, Tyler, TX",
     "https://kidcityusa.com/locations/texas/tyler",
     "6 weeks - 12 yrs", ""),

    ("Handprints Academy — Tyler", "Daycare / Preschool / After School",
     "Stephanie Wells (Regional Dir); Chad & Elizabeth Elias (Principals)",
     "(903) 593-9410", "",
     "13931 State Hwy 31 W, Tyler, TX",
     "https://handprintschildcare.com/tyler",
     "Infant - school age. Capacity: 345", "Also: (214) 484-1018"),

    # ── INDEPENDENT / LOCAL CENTERS ──────────────────────────────
    ("INDEPENDENT & LOCAL CENTERS", None),

    ("Tyler Day Nursery", "Nonprofit Daycare (est. 1936)",
     "", "(903) 592-4861", "",
     "2901 W Gentry Pkwy, Tyler, TX 75702",
     "https://www.tylerdaynursery.org",
     "6 weeks - 5 yrs", "Oldest licensed nonprofit childcare in TX"),

    ("North Tyler Developmental Academy", "Nonprofit Daycare",
     "Sonja L. Watson (Exec Dir)", "(903) 592-3671", "",
     "3000 N Border Ave, Tyler, TX 75702",
     "https://www.northtylerday.org",
     "6 weeks - 5 yrs (+ after-school)", "TX Rising Star 3-Star. Sliding scale fees"),

    ("Stepping Stone School", "Daycare / Preschool / After School",
     "Camille Brown (Owner/Exec Dir)", "(903) 566-1851", "monica.steppingstone@yahoo.com",
     "5113 Timbercreek Dr, Tyler, TX",
     "https://www.steppingstonetyler.org",
     "K-8th", "Since 1968. Montessori Math classes"),

    ("Creative Kids Learning Center", "Daycare / Preschool / After School",
     "Gary Bender (Co-Owner)", "(903) 253-7273", "",
     "15384 State Hwy 64 East, Tyler, TX 75707",
     "https://creativekidstyler.com",
     "6 weeks - 12 yrs", ""),

    ("Creative Children's Academy", "Daycare / Preschool",
     "Dana Greenwood (Dir/Owner)", "(903) 747-8284", "",
     "Tyler, TX 75701",
     "https://www.creativechildrens-academy.com",
     "8 weeks - 5 yrs. Capacity: 38", "$135-$140/week"),

    ("Tyler Christian Preschool", "Faith-based Daycare / Preschool",
     "Shannon Bedsole (Dir of Operations)", "(903) 534-9987", "",
     "225 Winchester Dr, Tyler, TX 75701",
     "https://www.tylerchristianpreschool.net",
     "6 weeks - 12 yrs. Capacity: 128", "TX Rising Star accredited. Meals included"),

    ("Kidz Depot Learning Academy", "Daycare / Preschool / After School",
     "", "", "",
     "Tyler, TX 75702",
     "https://kidzdepotlearning.com",
     "6 weeks - school age. Capacity: 39+34", "2 locations"),

    ("Bright Beginnings Early Education", "Faith-based Daycare (Harvest Time Church)",
     "", "(903) 253-6250", "",
     "17199 Old Jacksonville Hwy, Flint, TX 75762",
     "https://daycaresoftyler.org",
     "", ""),

    # ── MONTESSORI ───────────────────────────────────────────────
    ("MONTESSORI", None),

    ("Acute Children's Montessori Academy", "Montessori Preschool",
     "", "(903) 526-7084", "mcacutemon@aol.com",
     "1709 E 5th St, Tyler, TX 75701",
     "",
     "18 months - 6 yrs", "Mon-Fri 7am-6pm"),

    ("Oak Hill Montessori / The Leadership Academy", "Montessori / Christian School",
     "", "(903) 561-1002", "",
     "6720 Oak Hill Blvd, Tyler, TX 75703",
     "https://oakhillschooltyler.com",
     "", "7:1 student-teacher ratio"),

    # ── HALF-DAY PRESCHOOL ───────────────────────────────────────
    ("HALF-DAY PRESCHOOL", None),

    ("Oak Tree Academy", "Half-Day Preschool",
     "Sunday Hooper & Cathy Champion (Owners)", "", "",
     "1525 E Grande Blvd, Tyler, TX 75703",
     "https://www.oaktreeacademypreschool.com",
     "15 months - K. Capacity: 119 (175 families)", "MWF, TTH, or M-F 9:30am-1pm (Sept-May)"),

    # ── CHURCH-BASED PRESCHOOLS ──────────────────────────────────
    ("CHURCH-BASED PRESCHOOLS & EARLY EDUCATION", None),

    ("Green Acres Baptist — Early Education Center", "Church Daycare / Preschool",
     "", "(903) 525-1104", "marlac@mail.gabc.org",
     "1607 Troup Hwy, Tyler, TX 75701",
     "https://www.gabc.org/kids",
     "3 weeks - 12 yrs. Capacity: 350", "Founded 1974. After-school pickup from local schools"),

    ("Grace Community School — Early Education", "Christ-centered Early Ed (PK-12 school)",
     "Jay Ferguson (Head of School)", "", "admissions@gracetyler.org",
     "3215 Old Jacksonville Hwy, Tyler, TX 75701",
     "https://www.gracetyler.org",
     "8 weeks - 4 yrs", "Waitlist required ($25 fee). Also Lindale campus MDO"),

    ("Pollard UMC — Kids' Kaleidoscope", "Church Preschool / Daycare",
     "", "(903) 597-2571", "",
     "3030 New Copeland Rd, Tyler, TX 75701",
     "https://www.pollardumc.com",
     "18 months - K", "Play-based learning. 47+ years"),

    # ── MOTHER'S DAY OUT ─────────────────────────────────────────
    ("MOTHER'S DAY OUT PROGRAMS", None),

    ("Green Acres Baptist Church MDO", "Mother's Day Out",
     "", "(903) 525-1104", "",
     "1607 Troup Hwy, Tyler, TX 75701",
     "https://www.gabc.org/kids",
     "Babies - Pre-K", "Mon & Thu 9am-2:30pm. Waitlist resets yearly"),

    ("CrossPointe Community Church PDO", "Parents Day Out",
     "", "(903) 939-2266", "",
     "4642 FM 2813, Tyler, TX 75703",
     "https://www.cptyler.church/pdo",
     "18 months - Pre-K", "Tue & Thu 8:30am-2:30pm. $200/month"),

    ("Southern Oaks Baptist Church MDO", "Mother's Day Out",
     "Renee French (Dir since 2012)", "", "",
     "601 E Amherst Dr, Tyler, TX 75701",
     "https://welovethegospel.com/mothers-day-out",
     "", "Tue & Thu 9:30am-2:30pm"),

    # ── COLLEGE-BASED ────────────────────────────────────────────
    ("COLLEGE-BASED CHILDCARE", None),

    ("TJC Family Learning Center", "College Lab School / Daycare",
     "Kayla Hartweg (Director)", "(903) 510-2200", "",
     "1508 N. Haynie, Tyler, TX 75702",
     "https://tjc.edu/community/family-learning-center",
     "6 weeks - 5 yrs. Capacity: 66", "Collab of Literacy Council, Tyler ISD, TJC. Meals included. 7:15am-5:15pm"),

    # ── HEAD START / PUBLIC PRE-K ────────────────────────────────
    ("HEAD START & PUBLIC PRE-K", None),

    ("Tyler ISD Head Start", "Free Head Start (income-eligible)",
     "", "(903) 262-1000", "",
     "1319 Earl Campbell Pkwy, Tyler, TX 75701",
     "https://www.tylerisd.org/page/head-start",
     "Ages 3-5", "Multiple locations: Kennedy (531-3760), Saunders (531-3765), Patton (531-3770), Border (531-3785)"),

    ("GETCAP Head Start / Early Head Start", "Head Start",
     "", "(903) 962-2670", "",
     "1010 W SW Loop 323, Tyler, TX 75701",
     "https://get-cap.org/head-start",
     "", "Serves Smith County and surrounding"),

    # ── AFTER-SCHOOL PROGRAMS ────────────────────────────────────
    ("AFTER-SCHOOL PROGRAMS", None),

    ("The Mentoring Alliance", "After-School / Summer Camp",
     "", "(903) 593-9211", "",
     "1909 S Broadway Ave, Tyler, TX 75701",
     "https://thementoringalliance.com",
     "", "$24-$59/week. Financial aid available. Multiple elementary schools"),

    ("Tyler After School Program", "After-School Care",
     "", "", "",
     "Tyler, TX",
     "https://tylerafterschool.com",
     "", "Serves Jack, Owens, Rice, University Academy, Brown, Cumberland schools"),

    # ── LINDALE / WHITEHOUSE / BULLARD ───────────────────────────
    ("LINDALE / WHITEHOUSE / BULLARD", None),

    ("New Generation Learning Center (Whitehouse)", "Daycare / Preschool / After School",
     "Mike Dolan (Owner), Allyce Dolan (Director)", "(903) 326-8748", "",
     "14172 TX-110, Whitehouse, TX 75707",
     "https://www.newgenlearningcenter.com",
     "", "Founded 1996. Mon-Fri 7am-6pm"),

    ("Bullard Early Education LLC", "Daycare / Preschool / After School",
     "Lindsey Porter (Exec Dir)", "(903) 894-7222", "bee@suddenlink.net",
     "111 N Rather St, Bullard, TX 75757",
     "",
     "Capacity: 257", "Est. 1997. Mon-Fri 7am-6pm"),

    ("Grace Early Education — Lindale Campus", "Mother's Day Out (2 days/week)",
     "", "", "admissions@gracetyler.org",
     "Lindale, TX",
     "https://www.gracetyler.org",
     "", ""),
]

ws = wb.active
ws.title = "Daycares & Childcare"

col_widths = [42, 32, 40, 18, 32, 46, 48, 32, 56]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

row_num = 2
for record in centers:
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=record[0])
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 28
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

output = "/home/user/Claude/Smith_County_Daycares_Childcare_Centers.xlsx"
wb.save(output)
data_rows = sum(1 for r in centers if r[1] is not None)
print(f"Saved {data_rows} contacts to {output}")
