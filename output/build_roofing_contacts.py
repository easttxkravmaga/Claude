#!/usr/bin/env python3
"""Build Smith County TX Roofing Companies Contact List (.xlsx)
Merged from 3 agent search results, deduplicated."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Master contact list — merged from all 3 agents, deduplicated
# Format: (First Name, Last Name, Phone, Email, Company, Title, City)
contacts = [
    # ===== TYLER =====
    # Stonewater Roofing
    ("Roland", "Browne", "(903) 266-1205", "tyler@stonewaterroofing.com", "Stonewater Roofing", "CEO / Founder", "Tyler"),
    # Cable's Roofing & Construction
    ("Darren", "Cable", "(903) 581-5599", "sales@cablesroofing.com", "Cable's Roofing & Construction", "Co-Owner", "Tyler"),
    ("Macy", "Cable", "(903) 581-5599", "sales@cablesroofing.com", "Cable's Roofing & Construction", "Co-Owner / Office Manager", "Tyler"),
    # Steele Roofing
    ("John", "Steele", "(903) 630-5751", "steeleroofingtyler@gmail.com", "Steele Roofing LLC", "Owner", "Tyler"),
    # Yosemite Roofing
    ("Josiah", "Rosebury", "(903) 326-7255", "", "Yosemite Roofing LLC", "Owner / Founder", "Tyler"),
    # AVCO Roofing
    ("Heath", "Hicks", "(903) 534-8700", "office@avcoroofing.com", "AVCO Roofing", "Owner", "Tyler"),
    ("Ronnie", "Lollar", "(903) 534-8700", "office@avcoroofing.com", "AVCO Roofing", "CEO", "Tyler"),
    ("Bruce", "Webster", "(903) 534-8700", "office@avcoroofing.com", "AVCO Roofing", "COO", "Tyler"),
    ("Emerald", "Bragg", "(903) 534-8700", "office@avcoroofing.com", "AVCO Roofing", "Executive Assistant", "Tyler"),
    # Admire Roofing
    ("Jonathan", "Admire", "(903) 780-7245", "jadmire@admireroofing.com", "Admire Roofing LLC", "Owner", "Tyler"),
    # Dickson Roofing
    ("Shon", "Dickson", "(903) 508-5011", "sales@dicksonroofs.com", "Dickson Roofing", "Owner", "Tyler"),
    ("Talya", "Dickson", "(903) 508-5011", "sales@dicksonroofs.com", "Dickson Roofing", "Director of Marketing", "Tyler"),
    ("Cheri", "Dickson", "(903) 508-5011", "sales@dicksonroofs.com", "Dickson Roofing", "Office Manager", "Tyler"),
    # Tyler Roofing Company Inc.
    ("Tommy Ray", "Sukiennik", "(903) 597-4152", "tylerroofingco@gmail.com", "Tyler Roofing Company Inc.", "President", "Tyler"),
    ("Mark", "Sukiennik", "(903) 597-4152", "MarkSukiennik@suddenlink.net", "Tyler Roofing Company Inc.", "Vice President", "Tyler"),
    ("Cris", "Sukiennik", "(903) 597-4152", "tylerroofingco@gmail.com", "Tyler Roofing Company Inc.", "General Manager", "Tyler"),
    ("Tracy", "Sukiennik", "(903) 597-4152", "tylerroofingco@gmail.com", "Tyler Roofing Company Inc.", "Office Manager", "Tyler"),
    # Cover 3 Roofing
    ("Shane", "Day", "(903) 765-7089", "cover3roofing@gmail.com", "Cover 3 Roofing and Construction", "Co-Owner / CEO", "Tyler"),
    ("Eva", "Day", "(903) 765-7089", "cover3roofing@gmail.com", "Cover 3 Roofing and Construction", "Co-Owner", "Tyler"),
    # Indemnity Roofing
    ("Kyle", "Needham", "(903) 363-9155", "info@indemnityroofing.net", "Indemnity Roofing Inc.", "Co-Owner / Founder", "Tyler"),
    ("Josh", "Hays", "(903) 363-9155", "info@indemnityroofing.net", "Indemnity Roofing Inc.", "Co-Owner", "Tyler"),
    # BXC Roofing
    ("Blake", "Wilabay", "(903) 320-3030", "", "BXC Roofing LLC", "Partner", "Tyler"),
    # Weaver Action Roofing
    ("Robert", "Weaver", "(903) 592-5892", "", "Weaver Action Roofing Company", "Owner / President", "Tyler"),
    ("Janet", "Weaver", "(903) 592-5892", "", "Weaver Action Roofing Company", "Co-Owner", "Tyler"),
    # A&W Hill Roofing
    ("Ashley", "Hill", "(903) 630-4688", "contact@awhill.org", "A&W Hill Roofing & Construction", "Co-Owner", "Tyler"),
    ("William", "Hill", "(903) 630-4688", "contact@awhill.org", "A&W Hill Roofing & Construction", "Co-Owner", "Tyler"),
    # Quick Roofing & Restoration
    ("Andrew", "Quick", "(903) 505-5720", "", "Quick Roofing & Restoration LLC", "Owner", "Tyler"),
    # Hail King Professional Roofing
    ("Landon", "Stokes", "(682) 235-2880", "", "Hail King Professional Roofing LLC", "Co-Owner", "Tyler"),
    ("Lori", "Anderson-Stokes", "(682) 235-2880", "", "Hail King Professional Roofing LLC", "Co-Owner", "Tyler"),
    # ASAP Roofing
    ("Joseph", "Kozel III", "(903) 363-9160", "email@asaproofing.com", "ASAP Roofing", "Owner", "Tyler"),
    ("Tina", "Anders", "(903) 363-9160", "email@asaproofing.com", "ASAP Roofing", "Office Manager", "Tyler"),
    # Hargrove Roofing
    ("Billy", "Hargrove", "(903) 412-2710", "", "Hargrove Roofing & Construction", "CEO", "Tyler"),
    ("Jordan Zachary", "Cagle", "(903) 412-2710", "", "Hargrove Roofing & Construction", "Branch Director", "Tyler"),
    ("Devin", "Angel", "(903) 412-2710", "", "Hargrove Roofing & Construction", "Sales Manager", "Tyler"),
    # Texas Direct Roofing
    ("Matthew Hunter", "Andrews", "(903) 710-4022", "", "Texas Direct Roofing & Construction", "Owner", "Tyler"),
    ("Ryan", "Stephens", "(903) 710-4022", "", "Texas Direct Roofing & Construction", "Director of Marketing", "Tyler"),
    # East Texas Roof Works
    ("Dewayne", "Gordon", "(903) 526-7419", "", "East Texas Roof Works & Sheet Metal", "Owner", "Tyler"),
    ("Barbara", "Gordon", "(903) 526-7419", "", "East Texas Roof Works & Sheet Metal", "Corporate Secretary", "Tyler"),
    # Accurate Roof Systems
    ("Gabriel", "Spencer", "(903) 894-6418", "", "Accurate Roof Systems", "Co-Owner", "Tyler"),
    ("Melissa", "Spencer", "(903) 894-6418", "", "Accurate Roof Systems", "Co-Owner", "Tyler"),
    ("David", "Jimenez", "(903) 894-6418", "", "Accurate Roof Systems", "Director of Marketing", "Tyler"),
    # Smith Roofing and Construction
    ("Chris", "Smith", "(903) 279-7000", "Csmithsrc@gmail.com", "Smith Roofing and Construction", "Owner", "Tyler"),
    # Tyler Roof Repair
    ("Justin Dale", "Deason", "(903) 426-1151", "sales@RoofRepairTyler.com", "Tyler Roof Repair", "Owner", "Tyler"),
    # Bison Construction Pros
    ("Justin", "", "(903) 520-4111", "", "Bison Construction Pros", "Owner", "Tyler"),

    # ===== LINDALE =====
    # Redline Roofing
    ("Myrick Brett", "Prince", "(903) 437-7049", "", "Redline Roofing Company", "Owner", "Lindale"),
    ("Scott", "Prince", "(903) 437-7049", "", "Redline Roofing Company", "Owner (Operations)", "Lindale"),
    # Priority Roofing
    ("Will", "Miller", "(903) 920-6524", "tyler@priorityroofs.com", "Priority Roofing of East Texas", "Owner", "Lindale"),
    ("John", "Faue", "(903) 920-6524", "tyler@priorityroofs.com", "Priority Roofing of East Texas", "General Manager", "Lindale"),
    ("Micah", "McQueen", "(903) 920-6524", "tyler@priorityroofs.com", "Priority Roofing of East Texas", "Vice President", "Lindale"),
    # Advantage Roofing
    ("Jay", "Welling", "(903) 939-3168", "", "Advantage Roofing Company", "President", "Lindale"),
    # Firehouse Roofing
    ("Mark", "Latham", "(903) 817-0219", "", "Firehouse Roofing", "Owner", "Lindale"),
    # Calvary Roofing
    ("Nicholas", "Gullatt", "(903) 714-3729", "", "Calvary Roofing and Construction", "Co-Owner", "Lindale"),
    ("April", "Gullatt", "(903) 714-3729", "", "Calvary Roofing and Construction", "Co-Owner", "Lindale"),
    # Myers Roofing Solutions
    ("Ronnie", "Myers", "(903) 573-3431", "RonMyersRoofing@gmail.com", "Myers Roofing Solutions", "Owner", "Lindale"),

    # ===== WHITEHOUSE =====
    # Steve's Quality Roofing
    ("Steve", "Rohus", "(903) 839-2060", "steverohus@gmail.com", "Steve's Quality Roofing & Construction", "Owner", "Whitehouse"),
    # Project One Roofing
    ("Ryan", "Moore", "(903) 534-3979", "", "Project One Roofing", "Co-Owner / Founder", "Whitehouse"),
    ("Cori", "Moore", "(903) 534-3979", "", "Project One Roofing", "Co-Owner", "Whitehouse"),
    ("Ted", "Long", "(903) 534-3979", "", "Project One Roofing", "Co-Owner", "Whitehouse"),
    ("Chris", "Hughes", "(903) 534-3979", "", "Project One Roofing", "Co-Owner", "Whitehouse"),
    ("Julia", "Hughes", "(903) 534-3979", "", "Project One Roofing", "Insurance Agent", "Whitehouse"),
    ("Meredith", "Bryans", "(903) 534-3979", "", "Project One Roofing", "Marketing Director", "Whitehouse"),
    # Goodfella's Roofing
    ("Dayton", "Longfellow", "(903) 245-2988", "", "Goodfella's Roofing & Construction", "Owner", "Whitehouse"),
    ("Macie", "Longfellow", "(903) 245-2988", "", "Goodfella's Roofing & Construction", "Administrative Director", "Whitehouse"),
    # Coleman Roofing
    ("Kurt", "Coleman", "(903) 530-0350", "kurt@colemanroofs.com", "Coleman Roofing LLC", "Owner", "Whitehouse"),

    # ===== BULLARD =====
    # Boyd's of Texas
    ("John", "Boyd", "(903) 810-7770", "", "Boyd's of Texas Contracting", "Owner / Director of Operations", "Bullard"),
    ("Summer", "Boyd", "(903) 810-7770", "", "Boyd's of Texas Contracting", "Co-Founder / Director of Marketing", "Bullard"),
    ("Christina", "Courville", "(903) 810-7770", "", "Boyd's of Texas Contracting", "Office Manager", "Bullard"),
    ("Danielle", "Sovia", "(903) 810-7770", "", "Boyd's of Texas Contracting", "Office Assistant", "Bullard"),
    # Reliant Roofing
    ("Jason", "Hadley", "(903) 707-3608", "", "Reliant Roofing", "Owner", "Bullard"),

    # ===== FLINT =====
    # Adams Family Roofing
    ("John Frank", "Adams III", "(903) 484-4231", "sales@adamsfamilyroofing.com", "Adams Family Roofing LLC", "Owner", "Flint"),
    # Bad Bear Roofing
    ("Cody", "Savell", "(903) 363-4140", "", "Bad Bear Roofing & Construction", "Owner", "Flint"),
    ("Valeri", "Savell", "(903) 363-4140", "", "Bad Bear Roofing & Construction", "Co-Owner", "Flint"),
    # Guardian Roofing Pros
    ("Richard Shelby", "Henley", "(903) 710-3132", "", "Guardian Roofing Pros", "President / Owner", "Flint"),
    # Roofing ETX
    ("Jonathan", "", "(903) 312-7145", "jonathan@roofingetx.com", "Roofing ETX", "Owner", "Flint"),
    # ValueGuard Roofing
    ("Jonathan", "", "(903) 343-5700", "", "ValueGuard Roofing", "Owner", "Flint"),

    # ===== TROUP =====
    # ETX Roofing & Construction
    ("Adrian", "", "(903) 969-0503", "", "ETX Roofing & Construction", "Owner", "Troup"),

    # ===== WINONA =====
    # J W Roofing & Construction
    ("Jerry", "Whitley", "(903) 371-1730", "", "J W Roofing & Construction", "Owner", "Winona"),
    # Busby's Roofing
    ("Dale", "Busby", "(903) 877-3100", "", "Busby's Roofing", "Owner", "Winona"),

    # ===== NEW CHAPEL HILL =====
    # Premier Roofing of East Texas
    ("Roy", "Drewett", "(903) 724-1511", "", "Premier Roofing of East Texas", "Owner", "New Chapel Hill"),

    # ===== ARP =====
    # Noble Roofing
    ("", "", "(903) 738-4591", "", "Noble Roofing", "", "Arp"),
]

# --- Build .xlsx ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Roofing Contacts Smith Co TX"

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

headers = ["First Name", "Last Name", "Phone", "Email", "Company", "Title", "City"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

data_font = Font(name="Arial", size=10)
data_alignment = Alignment(vertical="center", wrap_text=False)

for row_idx, contact in enumerate(contacts, 2):
    for col_idx, value in enumerate(contact, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

col_widths = {"A": 16, "B": 18, "C": 18, "D": 34, "E": 38, "F": 28, "G": 16}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{len(contacts) + 1}"

output_path = "/home/user/Claude/output/Smith_County_TX_Roofing_Contacts.xlsx"
wb.save(output_path)

# Stats
companies = set(c[4] for c in contacts)
cities = set(c[6] for c in contacts if c[6])
with_email = sum(1 for c in contacts if c[3])
with_phone = sum(1 for c in contacts if c[2])
with_name = sum(1 for c in contacts if c[0])

print(f"Saved {len(contacts)} contacts to {output_path}")
print(f"Companies: {len(companies)}")
print(f"Cities: {sorted(cities)}")
print(f"Contacts with names: {with_name}")
print(f"Contacts with email: {with_email}")
print(f"Contacts with phone: {with_phone}")
