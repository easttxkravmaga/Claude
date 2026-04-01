#!/usr/bin/env python3
"""Build Smith County TX Roofing Companies Contact List (.xlsx)"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

contacts = [
    # --- Tyler-based companies ---
    # Stonewater Roofing
    ("Roland", "Browne", "(903) 266-1205", "tyler@stonewaterroofing.com", "Stonewater Roofing", "CEO / Founder", "Tyler"),
    # Cover 3 Roofing and Construction
    ("Shane", "Day", "(903) 765-7089", "cover3roofing@outlook.com", "Cover 3 Roofing and Construction", "Owner / CEO", "Tyler"),
    ("Eva", "Day", "(903) 765-7089", "", "Cover 3 Roofing and Construction", "Co-Owner", "Tyler"),
    # A&W Hill Roofing & Construction
    ("William", "Hill", "(903) 630-4688", "contact@awhill.org", "A&W Hill Roofing & Construction", "Co-Owner", "Tyler"),
    ("Ashley", "Hill", "(903) 630-4688", "contact@awhill.org", "A&W Hill Roofing & Construction", "Co-Owner", "Tyler"),
    # Adams Family Roofing
    ("John Frank", "Adams III", "(903) 484-4231", "sales@adamsfamilyroofing.com", "Adams Family Roofing", "Owner", "Flint"),
    ("Kassidy", "Adams", "(903) 484-4231", "sales@adamsfamilyroofing.com", "Adams Family Roofing", "Project Inspector", "Flint"),
    # Yosemite Roofing
    ("Josiah", "Rosebury", "(903) 326-7255", "info@yosemiteroofing.com", "Yosemite Roofing", "Owner / Founder", "Tyler"),
    # Priority Roofing
    ("Will", "Miller", "(903) 920-6524", "tyler@priorityroofs.com", "Priority Roofing", "CEO / Founder", "Lindale"),
    ("John", "Faue", "(903) 920-6524", "tyler@priorityroofs.com", "Priority Roofing", "Managing Partner", "Tyler"),
    ("Micah", "McQueen", "(903) 920-6524", "", "Priority Roofing", "Vice President", "Tyler"),
    ("Cody", "Wilcox", "(903) 920-6524", "", "Priority Roofing", "Office Manager", "Tyler"),
    # Redline Roofing
    ("Myrick", "Prince", "(903) 437-7049", "", "Redline Roofing & Construction", "Owner", "Lindale"),
    # Quick Roofing & Restoration
    ("Andrew Allan", "Quick", "(903) 505-5720", "", "Quick Roofing & Restoration", "Owner", "Tyler"),
    # Hargrove Roofing
    ("Billy", "Hargrove", "(903) 747-8206", "", "Hargrove Roofing & Construction", "CEO", "Tyler"),
    ("Clyde", "Hargrove", "(903) 412-2710", "", "Hargrove Roofing & Construction", "Founder", "Tyler"),
    ("Mae", "Hargrove", "(903) 412-2710", "", "Hargrove Roofing & Construction", "COO", "Tyler"),
    ("Jordan Zachary", "Cagle", "(903) 412-2710", "", "Hargrove Roofing & Construction", "Branch Director", "Tyler"),
    ("Devin", "Angel", "(903) 412-2710", "", "Hargrove Roofing & Construction", "Tyler Sales Manager", "Tyler"),
    # Dickson Roofing
    ("Shon", "Dickson", "(903) 508-5011", "sales@dicksonroofs.com", "Dickson Roofing", "Owner", "Tyler"),
    ("Talya", "Dickson", "(903) 508-5011", "sales@dicksonroofs.com", "Dickson Roofing", "Director of Marketing", "Tyler"),
    ("Cheri", "Dickson", "(903) 508-5011", "sales@dicksonroofs.com", "Dickson Roofing", "Office Manager", "Tyler"),
    # Admire Roofing
    ("Jonathan", "Admire", "(903) 780-7245", "jadmire@admireroofing.com", "Admire Roofing LLC", "Owner", "Tyler"),
    # Indemnity Roofing
    ("Kyle", "Needham", "(903) 363-9155", "info@indemnityroofing.net", "Indemnity Roofing Inc.", "Co-Owner / Founder", "Tyler"),
    ("Josh", "Hays", "(903) 363-9155", "info@indemnityroofing.net", "Indemnity Roofing Inc.", "Co-Owner", "Tyler"),
    # AVCO Roofing
    ("Heath", "Hicks", "(903) 534-8700", "office@avcoroofing.com", "AVCO Roofing", "Owner", "Tyler"),
    ("Ronnie", "Lollar", "(903) 534-8700", "office@avcoroofing.com", "AVCO Roofing", "CEO", "Tyler"),
    ("Bruce", "Webster", "(903) 534-8700", "office@avcoroofing.com", "AVCO Roofing", "COO", "Tyler"),
    ("Emerald", "Bragg", "(903) 534-8700", "office@avcoroofing.com", "AVCO Roofing", "Executive Assistant", "Tyler"),
    # ASAP Roofing
    ("Joseph", "Kozel III", "(903) 363-9160", "email@asaproofing.com", "ASAP Roofing", "Owner", "Tyler"),
    ("Klint", "Koppel", "(903) 363-9160", "email@asaproofing.com", "ASAP Roofing", "Owner", "Tyler"),
    ("Tina", "Anders", "(903) 363-9160", "email@asaproofing.com", "ASAP Roofing", "Office Manager", "Tyler"),
    # Steele Roofing
    ("John", "Steele", "(903) 630-5751", "steeleroofingtyler@gmail.com", "Steele Roofing LLC", "Owner", "Tyler"),
    # R.B. Roofing
    ("Ramon", "Balderrama", "(903) 570-5441", "rbroofing@att.net", "R.B. Roofing LLC", "Owner", "Tyler"),
    # Tyler Roofing Company
    ("Tommy", "Sukiennik", "(903) 597-4152", "", "Tyler Roofing Company Inc.", "President", "Tyler"),
    ("Mark", "Sukiennik", "(903) 597-4152", "", "Tyler Roofing Company Inc.", "Vice President", "Tyler"),
    ("Cris", "Sukiennik", "(903) 597-4152", "", "Tyler Roofing Company Inc.", "General Manager", "Tyler"),
    ("Tracy", "Sukiennik", "(903) 597-4152", "", "Tyler Roofing Company Inc.", "Office Manager", "Tyler"),
    # Weaver Action Roofing
    ("Robert", "Weaver", "(903) 592-5892", "", "Weaver Action Roofing Company", "Owner / President", "Tyler"),
    ("Janet", "Weaver", "(903) 592-5892", "", "Weaver Action Roofing Company", "Co-Owner", "Tyler"),
    # East Texas Roof Works & Sheet Metal
    ("Dewayne", "Gordon", "(903) 526-7419", "", "East Texas Roof Works & Sheet Metal", "Owner / Founder", "Tyler"),
    ("Steve", "DeRaad", "(903) 526-7419", "", "East Texas Roof Works & Sheet Metal", "Owner", "Tyler"),
    ("Jimmy", "Gordon", "(903) 526-7419", "", "East Texas Roof Works & Sheet Metal", "Owner", "Tyler"),
    ("Barbara", "Gordon", "(903) 526-7419", "", "East Texas Roof Works & Sheet Metal", "Corporate Secretary", "Tyler"),
    # Evolution Roofing
    ("Michael", "Rhodes", "(903) 258-1210", "", "Evolution Roofing LLC", "Owner", "Tyler"),
    ("Zachary", "Reilly", "(903) 258-1210", "", "Evolution Roofing LLC", "Co-Owner", "Tyler"),
    # BXC Roofing
    ("Blake", "Wilabay", "(903) 320-3030", "bxcroofing@yahoo.com", "BXC Roofing LLC", "Partner", "Tyler"),
    # Texas Roof Masters Tyler
    ("Amelia", "Pratka", "(903) 787-9043", "Info@roofmasterstyler.com", "Texas Roof Masters Tyler LLC", "Owner / Founder", "Tyler"),
    # Accurate Roof Systems (Flint)
    ("Gabriel", "Spencer", "(903) 894-6418", "", "Accurate Roof Systems", "Co-Owner / Founder", "Flint"),
    ("Melissa", "Spencer", "(903) 894-6418", "", "Accurate Roof Systems", "Co-Owner", "Flint"),
    ("David", "Jimenez", "(903) 894-6418", "", "Accurate Roof Systems", "Director of Marketing", "Flint"),
    # CXR Roofing
    ("Zach", "Reilly", "(903) 258-1210", "office.cxrroofing@gmail.com", "CXR Roofing", "Founder", "Tyler"),
    # Roof Care Inc.
    ("Ricardo", "Gamboa", "(903) 399-5403", "contact@roof.care", "Roof Care Inc.", "Owner", "Tyler"),
    ("Michelle", "Gamboa", "(903) 399-5403", "contact@roof.care", "Roof Care Inc.", "Operations Director", "Tyler"),
    # Smith Roofing and Construction
    ("Chris", "Smith", "(903) 279-7000", "", "Smith Roofing and Construction", "Owner", "Tyler"),
    # Hail King Professional Roofing
    ("Landon", "Stokes", "(682) 235-2880", "", "Hail King Professional Roofing LLC", "Co-Owner", "Tyler"),
    ("Lori", "Anderson-Stokes", "(682) 235-2880", "", "Hail King Professional Roofing LLC", "Co-Owner", "Tyler"),
    # Sky High Roofing
    ("Roman", "Skylar", "", "", "Sky High Roofing", "Owner", "Tyler"),
    ("Hugo", "Escobar", "", "", "Sky High Roofing", "COO", "Tyler"),
    # --- Small town companies ---
    # Calvary Roofing (Lindale)
    ("Nick", "", "(903) 714-3729", "", "Calvary Roofing and Construction", "Owner", "Lindale"),
    ("April", "", "(903) 714-3729", "", "Calvary Roofing and Construction", "Co-Owner", "Lindale"),
    # Myers Roofing Solutions (Lindale)
    ("Ronnie", "Myers", "(903) 573-3431", "RonMyersRoofing@gmail.com", "Myers Roofing Solutions", "Owner", "Lindale"),
    # Steve's Quality Roofing (Whitehouse)
    ("Steve", "Rohus", "(903) 839-2060", "steverohus@gmail.com", "Steve's Quality Roofing", "Owner", "Whitehouse"),
    # Project One Roofing (Whitehouse)
    ("Cori", "Moore", "(903) 534-3979", "", "Project One Roofing", "Co-Owner", "Whitehouse"),
    # Boyd's of Texas (Bullard)
    ("John Oliver", "Boyd II", "(903) 810-7770", "", "Boyd's of Texas", "Co-Founder", "Bullard"),
    # Coleman Roofing
    ("Kurt", "Coleman", "(903) 530-0350", "", "Coleman Roofing", "Owner", "Tyler"),
    # Goodfella's Roofing (Whitehouse)
    ("Dayton", "Longfellow", "(903) 245-2988", "", "Goodfella's Roofing", "Owner", "Whitehouse"),
    # Bad Bear Roofing (Flint)
    ("Cody", "Savell", "(903) 363-4140", "", "Bad Bear Roofing & Construction", "Owner", "Flint"),
    # Premier Roofing ETX
    ("Roy", "Drewett", "(903) 724-1511", "", "Premier Roofing ETX", "Owner", "Tyler"),
    # ETX Roofing & Construction (Troup)
    ("", "", "(903) 969-0503", "", "ETX Roofing & Construction", "", "Troup"),
    # Advantage Roofing
    ("", "", "(903) 939-3168", "", "Advantage Roofing", "", "Tyler"),
    # ValueGuard Roofing (Flint)
    ("", "", "(903) 343-5700", "", "ValueGuard Roofing", "", "Flint"),
]

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Roofing Contacts Smith Co TX"

# Header style
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

headers = ["First Name", "Last Name", "Phone", "Email", "Company", "Title", "City"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Data style
data_font = Font(name="Arial", size=10)
data_alignment = Alignment(vertical="center", wrap_text=False)

for row_idx, contact in enumerate(contacts, 2):
    for col_idx, value in enumerate(contact, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

# Auto-fit column widths
col_widths = {"A": 16, "B": 16, "C": 18, "D": 34, "E": 36, "F": 24, "G": 14}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:G{len(contacts) + 1}"

output_path = "/home/user/Claude/output/Smith_County_TX_Roofing_Contacts.xlsx"
wb.save(output_path)
print(f"Saved {len(contacts)} contacts to {output_path}")
print(f"Companies represented: {len(set(c[4] for c in contacts))}")
print(f"Contacts with email: {sum(1 for c in contacts if c[3])}")
print(f"Contacts with names: {sum(1 for c in contacts if c[0])}")
