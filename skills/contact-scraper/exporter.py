"""
exporter.py — Export and deduplication layer
Deduplicates across all results, exports to CSV, JSON, and XLSX.
"""

import csv
import json
import os
from datetime import datetime

from extractor import ContactRecord
from scraper import ScrapeSession


def deduplicate_records(session):
    email_index = {}
    no_email_records = []
    for result in session.results:
        if not result.success or not result.record:
            continue
        record = result.record
        if record.is_empty():
            continue
        if record.emails:
            primary = record.emails[0].lower()
            if primary in email_index:
                email_index[primary].merge(record)
            else:
                email_index[primary] = record
        else:
            no_email_records.append(record)
    return list(email_index.values()) + no_email_records


def _flatten_record(record):
    return {
        "source_url": record.source_url,
        "page_title": record.page_title or "",
        "name":       record.name or "",
        "title":      record.title or "",
        "company":    record.company or "",
        "emails":     "; ".join(record.emails),
        "phones":     "; ".join(record.phones),
        "address":    record.address or "",
        "linkedin":   record.social_links.get("linkedin", ""),
        "twitter":    record.social_links.get("twitter", ""),
        "facebook":   record.social_links.get("facebook", ""),
        "instagram":  record.social_links.get("instagram", ""),
    }


FIELDNAMES = [
    "source_url", "page_title", "name", "title", "company",
    "emails", "phones", "address",
    "linkedin", "twitter", "facebook", "instagram",
]

COLUMN_HEADERS = [
    "Source URL", "Page Title", "Name", "Title", "Company",
    "Emails", "Phones", "Address",
    "LinkedIn", "Twitter", "Facebook", "Instagram",
]

COLUMN_WIDTHS = [40, 30, 20, 20, 25, 35, 18, 30, 40, 30, 35, 30]


def export_csv(session, output_path, deduplicate=True):
    records = deduplicate_records(session) if deduplicate else [
        r.record for r in session.results if r.success and r.record and not r.record.is_empty()
    ]
    if not records:
        return 0
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for record in records:
            writer.writerow(_flatten_record(record))
    return len(records)


def export_json(session, output_path, deduplicate=True):
    records = deduplicate_records(session) if deduplicate else [
        r.record for r in session.results if r.success and r.record and not r.record.is_empty()
    ]
    output = {
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "summary": session.summary(),
        "contacts": [r.to_dict() for r in records],
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    return len(records)


def export_xlsx(session, output_path, deduplicate=True):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    records = deduplicate_records(session) if deduplicate else [
        r.record for r in session.results if r.success and r.record and not r.record.is_empty()
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    header_fill  = PatternFill("solid", fgColor="1A1A1A")
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    row_font     = Font(name="Arial", size=10)
    alt_fill     = PatternFill("solid", fgColor="F5F5F5")
    link_font    = Font(name="Arial", size=10, color="CC0000", underline="single")
    left_align   = Alignment(horizontal="left",   vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(bottom=Side(style="thin", color="DDDDDD"))

    # Header row
    ws.row_dimensions[1].height = 22
    for col_idx, (header, width) in enumerate(zip(COLUMN_HEADERS, COLUMN_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDNAMES))}1"

    link_col_indices = {
        FIELDNAMES.index("source_url"),
        FIELDNAMES.index("linkedin"),
        FIELDNAMES.index("twitter"),
        FIELDNAMES.index("facebook"),
        FIELDNAMES.index("instagram"),
    }

    # Data rows
    for row_idx, record in enumerate(records, start=2):
        flat = _flatten_record(record)
        is_alt = (row_idx % 2 == 0)
        ws.row_dimensions[row_idx].height = 18
        for col_idx, key in enumerate(FIELDNAMES, start=1):
            value = flat[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = left_align
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill
            if (col_idx - 1) in link_col_indices and value and value.startswith("http"):
                cell.hyperlink = value
                cell.font = link_font
            else:
                cell.font = row_font

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    s = session.summary()
    summary_rows = [
        ("CONTACT SCRAPER — EXPORT SUMMARY", ""),
        ("", ""),
        ("Exported at",     datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("URLs processed",  s.get("total_urls", 0)),
        ("Successful",      s.get("successful", 0)),
        ("Failed",          s.get("failed", 0)),
        ("Total emails",    s.get("total_emails", 0)),
        ("Total phones",    s.get("total_phones", 0)),
        ("Unique contacts", len(records)),
        ("Duration (sec)",  s.get("duration_seconds", 0)),
    ]
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 30
    title_font = Font(name="Arial", bold=True, size=13, color="CC0000")
    label_font = Font(name="Arial", bold=True, size=10)
    value_font = Font(name="Arial", size=10)
    for row_idx, (label, value) in enumerate(summary_rows, start=1):
        ws2.row_dimensions[row_idx].height = 20
        ca = ws2.cell(row=row_idx, column=1, value=label)
        cb = ws2.cell(row=row_idx, column=2, value=value)
        ca.font = title_font if row_idx == 1 else label_font
        cb.font = value_font

    wb.save(output_path)
    return len(records)


def export_all(session, base_path, deduplicate=True):
    os.makedirs(os.path.dirname(base_path) if os.path.dirname(base_path) else ".", exist_ok=True)
    csv_path  = base_path + ".csv"
    json_path = base_path + ".json"
    xlsx_path = base_path + ".xlsx"
    return {
        "csv":  {"path": csv_path,  "records": export_csv(session,  csv_path,  deduplicate)},
        "json": {"path": json_path, "records": export_json(session, json_path, deduplicate)},
        "xlsx": {"path": xlsx_path, "records": export_xlsx(session, xlsx_path, deduplicate)},
    }

# Backward compat alias
def export_both(session, base_path, deduplicate=True):
    return export_all(session, base_path, deduplicate)


def print_summary(session):
    s = session.summary()
    records = deduplicate_records(session)
    print("\n" + "═" * 50)
    print("  CONTACT SCRAPER — SESSION SUMMARY")
    print("═" * 50)
    print(f"  URLs processed : {s['total_urls']}")
    print(f"  Successful     : {s['successful']}")
    print(f"  Failed         : {s['failed']}")
    print(f"  Emails found   : {s['total_emails']}")
    print(f"  Phones found   : {s['total_phones']}")
    print(f"  Unique contacts: {len(records)}")
    print(f"  Duration       : {s['duration_seconds']}s")
    print("═" * 50)
    if records:
        print("\n  SAMPLE CONTACTS:")
        for r in records[:5]:
            print(f"  • {r.emails[0] if r.emails else r.phones[0] if r.phones else '?'}"
                  f"  |  {r.company or r.page_title or r.source_url[:40]}")
    print()
