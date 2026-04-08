#!/usr/bin/env python3
"""Build .xlsx of documented realtor attacks (2001-present)."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Realtor Attacks 2001-Present"

# --- Column headers ---
headers = [
    "Year",
    "Victim Name",
    "Victim Age",
    "Type of Attack",
    "Outcome (Victim)",
    "Location (City, State)",
    "Suspect(s)",
    "Suspect Outcome",
    "Source / Article Link",
]

# Header styling
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# --- Data rows: one line per occurrence ---
data = [
    [
        2001,
        "Mike Emert",
        40,
        "Stabbed (19 times) / Murder",
        "Killed",
        "Woodinville, WA",
        "Gary Krueger (former police officer)",
        "Linked via DNA in 2011; died before arrest (drowned)",
        "https://unresolved.me/mike-emert",
    ],
    [
        2003,
        "Cynthia 'Cyndi' Williams",
        33,
        "Strangled / Shot in head / Murder",
        "Killed",
        "Powder Springs, GA (Cobb County)",
        "Stacey Ian Humphries",
        "Death sentence (2007); execution date set",
        "https://www.oxygen.com/the-real-murders-of-atlanta/crime-news/cyndi-williams-lori-brown-shot-death-stacey-humphries",
    ],
    [
        2003,
        "Lori Brown",
        21,
        "Shot in head / Murder",
        "Killed",
        "Powder Springs, GA (Cobb County)",
        "Stacey Ian Humphries",
        "Death sentence (2007); execution date set",
        "https://www.oxygen.com/the-real-murders-of-atlanta/crime-news/cyndi-williams-lori-brown-shot-death-stacey-humphries",
    ],
    [
        2006,
        "Sarah Anne Walker",
        40,
        "Stabbed (27+ times) / Beaten / Bitten / Murder",
        "Killed",
        "McKinney, TX (Dallas area)",
        "Kosoul Chanthakoummane",
        "Executed (August 2022)",
        "https://www.fox4news.com/news/texas-to-execute-man-for-slaying-of-dallas-real-estate-agent",
    ],
    [
        2008,
        "Lindsay Buziak",
        24,
        "Stabbed / Murder",
        "Killed",
        "Saanich, British Columbia, Canada",
        "Unknown (lured by couple posing as buyers)",
        "Unsolved",
        "https://en.wikipedia.org/wiki/Murder_of_Lindsay_Buziak",
    ],
    [
        2008,
        "Troy VanderStelt",
        None,
        "Shot (point blank) / Murder",
        "Killed",
        "Michigan",
        "Robert Johnson (73-year-old client)",
        "Convicted of murder",
        "https://stories.avvo.com/crime/murder/7-heinous-cases-of-violence-against-real-estate-agents.html",
    ],
    [
        2009,
        "Unnamed Female Agent",
        None,
        "Raped / Kidnapped",
        "Survived",
        "Cary, NC",
        "Michael Edward Sleeman",
        "Indicted on attempted first-degree rape, kidnapping",
        "https://www.wral.com/news/local/story/5952316/",
    ],
    [
        2010,
        "Brenda Wilburn",
        None,
        "Bound / Murdered / Robbed",
        "Killed",
        "Pulaski, TN",
        "Robert Wayne Garner",
        "Arrested and charged",
        "https://activerain.com/blogsview/1486453/arrest-made-in-murder-of-tennessee-real-estate-broker",
    ],
    [
        2010,
        "Vivian Martin",
        67,
        "Strangled / Arson to cover crime / Murder",
        "Killed",
        "Youngstown, OH",
        "Robert Brooks & Grant Cooper",
        "Both sentenced to life in prison",
        "https://www.wfmj.com/story/26681295/life-in-prison-for-man-who-murdered-youngstown-real-estate-agent",
    ],
    [
        2010,
        "Andrew VonStein",
        51,
        "Shot / Murder",
        "Killed",
        "Kent, OH (Twin Lakes / Portage County)",
        "Robert W. Grigelaitis",
        "Pleaded guilty; life in prison without parole",
        "https://www.nbcnews.com/id/wbna39418027",
    ],
    [
        2011,
        "Ashley Okland",
        27,
        "Shot (twice) / Murder",
        "Killed",
        "West Des Moines, IA",
        "Kristin Ramsey (arrested 2026)",
        "Indicted for first-degree murder (2026)",
        "https://www.nbcnews.com/news/us-news/arrest-made-death-iowa-real-estate-agent-was-killed-open-house-2011-rcna264226",
    ],
    [
        2013,
        "Vernon Holbrook",
        None,
        "Beaten / Throat slashed / Murder-for-hire",
        "Survived initially; died 8 months later from brain damage",
        "Cowiche, WA",
        "Daniel Blizzard (hired attackers for $10K)",
        "Convicted; accomplices also charged",
        "https://www.oxygen.com/mastermind-of-murder/crime-news/daniel-blizzard-hires-couple-to-kill-fellow-realtor-vernon-holbrook",
    ],
    [
        2014,
        "Beverly Carter",
        50,
        "Kidnapped / Suffocated (tape over face) / Murder",
        "Killed",
        "Scott, AR (Little Rock area)",
        "Arron Lewis & Crystal Lowery",
        "Lewis: 2 life sentences; Lowery: 30 years",
        "https://beverlycarterfoundation.org/beverly-s-story",
    ],
    [
        2015,
        "Unnamed Female Agent (69 yrs old)",
        69,
        "Attempted sexual assault / Kidnapping",
        "Survived (escaped)",
        "Whitfield Estates, Manatee County, FL",
        "Bruce Anthony Kotter (used alias 'Jim')",
        "Arrested; charged with kidnapping and attempted sexual battery",
        "https://www.fox13news.com/news/female-realtor-attacked-while-showing-manatee-home",
    ],
    [
        2016,
        "Jim Olsen",
        None,
        "Pistol-whipped / Robbed at gunpoint",
        "Survived (injured)",
        "Milwaukee, WI (15th & Burleigh)",
        "Two masked men (lured by fake tenant)",
        "Investigation; property stolen (wallet, ring, phone)",
        "https://www.housingwire.com/articles/36971-ambushed-realtor-shares-graphic-photo-of-vicious-attack/",
    ],
    [
        2016,
        "Ryan Vega",
        37,
        "Stabbed in neck",
        "Survived (non-life-threatening injuries)",
        "North Las Vegas, NV",
        "Daniel Mora (tenant's son)",
        "Arrested",
        "https://www.reviewjournal.com/crime/stabbings/real-estate-agent-stabbed-while-showing-north-las-vegas-home/",
    ],
    [
        2018,
        "Steven B. Wilson",
        33,
        "Shot / Robbed / Murder",
        "Killed",
        "Hanover, MD (Anne Arundel County)",
        "Dillon Augustyniak (18)",
        "Pleaded guilty; life sentence",
        "https://www.cbsnews.com/baltimore/news/dillion-augustyniak-gets-life-sentence-for-murder-of-real-estate-rep-steven-wilson-inside-hanover-model-home/",
    ],
    [
        2019,
        "Unnamed Female Agent",
        None,
        "Attacked with stun gun / Attempted sexual assault",
        "Survived (escaped)",
        "Tinley Park, IL",
        "Stanley Keller (50, ex-con)",
        "Arrested; $1M bail; charged with agg. battery, attempted assault",
        "https://abc7chicago.com/open-house-stun-gun-tinley-park-stanley-keller/5239939/",
    ],
    [
        2019,
        "Unnamed Female Agent",
        None,
        "Attacked / Groped / Sexually assaulted at open house",
        "Survived",
        "Encino, CA (Los Angeles)",
        "Alen Karaboghosian (45)",
        "Arrested; charged with assault w/ deadly weapon, intent to commit rape, 4 counts sexual battery",
        "https://www.nbcnews.com/news/us-news/man-arrested-after-video-shows-attack-los-angeles-real-estate-n1058526",
    ],
    [
        2019,
        "Monique Baugh",
        27,
        "Kidnapped / Shot (multiple times) / Murder",
        "Killed",
        "Minneapolis, MN (lured to Maple Grove)",
        "Lyndon Wiggins, Cedric Berry, Berry Davis, Elsa Segura",
        "Wiggins: life without parole; Berry & Davis: life without parole; Segura: 20 years",
        "https://ccxmedia.org/news/man-sentenced-to-life-in-prison-for-maple-grove-realtor-murder/",
    ],
    [
        2020,
        "Lenora Farrington",
        None,
        "Hit in head with wrench / Multiple skull fractures",
        "Survived (still active as realtor)",
        "Bedford County, VA (Smith Mountain Lake)",
        "Dustin Holdren",
        "Sentenced to 75 years (50 active); aggravated malicious wounding",
        "https://www.wsls.com/news/local/2023/09/27/man-sentenced-after-attacking-bedford-county-realtor-with-wrench-during-open-house-in-2020/",
    ],
    [
        2021,
        "Sara Trost",
        40,
        "Shot (multiple times) / Murder",
        "Killed",
        "Coral Springs, FL",
        "Raymond Wesley Reese (51, evicted tenant)",
        "Arrested; charged with first-degree murder",
        "https://www.local10.com/news/local/2021/12/24/real-estate-agent-killed-outside-home-in-coral-springs/",
    ],
    [
        2022,
        "Unnamed Female Agent (pregnant)",
        None,
        "Assaulted with weapon / Attempted kidnapping",
        "Survived; miscarried 3 days later",
        "Tucson, AZ",
        "Donasti Davonsiea (a.k.a. Juan Nunley Jr., 37)",
        "28 years in prison; guilty of agg. assault, kidnapping, burglary",
        "https://www.kold.com/2024/02/23/man-sentenced-attack-pregnant-tucson-real-estate-agent/",
    ],
    [
        2022,
        "Unnamed Female Agent",
        None,
        "Attacked / Attempted sexual assault / Kidnapping",
        "Survived",
        "Las Vegas, NV",
        "Richy Cervantes",
        "Charged with home invasion, attempted sexual assault, kidnapping, battery",
        "https://www.fox5vegas.com/2022/07/15/las-vegas-real-estate-agent-attacked-model-home-police-say/",
    ],
    [
        2024,
        "Whitney Hurd",
        32,
        "Stabbed / Robbed / Murder",
        "Killed",
        "Charlotte, NC (SouthPark)",
        "Brandon Braxton (33, former classmate)",
        "Confessed; charged with first-degree murder and robbery",
        "https://www.wbtv.com/2025/03/21/man-confesses-stabbing-death-charlotte-realtor-district-attorney-says/",
    ],
    [
        2025,
        "Unnamed Male Agent",
        None,
        "Assaulted / Kidnapped / Robbed during showing",
        "Survived (showed up at sheriff's office 'bloody and bruised')",
        "Grayson County, KY (Falls of the Rough)",
        "Keith S. Rowe (71)",
        "Charged with assault, kidnapping, robbery",
        "https://www.wave3.com/2025/06/20/suspect-charged-with-assault-realtor-during-home-showing/",
    ],
]

# Data styling
data_font = Font(name="Arial", size=10)
data_align = Alignment(vertical="center", wrap_text=True)
link_font = Font(name="Arial", size=10, color="0563C1", underline="single")

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = data_align
        if col_idx == 9:  # Link column
            cell.font = link_font
            if value:
                cell.hyperlink = value
        else:
            cell.font = data_font

# Column widths
col_widths = {
    "A": 8,   # Year
    "B": 28,  # Victim Name
    "C": 12,  # Age
    "D": 40,  # Type of Attack
    "E": 35,  # Outcome
    "F": 35,  # Location
    "G": 38,  # Suspect(s)
    "H": 50,  # Suspect Outcome
    "I": 70,  # Source Link
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:I{len(data) + 1}"

# --- Summary sheet ---
ws2 = wb.create_sheet("Industry Statistics")
ws2.sheet_properties.tabColor = "CC0000"

stats_headers = ["Statistic", "Value", "Source"]
for col_idx, h in enumerate(stats_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

stats_data = [
    [
        "Real estate professionals killed by violence on the job (2022)",
        23,
        "U.S. Bureau of Labor Statistics",
    ],
    [
        "Average deaths per year in real estate/rental/leasing (2003-2009)",
        75,
        "U.S. Bureau of Labor Statistics",
    ],
    [
        "Murders of real estate professionals since 2010",
        83,
        "U.S. Bureau of Labor Statistics",
    ],
    [
        "Non-fatal assaults on real estate professionals since 2010",
        940,
        "U.S. Bureau of Labor Statistics",
    ],
    [
        "Real estate professionals died on the job (2017)",
        48,
        "U.S. Dept. of Labor / BLS",
    ],
    [
        "REALTORS who were victims of crime while on duty (2022 survey)",
        "~30,000 (2% of 1.5M agents)",
        "NAR 2022 Safety Survey",
    ],
    [
        "REALTORS who were victims of crime while on duty (2023 survey)",
        "~56,000",
        "NAR 2023 Safety Survey",
    ],
    [
        "REALTORS who felt fear for safety on the job (2023 survey)",
        "~322,000",
        "NAR 2023 Safety Survey",
    ],
    [
        "Agents who feared for personal safety during showings (2021)",
        "41%",
        "NAR 2021 Member Safety Report",
    ],
    [
        "Percentage of assault victims who are women",
        "70%",
        "Industry reports",
    ],
    [
        "Percentage of attacks involving guns",
        "50%",
        "Industry reports",
    ],
    [
        "Agents who said they'd been attacked or threatened at work (2018)",
        "9%",
        "Inman 2018 Survey",
    ],
]

for row_idx, row_data in enumerate(stats_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border

ws2.column_dimensions["A"].width = 60
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 35
ws2.freeze_panes = "A2"

output_path = "/home/user/Claude/output/realtor-attacks-research.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Total individual incidents documented: {len(data)}")
