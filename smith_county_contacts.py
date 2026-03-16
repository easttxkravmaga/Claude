#!/usr/bin/env python3
"""Generate Smith County TX Emergency Services Contacts spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=12, color="1F3864")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Organization", "Type", "Name", "Title", "Phone", "Email", "Address", "Website"]

# ── Data ────────────────────────────────────────────────────────────────
contacts = [
    # ── LAW ENFORCEMENT ──
    ("LAW ENFORCEMENT", None, None, None, None, None, None, None),

    ("Smith County Sheriff's Office", "Law Enforcement",
     "Larry R. Smith", "Sheriff",
     "(903) 590-2661", "",
     "227 N. Spring Ave., Tyler, TX 75702",
     "https://www.smithcountysheriff.com"),

    ("Smith County Sheriff's Office - Dispatch", "Law Enforcement",
     "", "Non-Emergency Dispatch",
     "(903) 566-6600", "",
     "227 N. Spring Ave., Tyler, TX 75702",
     "https://www.smithcountysheriff.com"),

    ("Smith County Sheriff's Office - Jail", "Law Enforcement",
     "", "Jail Administration",
     "(903) 590-2800", "",
     "227 N. Spring Ave., Tyler, TX 75702",
     "https://www.smithcountysheriff.com"),

    ("Tyler Police Department", "Law Enforcement",
     "Jimmy Toler", "Chief of Police",
     "(903) 531-1015", "",
     "711 W. Ferguson, Tyler, TX 75702",
     "https://www.cityoftyler.org/government/departments/police-department"),

    ("Tyler Police Department - Patrol", "Law Enforcement",
     "", "Patrol Division",
     "(903) 526-7578", "",
     "711 W. Ferguson, Tyler, TX 75702",
     "https://www.cityoftyler.org/government/departments/police-department"),

    ("Tyler Police Department - Records", "Law Enforcement",
     "", "Records/FOIA",
     "(903) 531-1072", "",
     "711 W. Ferguson, Tyler, TX 75702",
     "https://www.cityoftyler.org/government/departments/police-department"),

    ("Tyler Police Department - Non-Emergency", "Law Enforcement",
     "", "Non-Emergency Line",
     "(903) 531-1000", "",
     "711 W. Ferguson, Tyler, TX 75702",
     "https://www.cityoftyler.org/government/departments/police-department"),

    ("Smith County Constable Pct. 1", "Law Enforcement",
     "Ralph Caraway Jr.", "Constable",
     "(903) 590-2611", "",
     "308 E. Ferguson, Tyler, TX 75702",
     "https://www.smith-county.com/government/elected-officials/constables"),

    ("Smith County Constable Pct. 2", "Law Enforcement",
     "Wayne Allen", "Constable",
     "(903) 590-4840", "",
     "15405 Highway 155 South, Tyler, TX 75703",
     "https://www.smith-county.com/government/elected-officials/constables"),

    ("Smith County Constable Pct. 3", "Law Enforcement",
     "Constable Blackmon", "Constable",
     "(903) 590-4745", "",
     "313 E. Duval Street, Troup, TX 75789",
     "https://www.smith-county.com/government/elected-officials/constables"),

    ("Smith County Constable Pct. 4", "Law Enforcement",
     "Josh Joplin", "Constable",
     "(903) 590-4871", "",
     "14152 Hwy 155 N, Winona, TX 75792",
     "https://www.smith-county.com/government/elected-officials/constables"),

    ("Smith County Constable Pct. 5", "Law Enforcement",
     "Wesley Hicks", "Constable",
     "(903) 590-4901", "",
     "2616 S. Main, P.O. Box 609, Lindale, TX 75771",
     "https://www.smith-county.com/government/elected-officials/constables"),

    ("Smith County District Attorney's Office", "Law Enforcement",
     "Jacob Putman", "Criminal District Attorney",
     "(903) 590-4600", "",
     "200 E. Ferguson, Suite 100, Tyler, TX 75702",
     "https://www.smith-county.com/government/elected-officials/district-attorney"),

    ("Lindale Police Department", "Law Enforcement",
     "Dan Somes", "Chief of Police",
     "(903) 882-3313", "dans@lindaletx.gov",
     "105 Ballard Drive, Lindale, TX 75771",
     "https://www.lindaletx.gov/159/Police-Department"),

    ("Lindale Police Department", "Law Enforcement",
     "Brent Chambers", "Captain",
     "(903) 882-3313", "brentc@lindaletx.gov",
     "105 Ballard Drive, Lindale, TX 75771",
     "https://www.lindaletx.gov/159/Police-Department"),

    ("Whitehouse Police Department", "Law Enforcement",
     "", "Police Department",
     "(903) 510-7550", "",
     "101 A Bascom Road, Whitehouse, TX 75791",
     "https://www.whitehousetx.org/165/Police-Department"),

    ("Bullard Police Department", "Law Enforcement",
     "Jeff Bragg", "Chief of Police",
     "(903) 894-7788", "",
     "P.O. Box 109, Bullard, TX 75757",
     "https://www.bullardtexas.net/308/Bullard-Police-Department"),

    ("Texas DPS - Tyler Office", "Law Enforcement",
     "", "Highway Patrol / DPS",
     "(903) 939-6000", "",
     "4700 University Blvd, Tyler, TX 75707",
     "https://www.dps.texas.gov"),

    ("Texas DPS - Tyler Recruiter", "Law Enforcement",
     "Cpl. Deja Moore", "DPS Recruiter",
     "(430) 247-3112", "",
     "4700 University Blvd, Tyler, TX 75707",
     "https://www.dps.texas.gov/tle/contact"),

    # ── FIRE ──
    ("FIRE DEPARTMENTS", None, None, None, None, None, None, None),

    ("Tyler Fire Department", "Fire",
     "David L. Coble", "Fire Chief / Managing Dir. of Emergency Services",
     "(903) 535-0005", "firechief@tylertexas.com",
     "1520 West Front St., Tyler, TX 75702",
     "https://www.cityoftyler.org/government/departments/fire-department"),

    ("Tyler Fire Department", "Fire",
     "Kelly Adkinson", "Assistant Chief of Operations",
     "(903) 535-0005", "",
     "1520 West Front St., Tyler, TX 75702",
     "https://www.cityoftyler.org/government/departments/fire-department"),

    ("Tyler Fire Department", "Fire",
     "Michael Frost", "Assistant Chief of Administration",
     "(903) 535-0005", "",
     "1520 West Front St., Tyler, TX 75702",
     "https://www.cityoftyler.org/government/departments/fire-department"),

    ("Smith County Fire Marshal's Office", "Fire",
     "Chad Hogue", "Fire Marshal",
     "(903) 590-2650", "",
     "11325 Spur 248, Tyler, TX 75707",
     "https://www.smith-county.com/210/Fire-Marshal"),

    ("Smith County Fire Marshal's Office", "Fire",
     "Brandon Moore", "Emergency Management Coordinator",
     "(903) 590-2649", "",
     "11325 Spur 248, Tyler, TX 75707",
     "https://www.smith-county.com/212/Emergency-Management"),

    ("Smith County Fire Marshal's Office", "Fire",
     "", "General Line",
     "(903) 590-2655", "",
     "11325 Spur 248, Tyler, TX 75707",
     "https://www.smith-county.com/210/Fire-Marshal"),

    ("Smith County ESD2 - Administration", "Fire",
     "Jeffrey Smith", "Assistant Chief of Administration",
     "(903) 617-6578", "",
     "PO Box 780, Whitehouse, TX 75791 / 14128 Hwy 110 S, Whitehouse, TX 75791",
     "https://smithcountyfire.org/scesd2/"),

    ("Smith County ESD1 (Lindale)", "Fire",
     "", "Emergency Services District 1",
     "(903) 882-6492", "",
     "P.O. Box 697, Lindale, TX 75771",
     "https://www.safe-d.org/smith-county-emergency-services-district-no-1/"),

    ("Lindale Fire Department", "Fire",
     "Jeremy LaRue", "Fire Chief",
     "(903) 881-9448", "jlarue@lindalevfd.com",
     "208 E Hubbard St, Lindale, TX 75771",
     "http://www.lindalevfd.com"),

    ("Whitehouse Fire Department (ESD2)", "Fire",
     "Gene Champion", "Fire Chief",
     "(903) 839-2884", "",
     "Whitehouse, TX 75791",
     "https://smithcountyfire.org/scesd2/index.php/ova_dep/whitehouse/"),

    ("Bullard Fire Department (ESD2)", "Fire",
     "Terry DeGuentz", "Fire Chief",
     "(903) 894-7143", "",
     "215 S Houston St, Bullard, TX 75757",
     "https://www.bullardvfd.org"),

    ("Bullard Fire Department (ESD2)", "Fire",
     "Greg Lugo", "Assistant Fire Chief",
     "(903) 894-7143", "",
     "215 S Houston St, Bullard, TX 75757",
     "https://www.bullardvfd.org"),

    ("Arp Volunteer Fire Department (ESD2)", "Fire",
     "Mack Arnold", "Fire Chief",
     "(903) 859-5232", "",
     "201 Arnold St, Arp, TX 75750",
     "https://smithcountyfire.org/scesd2/index.php/ova_dep/arp/"),

    ("Troup Volunteer Fire Department (ESD2)", "Fire",
     "Tim Mager", "Fire Chief",
     "(903) 842-0184", "",
     "204 E Wilkinson Dr, Troup, TX 75789",
     "https://smithcountyfire.org/scesd2/index.php/ova_dep/troup/"),

    ("Noonday Volunteer Fire Department (ESD2)", "Fire",
     "Mel Harper", "Fire Chief",
     "(903) 561-1170", "",
     "16619 HWY 155 S, Tyler, TX 75703",
     "https://smithcountyfire.org/scesd2/index.php/ova_dep/noonday/"),

    ("Noonday Volunteer Fire Department (ESD2)", "Fire",
     "Travis Barnett", "Assistant Fire Chief",
     "(903) 561-1170", "",
     "16619 HWY 155 S, Tyler, TX 75703",
     "https://smithcountyfire.org/scesd2/index.php/ova_dep/noonday/"),

    ("Winona Volunteer Fire Department (ESD2)", "Fire",
     "William E. Allen", "District Chief",
     "(903) 877-3711", "",
     "1111 Dallas St, Winona, TX 75792",
     "https://smithcountyfire.org/scesd2/index.php/ova_dep/winona/"),

    ("Chapel Hill Volunteer Fire Department (ESD2)", "Fire",
     "", "Fire Department",
     "(903) 566-3890", "",
     "13801 County Rd 220, Tyler, TX 75707",
     "https://smithcountyfire.org/scesd2/"),

    ("Flint-Gresham Volunteer Fire Department (ESD2)", "Fire",
     "", "Fire Department",
     "(903) 280-7436", "",
     "Flint, TX (Smith County)",
     "https://www.fgfd.org"),

    ("Jackson Heights Volunteer Fire Department (ESD2)", "Fire",
     "", "Fire Department",
     "(903) 566-8777", "",
     "Smith County, TX",
     "https://smithcountyfire.org/scesd2/"),

    ("Red Springs Volunteer Fire Department (ESD2)", "Fire",
     "", "Fire Department",
     "(903) 858-2412", "",
     "Smith County, TX",
     "https://smithcountyfire.org/scesd2/"),

    # ── EMS ──
    ("EMS / EMERGENCY MEDICAL SERVICES", None, None, None, None, None, None, None),

    ("UT Health East Texas EMS", "EMS",
     "", "EMS / Ambulance Service",
     "(903) 535-5800", "",
     "325 Glenwood Blvd., Tyler, TX 75701",
     "https://uthealtheasttexas.com/patients-and-visitors/emsair1/"),

    ("UT Health East Texas EMS - Membership", "EMS",
     "", "EMS Membership Line",
     "1-800-642-5646", "",
     "P.O. Box 6968, Tyler, TX 75711",
     "https://uthealtheasttexas.com/patients-and-visitors/ems/"),

    ("UT Health East Texas AIR 1", "EMS",
     "", "Air Ambulance / Helicopter EMS",
     "(903) 535-5800", "",
     "Tyler, TX",
     "https://uthealtheasttexas.com/patients-and-visitors/emsair1/"),

    # ── 911 / EMERGENCY COMMUNICATIONS ──
    ("911 / EMERGENCY COMMUNICATIONS", None, None, None, None, None, None, None),

    ("Smith County 9-1-1 Emergency Communications District", "911",
     "", "911 District Administration",
     "(903) 566-8911", "",
     "205 Shelly Drive, Tyler, TX 75701",
     "https://www.911district.com"),
]


# ── Build worksheet ─────────────────────────────────────────────────────
ws = wb.active
ws.title = "Smith County TX Contacts"

# Column widths
col_widths = [42, 18, 24, 38, 20, 28, 52, 55]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Write headers
for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

# Write data
row_num = 2
for record in contacts:
    org = record[0]
    # Section header rows
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=org)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 26
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:H{row_num - 1}"

# ── Save ────────────────────────────────────────────────────────────────
output_path = "/home/user/Claude/Smith_County_TX_Contacts.xlsx"
wb.save(output_path)
print(f"Saved {row_num - 2} rows to {output_path}")
