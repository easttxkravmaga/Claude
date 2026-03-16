#!/usr/bin/env python3
"""Generate Smith County Employers / HR Contacts spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=13, color="1B4332")
section_fill = PatternFill(start_color="B7E4C7", end_color="B7E4C7", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Company", "Industry", "HR Contact", "Title", "Phone", "Email", "Address", "Website", "Employees", "Notes"]

employers = [
    # ── HEALTHCARE ───────────────────────────────────────────────
    ("HEALTHCARE", None),

    ("UT Health East Texas", "Healthcare / Hospital System",
     "R. Joey Saxon", "Director of Human Resources",
     "(903) 877-7739", "humanresources@uttyler.edu",
     "1000 S. Beckham Ave, Tyler, TX 75701",
     "https://uthealtheasttexas.com",
     "4,000+",
     "Also: L. Tyler Perkins, Lead HR Generalist (903-877-7622). Hospital: (903) 597-0351"),

    ("CHRISTUS Trinity Mother Frances", "Healthcare / Hospital System",
     "", "HR Strategy Office",
     "(844) 257-6925", "communications@christushealth.org",
     "727 E Front St, Tyler, TX 75702",
     "https://www.christushealth.org",
     "4,000+",
     "Hospital: 800 E Dawson St. Main: (903) 593-8441. HR office Mon-Fri 8am-5pm"),

    ("Hospice of East Texas", "Healthcare / Hospice",
     "", "General Contact",
     "(903) 266-3400", "",
     "4111 University Blvd, Tyler, TX 75701",
     "https://hospiceofeasttexas.org",
     "",
     "Est. 1982. Pathways program: (903) 565-9230"),

    # ── EDUCATION ────────────────────────────────────────────────
    ("EDUCATION", None),

    ("Tyler ISD", "Public Education (K-12)",
     "Sharon Roy", "Chief Human Resources Officer",
     "(903) 262-1000", "",
     "1319 Earl Campbell Pkwy, Tyler, TX 75701",
     "https://www.tylerisd.org/page/human-resources",
     "3,000+",
     "HR Benefits Director: Alicia Yarbrough"),

    ("Tyler Junior College (TJC)", "Higher Education",
     "Stephanie Cecil", "Chief HR Officer",
     "(903) 510-2200", "",
     "1400 E 5th St, Tyler, TX 75701 (PO Box 9020, 75711)",
     "https://tjc.edu/faculty-staff/human-resources",
     "600+ staff",
     "Toll-free: (800) 687-5680. 13,000 students"),

    ("University of Texas at Tyler", "Higher Education",
     "Gracy Buentello", "Chief HR Officer",
     "(903) 566-7234", "humanresources@uttyler.edu",
     "STE 108, 3900 University Blvd, Tyler, TX 75799",
     "https://www.uttyler.edu/offices/human-resources",
     "1,000+",
     "Also: Monica Sanchez (903-566-7358), Michael Casias (903-877-7706), Allison McCrary (903-565-5647). benefits@uttyler.edu"),

    # ── MANUFACTURING ────────────────────────────────────────────
    ("MANUFACTURING", None),

    ("Tyler Pipe / McWane Inc", "Manufacturing (Iron Pipe & Fittings)",
     "Micayla Helms", "Human Resources Manager",
     "(800) 527-8478", "typ.hr@tylerpipe.com",
     "11910 County Road 492, Tyler, TX 75706",
     "https://www.tylerpipe.com",
     "~210",
     "Direct HR email. Strong workplace safety prospect (foundry)"),

    ("Trane Technologies (Tyler Plant)", "HVAC Manufacturing",
     "", "HR Department",
     "(903) 581-3200", "",
     "6200 Troup Highway, Tyler, TX 75707",
     "https://www.tranetechnologies.com",
     "~1,750",
     "Staffing partner: Remedy Staffing (903) 509-8367. Major safety prospect"),

    ("John Soules Foods", "Food Manufacturing (Meat Processing)",
     "Jorge Torres", "Human Resources Director",
     "", "",
     "10150 FM 14, Tyler, TX 75706",
     "https://www.soulesfoods.com",
     "~1,324",
     "Founded 1975. Contact form at soulesfoods.com/contact-us. HR team of 7"),

    ("Wayne-Sanderson Farms", "Poultry Processing",
     "Brent Glasgow", "Plant Manager",
     "(903) 630-9343", "",
     "13523 FM 2015, Tyler, TX 75708",
     "https://waynesandersonfarms.com",
     "500+",
     "Mon-Fri 7am-5pm. High safety/wellness need"),

    ("Delek US (Tyler Refinery)", "Oil & Gas Refining",
     "", "HR Department",
     "(903) 474-3743", "",
     "425 McMurrey Dr, Tyler, TX 75702",
     "https://delekus.com",
     "",
     "Refinery operation. Corporate HQ: Brentwood, TN (615) 771-6701"),

    # ── RETAIL / DISTRIBUTION ────────────────────────────────────
    ("RETAIL & DISTRIBUTION", None),

    ("Brookshire Grocery Company (BGC)", "Retail Grocery / Distribution",
     "", "Corporate HR",
     "(903) 534-3000", "",
     "1600 W Southwest Loop 323, Tyler, TX 75701",
     "https://www.brookshires.com",
     "17,000+",
     "Family-owned since 1928. 215+ stores (Brookshire's, Super 1, Spring Market, FRESH, Reasor's). Fax: (903) 534-2206"),

    ("Cavender's (Corporate Office)", "Retail (Western Wear)",
     "", "Corporate HR",
     "(903) 509-9509", "",
     "2019 WSW Loop 323, Tyler, TX 75703",
     "https://www.cavenders.com",
     "~1,500",
     "Customer Service: (888) 361-2555. Owner: James Cavender"),

    ("Target Distribution Center (T-578)", "Retail Distribution",
     "", "HR Department",
     "(903) 881-1000", "candidate.accommodations@HRHelp.Target.com",
     "13786 Harvey Rd, Tyler, TX 75706",
     "",
     "500-999",
     "Fax: (903) 881-1059"),

    ("Amazon (Tyler Delivery Station)", "E-commerce / Logistics",
     "", "HR / ERC",
     "(888) 892-7180", "",
     "North Tyler Commerce Park, off Hwy 155 / CR 334",
     "https://www.amazon.jobs",
     "100+ FT, several hundred PT",
     "Opened Oct 2025. ~$50M investment. 140,000 sq ft"),

    # ── GOVERNMENT ───────────────────────────────────────────────
    ("GOVERNMENT", None),

    ("Smith County Government", "County Government",
     "Esmeralda Corona", "Chief HR Officer",
     "(903) 590-4645", "",
     "200 E Ferguson St, Suite 202, Tyler, TX 75707",
     "https://www.smith-county.com/209/Human-Resources",
     "",
     "Main HR: (903) 590-4644. Fax: (903) 590-4640. Mon-Fri 8am-5pm"),

    ("City of Tyler", "Municipal Government",
     "", "Organizational Development / HR",
     "", "",
     "212 N Bonner Ave, Tyler, TX 75702 (PO Box 2039, 75710)",
     "https://www.cityoftyler.org/government/departments/organizational-development-hr",
     "",
     "Staff directory: cityoftyler.org/staff-directory-list"),

    # ── BANKING ──────────────────────────────────────────────────
    ("BANKING & FINANCE", None),

    ("Citizens 1st Bank", "Banking",
     "James I. Perkins", "CEO",
     "(903) 581-1900", "info@citizens1stbank.com",
     "2001 ESE Loop 323, Tyler, TX 75701",
     "https://www.citizens1stbank.com",
     "51-200",
     ""),

    ("Southside Bank", "Banking",
     "", "Customer Care",
     "(903) 531-7111", "customercare@southside.com",
     "1201 S Beckham Ave, Tyler, TX 75701",
     "https://www.southside.com",
     "501-1,000",
     "Toll-free: (877) 639-3511"),

    # ── MEDIA ────────────────────────────────────────────────────
    ("MEDIA", None),

    ("Tyler Morning Telegraph", "Media / Newspaper",
     "Justin Wilcox", "Regional Publisher",
     "(903) 596-6299", "feedback@tylerpaper.com",
     "100 E Ferguson, Suite 501, Tyler, TX 75702",
     "https://tylerpaper.com",
     "",
     "Main: (903) 597-8111. Carpenter Media Group"),

    # ── RELIGIOUS (LARGE EMPLOYER) ───────────────────────────────
    ("LARGE RELIGIOUS EMPLOYERS", None),

    ("Green Acres Baptist Church", "Religious Organization",
     "", "HR Department",
     "(903) 525-1100", "",
     "1607 Troup Hwy, Tyler, TX 75701",
     "https://www.gabc.org",
     "51-200",
     "Staff directory: gabc.org/our-team. Mon-Fri 9am-5pm"),

    # ── STAFFING AGENCIES ────────────────────────────────────────
    ("STAFFING AGENCIES (Partnership Contacts)", None),

    ("Express Employment Professionals", "Staffing",
     "Rocky Gill", "Contact",
     "(903) 592-9999", "rocky.gill@expresspros.com",
     "5604 Donnybrook Ave, Tyler, TX 75703",
     "https://www.expresspros.com/us-texas-tyler",
     "",
     "Serving East Texas since 1995. 2025 Best of Staffing winner. By appointment only. Fax: (903) 592-9898"),

    ("Remedy Intelligent Staffing", "Staffing",
     "", "Office Contact",
     "(903) 509-8367", "ftylertx@select.com",
     "909 ESE Loop 323, Suite 550, Tyler, TX 75701",
     "https://www.remedystaffing.com/locations/tyler-tx-ris",
     "",
     "Handles staffing for Trane Tyler. Since 1998. Fax: (903) 509-8366"),
]

ws = wb.active
ws.title = "Employers & HR Contacts"

col_widths = [42, 30, 22, 26, 18, 36, 52, 50, 14, 60]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

row_num = 2
for record in employers:
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=record[0])
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 28
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

output = "/home/user/Claude/Smith_County_Employers_HR_Contacts.xlsx"
wb.save(output)
data_rows = sum(1 for r in employers if r[1] is not None)
print(f"Saved {data_rows} contacts to {output}")
