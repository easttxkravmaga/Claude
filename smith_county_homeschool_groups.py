#!/usr/bin/env python3
"""Generate Smith County TX & Surrounding Area Homeschool Groups Contacts spreadsheet (email required)."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=13, color="1A5276")
section_fill = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Organization", "Type", "Name", "Title", "Phone", "Email", "Address", "Website"]

# ── Data — ONLY entries with a confirmed email ──────────────────────────
# (Organization, Type, Name, Title, Phone, Email, Address, Website)
# None in Type = section header row

groups = [
    # ═══════════════════════════════════════════════════════════
    ("TYLER / SMITH COUNTY HOMESCHOOL GROUPS", None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("TACHE - Tyler Area Christian Home Educators", "Co-op / Support Group",
     "Jeremiah Ballard", "Principal Officer",
     "", "tache.info@tachetexas.org",
     "PO Box 132144, Tyler, TX 75713",
     "https://www.tachetexas.org"),

    ("Venture Enrichment (Calvary Baptist Church)", "Co-op / Enrichment",
     "Jenny", "Contact",
     "", "jenny@venturetyler.com",
     "Tyler, TX (Ministry of Calvary Baptist Church)",
     "https://www.venturetyler.com"),

    ("Venture Enrichment (Calvary Baptist Church)", "Co-op / Enrichment",
     "Kim Strout", "Admissions",
     "", "kim@venturetyler.com",
     "Tyler, TX (Ministry of Calvary Baptist Church)",
     "https://www.venturetyler.com"),

    ("Cottage Garden Co-op", "Co-op",
     "", "Co-op Office",
     "", "info@cottagegardentyler.org",
     "5601 E Amherst Dr, Tyler, TX 75701",
     "https://www.cottagegardentyler.org"),

    # ═══════════════════════════════════════════════════════════
    ("LINDALE AREA HOMESCHOOL GROUPS", None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("Arrow's Way Academy", "Co-op",
     "", "Co-op Contact",
     "", "ArrowsWayLindale@gmail.com",
     "Lindale, TX 75771",
     "https://arrowswayacademy.webador.com"),

    # ═══════════════════════════════════════════════════════════
    ("LONGVIEW AREA HOMESCHOOL GROUPS", None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("LARHE - Longview Area Relaxed Home Educators", "Support Group",
     "", "Group Contact",
     "", "larhehomeschool@gmail.com",
     "Longview, TX",
     "https://www.larhe.com"),

    ("CHEC - Christian Home Educators Community of Longview", "Co-op / Support Group",
     "", "General Contact",
     "", "checlongview@gmail.com",
     "P.O. Box 224, Longview, TX 75606",
     "https://www.checlongview.com"),

    # ═══════════════════════════════════════════════════════════
    ("HOMESCHOOL ATHLETICS", None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("EXEL Athletic Association", "Athletics",
     "", "Athletic Association",
     "", "ExelAthletics@gmail.com",
     "Canton, TX (serves Van Zandt, Wood, Smith, Gregg, Cherokee counties)",
     "https://exelathletics.weebly.com"),

    # ═══════════════════════════════════════════════════════════
    ("HOMESCHOOL ENRICHMENT PROGRAMS", None, None, None, None, None, None, None),
    # ═══════════════════════════════════════════════════════════

    ("East Texas Arboretum - Homeschool Classes", "Enrichment Program",
     "", "Program Contact",
     "(903) 675-5630", "eta@easttexasarboretum.org",
     "1601 Patterson Rd., Athens, TX 75751",
     "https://www.easttexasarboretum.org/homeschool.html"),
]


# ── Build worksheet ─────────────────────────────────────────────────────
ws = wb.active
ws.title = "Homeschool Groups"

# Column widths
col_widths = [48, 24, 22, 22, 18, 34, 52, 44]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Write headers
for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

# Write data
row_num = 2
for record in groups:
    org = record[0]
    # Section header rows
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=org)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 28
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

# ── Save ────────────────────────────────────────────────────────────────
output_path = "/home/user/Claude/Smith_County_TX_Homeschool_Groups.xlsx"
wb.save(output_path)
print(f"Saved {row_num - 2} rows to {output_path}")
