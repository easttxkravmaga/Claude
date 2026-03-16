#!/usr/bin/env python3
"""Generate Smith County Youth Sports Organizations spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=13, color="1B4332")
section_fill = PatternFill(start_color="B7E4C7", end_color="B7E4C7", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Organization", "Sport(s)", "Contact Name", "Title", "Phone", "Email", "Address", "Website", "Notes"]

orgs = [
    # ── FOOTBALL ─────────────────────────────────────────────────
    ("FOOTBALL", None),

    ("Tyler Lions Youth Sports Association", "Football, Baseball, Softball",
     "", "General Contact",
     "", "tylerlionsyouthsports@gmail.com",
     "Tyler, TX",
     "https://www.leaguelineup.com/welcome.asp?url=tylerlions",
     "Ages 5-12. Founded 2001"),

    ("Chapel Hill Youth Football & Cheer Association", "Football, Cheerleading",
     "", "General Contact",
     "(469) 441-9047", "chapelhillyouthfootball@gmail.com",
     "PO Box 131287, Tyler, TX",
     "https://chapelhillyouthsports.com",
     ""),

    # ── BASEBALL & SOFTBALL ──────────────────────────────────────
    ("BASEBALL & SOFTBALL", None),

    ("Rose Capital East Little League (RCELL)", "Baseball, Softball",
     "", "League Contact",
     "(903) 392-9479", "rosecapitaleast@gmail.com",
     "2300 McDonald Rd, Tyler, TX 75701 (Golden Road Park)",
     "https://www.rosecapitaleast.com",
     "Ages 4-14. ~600 players/season"),

    ("Rose Capital West Little League (RCWLL)", "Baseball, Softball",
     "", "League Contact",
     "(903) 920-9101", "rcwllbaseball@gmail.com",
     "410 W Cumberland Rd, Tyler, TX (Faulkner Park). Mail: PO Box 7241, Tyler, TX 75711",
     "https://www.rcwll.org",
     ""),

    ("North Tyler Youth Baseball & Softball Association", "Baseball, Softball",
     "", "League Contact",
     "", "northtyleryouthbaseball@gmail.com",
     "Fun Forest Park, Tyler, TX",
     "https://www.leaguelineup.com/welcome.asp?url=northtyleryouthbaseball",
     "Ages 5-12. Founded 2022. Diamond Youth Baseball/Softball member"),

    ("Chapel Hill Baseball/Softball Association (CHBSA)", "Baseball, Softball",
     "", "League Contact",
     "(903) 805-7699", "Chbbayouth@gmail.com",
     "15261 County Road 220, Tyler, TX 75707",
     "http://www.chbba.com",
     "Ages 3-15"),

    # ── SOCCER ───────────────────────────────────────────────────
    ("SOCCER", None),

    ("Tyler Soccer Association (TSA)", "Soccer",
     "Beth Davis", "Contact",
     "(903) 939-9829", "bethtylersoccer@gmail.com",
     "5610 Old Bullard Road, Suite 205, Tyler, TX 75703",
     "https://www.tylersoccer.com",
     "Rec and academy leagues. Games at Lindsey Park"),

    ("Soccer Shots East Texas", "Soccer (ages 18mo-8yr)",
     "", "Program Contact",
     "(318) 215-6942", "etx@soccershots.com",
     "22230 Orchard Dale Dr, Tyler, TX 75703",
     "https://www.soccershots.com/etx",
     ""),

    # ── BASKETBALL ───────────────────────────────────────────────
    ("BASKETBALL", None),

    ("East Texas Hoops", "Basketball (tournaments/leagues)",
     "Bart Millsap", "Founder",
     "(903) 434-3412", "",
     "Mt. Vernon, TX (hosts events in Tyler)",
     "http://www.easttexashoops.com",
     "Tournament of Champions in Tyler"),

    # ── MULTI-SPORT / RECREATION ─────────────────────────────────
    ("MULTI-SPORT & RECREATION", None),

    ("Boys & Girls Club of East Texas", "General youth programs",
     "", "General Contact",
     "(903) 593-9211", "",
     "504 West 32nd Street, Tyler, TX 75702",
     "http://www.bgcet.org",
     "Satellite at Owens Elementary: 11780 CR 168, Tyler (903) 262-2175"),

    ("i9 Sports — Greater Tyler", "Flag Football, Basketball, Soccer, Baseball",
     "", "General Contact",
     "(903) 484-4920", "wecare@i9sports.com",
     "1400 W. Barrett St. #2102, Tyler, TX 75702",
     "https://www.i9sports.com/franchises/greater-tyler-tx/471",
     "Serves Flint, Whitehouse, Bullard. One-day/week commitment"),

    ("City of Tyler — Athletics Division", "Various youth & adult leagues",
     "Allan Piedra", "Athletic Coordinator",
     "(903) 595-7217", "apiedra@TylerTexas.com",
     "1718 W Houston St, Tyler, TX 75702",
     "https://www.cityoftyler.org/government/departments/parks-rec/programs-services/athletics",
     "Parks Admin: (903) 531-1370"),

    ("GABC IMPACT Sports (Green Acres Baptist)", "Volleyball, Flag Football, Basketball, Soccer, Disc Golf, Pickleball",
     "", "IMPACT Sports",
     "(903) 525-3222", "",
     "18700 US-69, Tyler, TX 75703",
     "https://impactsports.win",
     "Church: 1607 Troup Hwy. Family Life Center: (903) 525-1116"),

    ("Colonial Hills Baptist — Upward Sports", "Basketball, Cheerleading, Soccer",
     "Jeremy Deimund", "Upward Director",
     "(903) 561-9995", "jeremy@colonialhills.com",
     "7330 South Broadway Ave, Tyler, TX 75703",
     "https://www.colonialhills.com/upward-sports",
     "25+ years. Soccer Aug-Oct; Basketball Jan-Feb"),

    # ── SWIMMING / AQUATICS ──────────────────────────────────────
    ("SWIMMING & AQUATICS", None),

    ("East Texas Aquatics / Tyler Swim Academy", "Competitive swimming, lessons",
     "", "General Contact",
     "(903) 566-6595", "",
     "11594 TX-248 Spur, Tyler, TX 75707",
     "https://www.gomotionapp.com/team/ntstst/page/home",
     "Since 1972. 10-lane, 25-yard pool. Tyler Titans summer team"),

    ("Tyler Rose Aquatic Club", "Competitive swimming, lessons",
     "Abbi Roeland", "Coach",
     "(903) 262-1897", "abbi.roeland@tylerisd.org",
     "3013 Earl Campbell Pkwy, Tyler, TX 75701",
     "http://www.tylerroseaquatics.com",
     "Also: Jason Petty (jason.petty@tylerisd.org). Office: (903) 262-1024"),

    ("Tyler Swim School", "Swim lessons (kids & adults)",
     "", "General Contact",
     "(903) 595-1222", "",
     "Tyler, TX 75702",
     "https://www.tylerswimschool.com",
     "Max 3 students/class. Indoor heated salt water pool"),

    # ── VOLLEYBALL ───────────────────────────────────────────────
    ("VOLLEYBALL", None),

    ("Tyler Juniors Volleyball", "Club volleyball (ages 13-17)",
     "Cristy O'Bannon", "Director",
     "", "",
     "413 W Cumberland Rd #302, Tyler, TX 75703",
     "https://www.tylerjuniorsvb.com",
     "USA Volleyball, North Texas Region"),

    ("Tyler HEAT Volleyball", "Volleyball (grades 6-12, homeschool)",
     "", "Program Contact",
     "", "",
     "315 N. Broadway, Ste 210, Tyler, TX 75702",
     "https://etxheat.org/volleyball",
     "$275 middle school, $375 high school"),

    # ── GYMNASTICS ───────────────────────────────────────────────
    ("GYMNASTICS", None),

    ("Texas East Kids (formerly Texas East Gymnastics)", "Gymnastics, competitive, cheer tumbling",
     "Martin Parsley", "Owner",
     "(903) 509-3547", "theoffice@texaseastkids.com",
     "1914 Deerbrook Dr, Tyler, TX 75703",
     "https://texaseastkids.com",
     "22,500 sq ft. 50+ years. Also: Stacy Parsley, Marlynne Parsley Finch"),

    ("GymTyler Gymnastics & Fitness", "Gymnastics, tumbling, cheer",
     "Kim Johnson", "Owner",
     "(903) 593-2931", "gymtylergymnasticsandfitness@gmail.com",
     "4598 Old Troup Hwy, Tyler, TX",
     "https://www.gymtyler.com",
     "Since 1970. 3 locations"),

    ("The Little Gym of Tyler", "Gymnastics (ages 4mo - grade 6)",
     "", "General Contact",
     "(903) 871-1152", "",
     "7596 Old Jacksonville Hwy, Ste 100, Tyler, TX 75703",
     "https://www.thelittlegym.com/texas-tyler",
     ""),

    # ── CHEERLEADING & DANCE ─────────────────────────────────────
    ("CHEERLEADING & DANCE", None),

    ("Spirit of Tyler Cheerleading and Dance", "Competitive cheer, tumbling, hip hop, dance",
     "", "General Contact",
     "(903) 521-8183", "spiritoftyler@hotmail.com",
     "10595 FM 2813, Tyler, TX",
     "https://spiritoftyler.com",
     "28 seasons. 132 national championships, 32 grand championship titles"),

    ("Tyler Rose Athletics", "All-star cheer, tumbling",
     "Neal Mark", "Owner",
     "(903) 484-6610", "info@tracheer.com",
     "15294 Hwy 110, Whitehouse, TX 75791",
     "https://www.tracheer.com",
     ""),

    # ── MARTIAL ARTS ─────────────────────────────────────────────
    ("MARTIAL ARTS (Youth Programs)", None),

    ("Tiger-Rock Martial Arts of Tyler", "Taekwondo, self-defense (ages 4+)",
     "Amy & Charles Lauffer", "Owners",
     "(903) 509-8782", "info@trmatyler.com",
     "535 W Southwest Loop 323, #300, Tyler, TX 75701",
     "https://trmatyler.com",
     ""),

    ("Bares Taekwondo Fitness", "Taekwondo, kickboxing, Krav Maga",
     "Matt Bares", "Founder",
     "(903) 561-2966", "",
     "1901 Deerbrook Dr, Tyler, TX 75703",
     "https://www.barestkd.fit",
     "Est. 2006"),

    ("Inspire Youth Sports", "Martial arts, dance (ages 3+)",
     "Anthony Splawn", "Owner",
     "(903) 239-2480", "",
     "1505 E Grande Blvd, Tyler, TX 75703",
     "https://inspiremartialartstyler.com",
     ""),

    # ── TENNIS ───────────────────────────────────────────────────
    ("TENNIS", None),

    ("Tyler Community Tennis Association (TCTA)", "Tennis (all ages)",
     "", "General Contact",
     "", "",
     "Tyler, TX 75701",
     "https://playtennis.usta.com/TylerCTA",
     "501(c)(3). USTA affiliated"),

    # ── HOMESCHOOL ATHLETICS ─────────────────────────────────────
    ("HOMESCHOOL ATHLETICS", None),

    ("Tyler HEAT (Home Education Athletic Teams)", "Football, Volleyball, Cross Country, Soccer, Basketball, Baseball",
     "", "General Contact",
     "", "",
     "315 N. Broadway, Ste 210, Tyler, TX 75702",
     "https://etxheat.org",
     "Grades 6-12, Christian homeschool students"),
]

ws = wb.active
ws.title = "Youth Sports Organizations"

col_widths = [46, 36, 24, 20, 18, 34, 52, 50, 56]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[1].height = 22

row_num = 2
for record in orgs:
    if record[1] is None:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(HEADERS))
        cell = ws.cell(row=row_num, column=1, value=record[0])
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 28
        row_num += 1
        continue

    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)
    row_num += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_num - 1}"

output = "/home/user/Claude/Smith_County_Youth_Sports_Organizations.xlsx"
wb.save(output)
data_rows = sum(1 for r in orgs if r[1] is not None)
print(f"Saved {data_rows} contacts to {output}")
